import streamlit as st
import pandas as pd
import math
import plotly.graph_objects as go
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import PatternFill
import base64
import pytz
# ==============================================================================
# 1. CONFIGURAÇÕES E CONSTANTES GLOBAIS APRIMORADAS
# ==============================================================================

st.set_page_config(
    page_title="🏗️ Calculadora Estrutural - Perfis Metálicos",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.abnt.org.br',
        'Report a bug': None,
        'About': "# Calculadora Estrutural\nCálculos baseados na ABNT NBR 8800:2008"
    }
)

class Config:
    NOME_NORMA = 'ABNT NBR 8800:2008'
    GAMMA_A1 = 1.10
    FATOR_SIGMA_R = 0.3
    FATOR_LAMBDA_P_FLT = 1.76
    FATOR_LAMBDA_P_FLM = 0.38
    FATOR_LAMBDA_R_FLM_LAMINADO = 0.83
    FATOR_LAMBDA_R_FLM_SOLDADO = 0.95
    FATOR_LAMBDA_P_FLA = 3.76
    FATOR_LAMBDA_R_FLA = 5.70
    KV_ALMA_SEM_ENRIJECEDORES = 5.0
    FATOR_VP = 0.60
    FATOR_LAMBDA_P_VRD = 1.10
    FATOR_LAMBDA_R_VRD = 1.37
    FATOR_VRD_ELASTICO = 1.24

PROFILE_TYPE_MAP = {
    "Laminados": "Perfis Laminados",
    "CS": "Perfis Compactos Soldados",
    "CVS": "Perfil de Seção Variável",
    "VS": "Perfis Soldados"
}


# Substitua a variável HTML_TEMPLATE_CSS_PRO inteira por esta:
HTML_TEMPLATE_CSS_PRO = """
<style>
    /* Google Fonts: Inter & Poppins & JetBrains Mono */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Poppins:wght@600;700;800&display=swap');
    @import url('https://fonts.com/css2?family=JetBrains+Mono:wght@400;500&display=swap');

    :root {
        --background-dark: #0A0A0F; /* Mais escuro */
        --surface-medium: #15151F; /* Fundo principal */
        --surface-light: #20202A;  /* Cards e elementos */
        --border-color: #30303A;   /* Linhas de borda */
        --text-primary: #F0F0F0;   /* Texto claro */
        --text-secondary: #AAAAAA; /* Texto secundário */
        --accent-gold: #FFD700;    /* Ouro puro */
        --accent-amber: #FFBF00;   /* Âmbar mais vibrante */
        --button-primary: #FFBF00; /* Botão principal */
        --button-hover: #FFD700;   /* Hover do botão */
        --button-text: #0A0A0F;    /* Texto escuro no botão */
        --success: #32CD32;        /* Verde para sucesso */
        --error: #FF4500;          /* Laranja avermelhado para erro */
        --shadow-color: rgba(0, 0, 0, 0.4); /* Sombra mais escura */
    }

    /* --- Base Body & Streamlit Overrides --- */
    html, body {
        font-family: 'Inter', sans-serif;
        background-color: var(--background-dark);
        color: var(--text-primary);
        line-height: 1.6;
    }

    .stApp {
        background-color: var(--background-dark);
        color: var(--text-primary);
    }

    /* Remove Streamlit default padding and margin */
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
        padding-left: 3rem;
        padding-right: 3rem;
    }

    /* Streamlit Sidebar */
    .css-1d391kg, .css-1dp5vir { /* Targets sidebar content and overall sidebar */
        background-color: var(--surface-medium);
        color: var(--text-primary);
    }
    .css-1dp5vir .css-1v3fvcr { /* Sidebar header */
        color: var(--accent-gold);
        font-family: 'Poppins', sans-serif;
        font-weight: 700;
        border-bottom: 1px solid var(--border-color);
        padding-bottom: 1rem;
        margin-bottom: 1rem;
    }
    .css-1dp5vir .stSelectbox > label, 
    .css-1dp5vir .stNumberInput > label,
    .css-1dp5vir .stDateInput > label,
    .css-1dp5vir .stTextInput > label,
    .css-1dp5vir .stRadio > label,
    .css-1dp5vir .stCheckbox > label {
        color: var(--text-primary);
        font-weight: 600;
    }
    .css-1dp5vir .stSelectbox div[data-baseweb="select"] button,
    .css-1dp5vir .stNumberInput input,
    .css-1dp5vir .stTextInput input,
    .css-1dp5vir .stDateInput input {
        background-color: var(--surface-light);
        color: var(--text-primary);
        border: 1px solid var(--border-color);
        border-radius: 6px;
    }
    .css-1dp5vir .stSelectbox div[data-baseweb="select"] button:hover,
    .css-1dp5vir .stNumberInput input:hover,
    .css-1dp5vir .stTextInput input:hover,
    .css-1dp5vir .stDateInput input:hover {
        border-color: var(--accent-gold);
    }
    .css-1dp5vir .stSelectbox [data-baseweb="menu"] {
        background-color: var(--surface-light);
        color: var(--text-primary);
        border: 1px solid var(--border-color);
    }
    .css-1dp5vir .stSelectbox [data-baseweb="menu"] li {
        color: var(--text-primary);
    }
    .css-1dp5vir .stSelectbox [data-baseweb="menu"] li:hover {
        background-color: var(--background-dark);
        color: var(--accent-gold);
    }
    
    /* --- Header Profissional com Logo --- */
    .pro-header {
        background: var(--surface-light);
        color: var(--text-primary);
        padding: 3rem 2rem; /* Mais padding */
        border-radius: 12px;
        border: 1px solid var(--border-color);
        text-align: center;
        margin-bottom: 2.5rem;
        box-shadow: 0 8px 25px var(--shadow-color); /* Sombra proeminente */
        position: relative;
        overflow: hidden; /* Para o efeito de borda */
    }

    /* Efeito de borda dourada para o cabeçalho */
    .pro-header::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0; bottom: 0;
        border-radius: 12px;
        padding: 2px; /* Espessura da borda */
        background: linear-gradient(45deg, var(--accent-amber), var(--accent-gold), var(--accent-amber));
        -webkit-mask: 
            linear-gradient(#fff 0 0) content-box, 
            linear-gradient(#fff 0 0);
        -webkit-mask-composite: xor;
        mask-composite: exclude;
    }

    .header-content {
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 1.5rem; /* Mais espaçamento */
        position: relative; /* Para ficar acima do ::before */
        z-index: 1;
    }

    .pro-header img {
        height: 100px; /* Logo maior */
        width: auto;
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.3); /* Sombra na logo */
    }

    .pro-header h1 {
        font-size: 3rem; /* Título maior */
        font-weight: 800;
        font-family: 'Poppins', sans-serif;
        margin: 0;
        line-height: 1.2;
    }
    
    .pro-header p {
        font-size: 1.2rem; /* Subtítulo maior */
        margin: 0;
        color: var(--text-secondary);
    }
    
    .gradient-text {
        background: linear-gradient(135deg, var(--accent-amber) 0%, var(--accent-gold) 50%, #FFD700 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }

    /* --- Títulos --- */
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Poppins', sans-serif;
        color: var(--text-primary);
        font-weight: 700;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }

    h2 {
        border-bottom: 2px solid var(--border-color);
        padding-bottom: 15px;
        color: var(--accent-gold); /* H2 em dourado */
        font-size: 1.8rem;
    }
    
    h3 { 
        color: var(--accent-amber); 
        font-size: 1.5rem;
        margin-top: 1.5rem;
    }

    /* --- Seções de Parâmetros e Modos --- */
    .stColumns > div > div {
        background: var(--surface-light);
        border: 1px solid var(--border-color);
        border-radius: 10px;
        padding: 1.5rem;
        box-shadow: 0 4px 15px var(--shadow-color); /* Sombra nos cards */
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        gap: 0.5rem;
    }
    .stColumns > div > div h3 {
        margin-top: 0;
        font-size: 1.3rem;
        display: flex;
        align-items: center;
        gap: 0.75rem;
        color: var(--text-primary); /* Título do card */
    }
    .stColumns > div > div h3 .icon {
        color: var(--accent-gold); /* Ícone dourado */
        font-size: 1.5rem;
    }
    .stColumns > div > div p {
        margin: 0;
        font-size: 1.1rem;
        color: var(--text-secondary);
    }
    .stColumns > div > div p strong {
        color: var(--text-primary);
        font-weight: 600;
    }

    /* --- Icones --- */
    .icon {
        font-size: 1.2rem;
        margin-right: 0.5rem;
        color: var(--accent-gold); /* Ícones com a cor dourada */
    }
    
    /* --- Botões --- */
    .stButton button {
        background-color: var(--button-primary);
        color: var(--button-text);
        font-weight: bold;
        padding: 0.75rem 1.5rem;
        border-radius: 8px;
        border: none;
        cursor: pointer;
        transition: all 0.3s ease;
        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.2);
    }
    .stButton button:hover {
        background-color: var(--button-hover);
        box-shadow: 0 6px 15px rgba(0, 0, 0, 0.3);
        transform: translateY(-2px);
    }
    .stButton button:active {
        transform: translateY(0);
        box-shadow: 0 2px 5px rgba(0, 0, 0, 0.2);
    }
    .stButton button.secondary-button { /* Para botões secundários */
        background-color: var(--surface-light);
        color: var(--accent-gold);
        border: 1px solid var(--accent-gold);
    }
    .stButton button.secondary-button:hover {
        background-color: var(--accent-gold);
        color: var(--button-text);
    }
    
    /* --- Entradas de Texto/Número e Seletores --- */
    .stNumberInput label, .stSelectbox label, .stTextInput label, .stDateInput label {
        color: var(--text-primary);
        font-weight: 600;
        font-size: 1rem;
        margin-bottom: 0.5rem;
    }
    .stNumberInput input, .stTextInput input {
        background-color: var(--surface-medium);
        color: var(--text-primary);
        border: 1px solid var(--border-color);
        border-radius: 6px;
        padding: 0.75rem;
        font-size: 1rem;
        transition: border-color 0.2s ease;
    }
    .stNumberInput input:focus, .stTextInput input:focus {
        border-color: var(--accent-gold);
        box-shadow: 0 0 0 2px rgba(255, 191, 0, 0.3); /* Dourado suave */
        outline: none;
    }
    .stSelectbox div[data-baseweb="select"] button {
        background-color: var(--surface-medium);
        color: var(--text-primary);
        border: 1px solid var(--border-color);
        border-radius: 6px;
        padding: 0.75rem;
        font-size: 1rem;
        transition: border-color 0.2s ease;
    }
    .stSelectbox div[data-baseweb="select"] button:hover {
        border-color: var(--accent-gold);
    }
    .stSelectbox [data-baseweb="menu"] {
        background-color: var(--surface-medium);
        color: var(--text-primary);
        border: 1px solid var(--border-color);
        border-radius: 6px;
    }
    .stSelectbox [data-baseweb="menu"] li {
        color: var(--text-primary);
        padding: 0.75rem 1rem;
    }
    .stSelectbox [data-baseweb="menu"] li:hover {
        background-color: var(--surface-light);
        color: var(--accent-gold);
    }


    /* --- Tabelas de Resumo (se houver) --- */
    .summary-table {
        border-collapse: separate;
        border-spacing: 0;
        border: 1px solid var(--border-color);
        border-radius: 8px;
        overflow: hidden;
        margin-top: 1.5rem;
        box-shadow: 0 4px 15px var(--shadow-color);
    }

    .summary-table thead tr th {
        background: var(--surface-light);
        color: var(--accent-gold);
        padding: 1rem;
        font-weight: 600;
        font-size: 0.95rem;
        text-transform: uppercase;
        border-bottom: 2px solid var(--border-color);
        text-align: left;
    }

    .summary-table tbody tr td {
        padding: 1rem;
        background: var(--surface-medium);
        border-bottom: 1px solid var(--border-color);
        color: var(--text-primary);
    }
    .summary-table tbody tr:last-child td { border-bottom: none; }
    .summary-table tbody tr:nth-child(even) td { background: var(--surface-light); } /* Linhas alternadas */

    /* --- Blocos de Detalhes e Fórmulas --- */
    .info-card, .formula-block {
        background: var(--surface-light);
        border: 1px solid var(--border-color);
        padding: 1.5rem;
        margin: 1.5rem 0;
        border-radius: 10px;
        box-shadow: 0 4px 15px var(--shadow-color);
    }
    .formula-block {
        background-color: var(--background-dark);
        border-left: 5px solid var(--accent-amber); /* Borda mais grossa */
    }
    
    .formula {
        font-family: 'JetBrains Mono', monospace;
        background: var(--surface-medium);
        padding: 1rem;
        border-radius: 6px;
        border: 1px solid var(--border-color);
        margin: 1rem 0;
        color: var(--text-primary);
        font-size: 0.95rem;
    }
    
    .verification-step {
        background-color: var(--surface-medium);
        border: 1px solid var(--border-color);
        padding: 1.5rem;
        margin-top: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 10px var(--shadow-color);
    }

    /* --- Indicadores de Status --- */
    .pass { color: var(--success); font-weight: 700; }
    .fail { color: var(--error); font-weight: 700; }
    .conclusion.pass { color: var(--success); }
    .conclusion.fail { color: var(--error); }

    .final-status.pass {
        background-color: rgba(50, 205, 50, 0.15); /* Fundo verde translúcido */
        color: var(--success);
        border: 1px solid var(--success);
    }
    .final-status.fail {
        background-color: rgba(255, 69, 0, 0.15); /* Fundo vermelho translúcido */
        color: var(--error);
        border: 1px solid var(--error);
    }
    .final-status {
        font-size: 1.3em; font-weight: bold; text-align: center;
        padding: 1rem; border-radius: 8px; margin: 1.5rem 0;
        box-shadow: 0 4px 15px var(--shadow-color);
    }

    /* --- Rodapé (Opcional, se você tiver um) --- */
    .footer {
        text-align: center;
        margin-top: 3rem;
        padding-top: 1.5rem;
        border-top: 1px solid var(--border-color);
        color: var(--text-secondary);
        font-size: 0.9rem;
    }

    /* Remove padding default de colunas para melhor controle do card */
    .st-emotion-cache-1uj3h0e { /* Target column content */
        padding: 0px !important;
    }
    
    /* Small adjustments to column gaps if needed */
    .st-emotion-cache-1sdc03s { /* Parent of columns */
        gap: 20px; /* Adjust gap between cards */
    }

    /* Icones dentro dos titulos h3 nos cards */
    h3 > span:first-child {
        color: var(--accent-gold);
    }

</style>
"""

# ==============================================================================
# 2. FUNÇÕES DE CÁLCULO E UTILITÁRIAS
# ==============================================================================

@st.cache_data
def load_data_from_local_file():
    """Carrega os dados da planilha de perfis."""
    try:
        caminho_arquivo_excel = 'perfis.xlsx'
        return pd.read_excel(caminho_arquivo_excel, sheet_name=None)
    except FileNotFoundError:
        st.error(f"Erro: Arquivo '{caminho_arquivo_excel}' não foi encontrado. Verifique se ele está na mesma pasta que o seu script Python.")
        return None
    except Exception as e:
        st.error(f"Erro ao ler o arquivo Excel: {e}")
        return None

def calcular_esforcos_viga(tipo_viga, L_cm, q_kn_cm=0, p_load=None):
    msd_q, vsd_q, msd_p, vsd_p = 0, 0, 0, 0
    L = L_cm
    detalhes_esforcos = {
        'Msd_q': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN.cm'},
        'Vsd_q': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN'},
        'Msd_p': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN.cm'},
        'Vsd_p': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN'}
    }

    detalhes_carga_memorial = {'q_ult': {'valor': 0, 'unidade': 'kN/cm'}, 'p_ult': {'valor': 0, 'unidade': 'kN'}}

    if q_kn_cm > 0:
        detalhes_carga_memorial['q_ult']['valor'] = q_kn_cm
        if tipo_viga == 'Bi-apoiada':
            msd_q = (q_kn_cm * L**2) / 8
            vsd_q = (q_kn_cm * L) / 2
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{8}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = \\frac{q_{ultima} \\times L}{2}"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            msd_q = (q_kn_cm * L**2) / 2
            vsd_q = q_kn_cm * L
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{2}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = q_{ultima} \\times L"
        elif tipo_viga == 'Bi-engastada':
            msd_q = (q_kn_cm * L**2) / 12
            vsd_q = (q_kn_cm * L) / 2
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{12}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = \\frac{q_{ultima} \\times L}{2}"
        elif tipo_viga == 'Engastada e Apoiada':
            msd_q = (q_kn_cm * L**2) / 8
            vsd_q = (5 * q_kn_cm * L) / 8
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{8}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = \\frac{5 \\times q_{ultima} \\times L}{8}"

    if p_load:
        P, x = p_load
        detalhes_carga_memorial['p_ult']['valor'] = P
        P_kn = P
        L_m = L / 100.0
        x_m = x / 100.0
        a_val = x_m
        b_val = L_m - a_val
        if tipo_viga == 'Bi-apoiada':
            msd_p = (P_kn * a_val * b_val) / L_m
            vsd_p = max((P_kn * b_val) / L_m, (P_kn * a_val) / L_m)
            detalhes_esforcos['Msd_p']['formula_simbolica'] = "M_{sd, P} = \\frac{P_{ultima} \\times a \\times b}{L}"
            detalhes_esforcos['Vsd_p']['formula_simbolica'] = "V_{sd, P} = \\max(\\frac{P_{ultima} \\times b}{L}, \\frac{P_{ultima} \\times a}{L})"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            msd_p = P_kn * a_val
            vsd_p = P_kn
            detalhes_esforcos['Msd_p']['formula_simbolica'] = "M_{sd, P} = P_{ultima} \\times a"
            detalhes_esforcos['Vsd_p']['formula_simbolica'] = "V_{sd, P} = P_{ultima}"
        msd_p *= 100

    detalhes_esforcos['Msd_q']['valor'] = msd_q
    detalhes_esforcos['Vsd_q']['valor'] = vsd_q
    detalhes_esforcos['Msd_p']['valor'] = msd_p
    detalhes_esforcos['Vsd_p']['valor'] = vsd_p

    msd_total = msd_q + msd_p
    vsd_total = vsd_q + vsd_p

    return msd_total, vsd_total, detalhes_esforcos, detalhes_carga_memorial

def calcular_cb(tipo_viga, L_cm, q_kn_cm=0, p_load=None):
    L_m = L_cm / 100
    q_ult_kn_m = q_kn_cm * 100

    detalhes_cb = {
        'formula_simbolica': 'C_b = \\frac{12,5 M_{máx}}{|2,5 M_{máx} + 3 M_A + 4 M_B + 3 M_C|}',
        'momentos': {
            'M_max': {'final_value': 0, 'components': []},
            'M_A': {'final_value': 0, 'components': []},
            'M_B': {'final_value': 0, 'components': []},
            'M_C': {'final_value': 0, 'components': []}
        }
    }

    momentos_q = {'M_max': 0, 'M_A': 0, 'M_B': 0, 'M_C': 0}
    momentos_p = {'M_max': 0, 'M_A': 0, 'M_B': 0, 'M_C': 0}

    # --- Cálculo devido à Carga Distribuída (q) ---
    if q_kn_cm > 0:
        if tipo_viga == 'Bi-apoiada':
            momentos_q['M_max'] = (q_ult_kn_m * L_m**2) / 8
            momentos_q['M_A'] = (3 * q_ult_kn_m * L_m**2) / 32
            momentos_q['M_B'] = momentos_q['M_max']
            momentos_q['M_C'] = momentos_q['M_A']

            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de q', 'formula': 'M_{max,q} = \\frac{q_{ultima} \\times L^2}{8}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de q', 'formula': 'M_{A,q} = \\frac{3 \\times q_{ultima} \\times L^2}{32}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de q', 'formula': 'M_{B,q} = M_{max,q}', 'valores': {}, 'valor': momentos_q['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de q', 'formula': 'M_{C,q} = M_{A,q}', 'valores': {}, 'valor': momentos_q['M_C']})

        elif tipo_viga == 'Engastada e Livre (Balanço)':
            momentos_q['M_max'] = (q_ult_kn_m * L_m**2) / 2
            momentos_q['M_A'] = (q_ult_kn_m * (0.75 * L_m)**2) / 2
            momentos_q['M_B'] = (q_ult_kn_m * (0.5 * L_m)**2) / 2
            momentos_q['M_C'] = (q_ult_kn_m * (0.25 * L_m)**2) / 2

            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de q', 'formula': 'M_{max,q} = \\frac{q_{ultima} \\times L^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de q', 'formula': 'M_{A,q} = \\frac{q_{ultima} \\times (0,75L)^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de q', 'formula': 'M_{B,q} = \\frac{q_{ultima} \\times (0,5L)^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de q', 'formula': 'M_{C,q} = \\frac{q_{ultima} \\times (0,25L)^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_C']})
        else:
              return 1.0, {'formula_simbolica': 'N/A', 'momentos': {}}

    # --- Cálculo devido à Carga Pontual (P) ---
    if p_load:
        P_kn, x_cm = p_load
        x_m = x_cm / 100.0
        
        if tipo_viga == 'Bi-apoiada':
            a, b = x_m, L_m - x_m
            momentos_p['M_max'] = (P_kn * a * b) / L_m if L_m > 0 else 0
            momentos_p['M_A'] = (P_kn * b * (0.25*L_m)) / L_m if x_m > 0.25*L_m else (P_kn * a * (L_m - 0.25*L_m)) / L_m
            momentos_p['M_B'] = (P_kn * b * (0.5*L_m)) / L_m if x_m > 0.5*L_m else (P_kn * a * (L_m - 0.5*L_m)) / L_m
            momentos_p['M_C'] = (P_kn * b * (0.75*L_m)) / L_m if x_m > 0.75*L_m else (P_kn * a * (L_m - 0.75*L_m)) / L_m
            
            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de P', 'formula': 'M_{max,P} = \\frac{P_{ultima} \\times a \\times b}{L}', 'valores': {'P_{ultima}': P_kn, 'a': a, 'b': b, 'L': L_m}, 'valor': momentos_p['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de P', 'formula': 'M_{A,P} \\text{ (calculado em L/4)}', 'valores': {}, 'valor': momentos_p['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de P', 'formula': 'M_{B,P} \\text{ (calculado em L/2)}', 'valores': {}, 'valor': momentos_p['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de P', 'formula': 'M_{C,P} \\text{ (calculado em 3L/4)}', 'valores': {}, 'valor': momentos_p['M_C']})

        elif tipo_viga == 'Engastada e Livre (Balanço)':
            momentos_p['M_max'] = P_kn * x_m
            momentos_p['M_A'] = P_kn * x_m if x_m < (0.25 * L_m) else P_kn * (x_m - 0.25 * L_m)
            momentos_p['M_B'] = P_kn * x_m if x_m < (0.5 * L_m) else P_kn * (x_m - 0.5 * L_m)
            momentos_p['M_C'] = P_kn * x_m if x_m < (0.75 * L_m) else P_kn * (x_m - 0.75 * L_m)
            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de P', 'formula': 'M_{max,P} = P_{ultima} \\times x', 'valores': {'P_{ultima}': P_kn, 'x': x_m}, 'valor': momentos_p['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de P', 'formula': 'M_{A,P} = P_{ultima} \\times (x-0.25L)', 'valores': {}, 'valor': momentos_p['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de P', 'formula': 'M_{B,P} = P_{ultima} \\times (x-0.5L)', 'valores': {}, 'valor': momentos_p['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de P', 'formula': 'M_{C,P} = P_{ultima} \\times (x-0.75L)', 'valores': {}, 'valor': momentos_p['M_C']})
        else:
              return 1.0, {'formula_simbolica': 'N/A', 'momentos': {}}

    # --- Soma das componentes e cálculo final do Cb ---
    M_finais = {}
    for m_key in ['M_max', 'M_A', 'M_B', 'M_C']:
        total = sum(c['valor'] for c in detalhes_cb['momentos'][m_key]['components'])
        detalhes_cb['momentos'][m_key]['final_value'] = abs(total)
        M_finais[m_key] = abs(total)

    denominador = (2.5 * M_finais['M_max'] + 3 * M_finais['M_A'] + 4 * M_finais['M_B'] + 3 * M_finais['M_C'])
    Cb = min(12.5 * M_finais['M_max'] / abs(denominador), 3.0) if denominador != 0 else 1.0

    return Cb, detalhes_cb

def calcular_flecha_maxima(tipo_viga, L_cm, E, Ix, q_serv_kn_cm=0, p_serv_load=None):
    delta_q, delta_p = 0, 0
    L = L_cm

    detalhes = {
        'delta_q': {'formula_simbolica': '', 'formula_numerica': '', 'valor': 0, 'unidade': 'cm'},
        'delta_p': {'formula_simbolica': '', 'formula_numerica': '', 'valor': 0, 'unidade': 'cm'},
        'delta_total': 0
    }

    if q_serv_kn_cm > 0:
        q_serv_val = q_serv_kn_cm
        if tipo_viga == 'Bi-apoiada':
            delta_q = (5 * q_serv_val * L**4) / (384 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{5 \\times q_{serv} \\times L^4}{384 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{5 \\times \\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{384 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            delta_q = (q_serv_val * L**4) / (8 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{q_{serv} \\times L^4}{8 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{\\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{8 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Bi-engastada':
            delta_q = (q_serv_val * L**4) / (384 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{q_{serv} \\times L^4}{384 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{\\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{384 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Engastada e Apoiada':
            delta_q = (q_serv_val * L**4) / (185 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{q_{serv} \\times L^4}{185 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{\\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{185 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        detalhes['delta_q']['valor'] = delta_q

    if p_serv_load:
        P, x = p_serv_load
        a = x
        b = L - a
        if tipo_viga == 'Bi-apoiada':
            delta_p = (P * a**2 * b**2) / (3 * E * Ix * L)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^2 \\times b^2}{3 \\times E \\times I_x \\times L}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^2 \\times \\mathbf{{{b:.2f}}}^2}}{{3 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}} \\times \\mathbf{{{L:.2f}}}}}"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            delta_p = (P * a**3) / (3 * E * Ix)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^3}{3 \\times E \\times I_x}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^3}}{{3 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Bi-engastada':
            delta_p = (P * a**3 * b**3) / (3 * E * Ix * L**3)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^3 \\times b^3}{3 \\times E \\times I_x \\times L^3}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^3 \\times \\mathbf{{{b:.2f}}}^3}}{{3 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}} \\times \\mathbf{{{L:.2f}}}^3}}"
        elif tipo_viga == 'Engastada e Apoiada':
            delta_p = (P * a**3 * b**2) / (12 * E * Ix * L**3) * (a + 2*L)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^3 \\times b^2 \\times (a + 2L)}{12 \\times E \\times I_x \\times L^3}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^3 \\times \\mathbf{{{b:.2f}}}^2 \\times (\\mathbf{{{a:.2f}}} + 2 \\times \\mathbf{{{L:.2f}}})}}{{12 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}} \\times \\mathbf{{{L:.2f}}}^3}}"

        detalhes['delta_p']['valor'] = delta_p

    detalhes['delta_total'] = delta_q + delta_p
    return detalhes

def get_profile_properties(profile_series):
    props = {"d": profile_series.get('d (mm)'),"bf": profile_series.get('bf (mm)'),"tw": profile_series.get('tw (mm)'),"tf": profile_series.get('tf (mm)'),"h": profile_series.get('h (mm)'),"Area": profile_series.get('Área (cm2)'),"Ix": profile_series.get('Ix (cm4)'),"Wx": profile_series.get('Wx (cm3)'),"rx": profile_series.get('rx (cm)'),"Zx": profile_series.get('Zx (cm3)'),"Iy": profile_series.get('Iy (cm4)'),"Wy": profile_series.get('Wy (cm3)'),"ry": profile_series.get('ry (cm)'),"Zy": profile_series.get('Zy (cm3)'),"J": profile_series.get('It (cm4)'),"Cw": profile_series.get('Cw (cm6)'),"Peso": profile_series.get('Massa Linear (kg/m)', profile_series.get('Peso (kg/m)'))}
    required_keys = ["d", "bf", "tw", "tf", "h", "Area", "Ix", "Wx", "rx", "Zx", "Iy", "ry", "J", "Cw", "Peso"]
    profile_name = profile_series.get('Bitola (mm x kg/m)', 'Perfil Desconhecido')
    for key in required_keys:
        value = props.get(key)
        if value is None or pd.isna(value) or (isinstance(value, (int, float)) and value <= 0):
            raise ValueError(f"Propriedade ESSENCIAL '{key}' inválida ou nula no Excel para '{profile_name}'. Verifique a planilha.")
    for key in ['d', 'bf', 'tw', 'tf', 'h']: props[key] /= 10.0
    return props

def _calcular_mrdx_flt(props, Lb, Cb, fy, E):
    Zx, ry, Iy, Cw, J, Wx = props['Zx'], props['ry'], props['Iy'], props['Cw'], props['J'], props['Wx']
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Mp = Zx * fy
    detalhes['passos_calculo'].append({
        'desc': 'Momento de Plastificação',
        'formula': 'M_p = Z_x \\times f_y',
        'valores': {'Z_x': Zx, 'f_y': fy},
        'valor': Mp,
        'unidade': 'kN.cm',
        'verif_id': 'Mp'
    })

    lambda_val = Lb / ry if ry > 0 else float('inf')
    detalhes['passos_calculo'].append({
        'desc': 'Índice de Esbeltez (λ = Lb/ry)',
        'formula': '\\lambda = \\frac{{L_b}}{{r_y}}',
        'valores': {'L_b': Lb, 'r_y': ry},
        'valor': lambda_val,
        'verif_id': 'lambda'
    })

    lambda_p = Config.FATOR_LAMBDA_P_FLT * math.sqrt(E / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)',
        'formula': '\\lambda_p = 1,76 \\sqrt{\\frac{{E}}{{f_y}}}',
        'valores': {'E': E, 'f_y': fy},
        'valor': lambda_p,
        'ref': 'Tabela F.1',
        'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        verificacao_texto = f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}"
        conclusao_texto = "SEÇÃO COMPACTA - O regime de flambagem é Plástico."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'REGIME PLÁSTICO',
            'verif_for_calc': 'lambda_p'
        })

        Mrdx = Mp / Config.GAMMA_A1
        detalhes['Mrdx_calc'] = {
            'desc': 'Momento Resistente (Regime Plástico)',
            'formula': 'M_{rd} = \\frac{{M_p}}{{\\gamma_{{a1}}}}',
            'valores': {'M_p': Mp, '\\gamma_{{a1}}': Config.GAMMA_A1},
            'valor': Mrdx,
            'unidade': 'kN.cm',
            'ref': 'Eq. F-1'
        }
    else:
        verificacao_texto = f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}"
        conclusao_texto = "SEÇÃO NÃO COMPACTA - O regime de flambagem é Inelástico ou Elástico."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'NECESSÁRIO VERIFICAR REGIME',
            'verif_for_calc': 'lambda_p'
        })

        sigma_r = Config.FATOR_SIGMA_R * fy
        detalhes['passos_calculo'].append({
            'desc': 'Tensão Residual (σr)',
            'formula': '\\sigma_r = 0,3 \\times f_y',
            'valores': {'f_y': fy},
            'valor': sigma_r,
            'unidade': 'kN/cm²',
            'verif_id': 'sigma_r'
        })

        Mr = (fy - sigma_r) * Wx
        detalhes['passos_calculo'].append({
            'desc': 'Momento de Escoamento Residual (Mr)',
            'formula': 'M_r = (f_y - \\sigma_r) \\times W_x',
            'valores': {'f_y': fy, '\\sigma_r': sigma_r, 'W_x': Wx},
            'valor': Mr,
            'unidade': 'kN.cm',
            'verif_id': 'Mr'
        })

        beta1 = ((fy - sigma_r) * Wx) / (E * J) if E * J != 0 else 0
        detalhes['passos_calculo'].append({
            'desc': 'Parâmetro β1',
            'formula': '\\beta_1 = \\frac{(f_y - \\sigma_r) \\times W_x}{E \\times J}',
            'valores': {'f_y': fy, '\\sigma_r': sigma_r, 'W_x': Wx, 'E': E, 'J': J},
            'valor': beta1,
            'unidade': '',
            'verif_id': 'beta1'
        })

        lambda_r = float('inf')
        if ry > 0 and beta1 > 0 and J > 0 and Cw > 0 and Iy > 0:
            termo_sqrt1 = 1 + (27 * Cw * (beta1**2) / Iy)
            termo_sqrt2 = 1 + math.sqrt(termo_sqrt1) if termo_sqrt1 >= 0 else 1
            lambda_r = (1.38 * math.sqrt(Iy * J) / (ry * beta1 * J)) * math.sqrt(termo_sqrt2)

        detalhes['passos_calculo'].append({
            'desc': 'Esbeltez Limite Inelástica (λr)',
            'formula': '\\lambda_r = 1,38 \\frac{\\sqrt{I_y \\times J}}{r_y \\times \\beta_1 \\times J} \\sqrt{1 + \\sqrt{1+\\frac{27 \\times C_w \\times \\beta_1^2}{I_y}}}',
            'valores': {'I_y': Iy, 'J': J, 'r_y': ry, '\\beta_1': beta1, 'C_w': Cw},
            'valor': lambda_r,
            'verif_id': 'lambda_r'
        })

        if lambda_val <= lambda_r:
            verificacao_texto = f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}"
            conclusao_texto = "REGIME INELÁSTICO - Flambagem no regime inelástico."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME INELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx_calc = (Cb / Config.GAMMA_A1) * (Mp - (Mp - Mr) * ((lambda_val - lambda_p) / (lambda_r - lambda_p)))
            Mp_gamma = Mp / Config.GAMMA_A1
            Mrdx = min(Mrdx_calc, Mp_gamma)

            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Regime Inelástico)',
                'formula': 'M_{rd,calc} = \\frac{{C_b}}{{\\gamma_{{a1}}}} [M_p - (M_p - M_r) (\\frac{{\\lambda - \\lambda_p}}{{\\lambda_r - \\lambda_p}})]',
                'valores': {'C_b': Cb, '\\gamma_{{a1}}': Config.GAMMA_A1, 'M_p': Mp, 'M_r': Mr,  '\\lambda': lambda_val, '\\lambda_p': lambda_p, '\\lambda_r': lambda_r},
                'valor': Mrdx_calc,
                'unidade': 'kN.cm',
                'ref': 'Eq. F-2'
            }

            detalhes['verificacao_limite'] = {
                'desc': 'Verificação do Limite de Plastificação',
                'texto': f"""
                    $M_{{rd,calc}} = {Mrdx_calc/100:.2f} \\, kNm$
                    $M_{{p,rd}} = \\frac{{M_p}}{{\\gamma_{{a1}}}} = {Mp_gamma/100:.2f} \\, kNm$
                    $M_{{rd}} = \\min(M_{{rd,calc}}; M_{{p,rd}}) = \\mathbf{{{Mrdx/100:.2f}}} \\, kNm$"""
            }
        else:
            verificacao_texto = f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}"
            conclusao_texto = "REGIME ELÁSTICO - Flambagem no regime elástico."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME ELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mcr = 0
            if Lb**2 > 0 and Iy > 0 and Cw > 0 and J > 0:
                Mcr = ((Cb * (math.pi**2) * E * Iy) / (Lb**2)) * math.sqrt((Cw/Iy) * (1 + (0.039 * J * (Lb**2) / Cw)))

            Mrdx = Mcr / Config.GAMMA_A1

            detalhes['passos_calculo'].append({
                'desc': 'Momento Crítico Elástico (Mcr)',
                'formula': 'M_{cr} = \\frac{{C_b \\times \\pi^2 \\times E \\times I_y}}{{L_b^2}} \\sqrt{{\\frac{{C_w}}{{I_y}}(1 + 0,039 \\times \\frac{{J \\times L_b^2}}{{C_w}})}}',
                'valores': {'C_b': Cb, '\\pi^2': math.pi**2, 'E': E, 'I_y': Iy, 'L_b': Lb, 'C_w': Cw, 'J': J},
                'valor': Mcr,
                'unidade': 'kN.cm',
                'ref': 'Eq. F-4',
                'verif_id': 'Mcr'
            })

            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Regime Elástico)',
                'formula': 'M_{rd} = \\frac{{M_{{cr}}}}{{\\gamma_{{a1}}}}',
                'valores': {'M_{{cr}}': Mcr, '\\gamma_{{a1}}': Config.GAMMA_A1},
                'valor': Mrdx,
                'unidade': 'kN.cm',
                'ref': 'Eq. F-1'
            }

    detalhes['Mrdx'] = Mrdx
    return detalhes

def _calcular_mrdx_flm(props, fy, tipo_fabricacao, E):
    bf, tf, Zx, Wx, h, tw = props['bf'], props['tf'], props['Zx'], props['Wx'], props['h'], props['tw']
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Mp = Zx * fy
    detalhes['passos_calculo'].append({
        'desc': 'Momento de Plastificação',
        'formula': 'M_p = Z_x \\times f_y',
        'valores': {'Z_x': Zx, 'f_y': fy},
        'valor': Mp,
        'unidade': 'kN.cm',
        'verif_id': 'Mp'
    })

    lambda_val = (bf / 2) / tf if tf > 0 else float('inf')
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez da Mesa (λ = bf/2tf)',
        'formula': '\\lambda = \\frac{{b_f/2}}{{t_f}}',
        'valores': {'b_f': bf, 't_f': tf},
        'valor': lambda_val,
        'verif_id': 'lambda'
    })

    lambda_p = Config.FATOR_LAMBDA_P_FLM * math.sqrt(E / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)',
        'formula': '\\lambda_p = 0,38 \\sqrt{{\\frac{{E}}{{f_y}}}}',
        'valores': {'E': E, 'f_y': fy},
        'valor': lambda_p,
        'ref': 'Tabela F.1',
        'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        verificacao_texto = f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}"
        conclusao_texto = "MESA COMPACTA - Não ocorre flambagem local da mesa."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'REGIME PLÁSTICO',
            'verif_for_calc': 'lambda_p'
        })

        Mrdx = Mp / Config.GAMMA_A1
        detalhes['Mrdx_calc'] = {
            'desc': 'Momento Resistente (Mesa Compacta)',
            'formula': 'M_{rd} = \\frac{{M_p}}{{\\gamma_{{a1}}}}',
            'valores': {'M_p': Mp, '\\gamma_{{a1}}': Config.GAMMA_A1},
            'valor': Mrdx,
            'unidade': 'kN.cm'
        }
    else:
        verificacao_texto = f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}"
        conclusao_texto = "MESA NÃO COMPACTA - Verificar se é semicompacta ou esbelta."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'NECESSÁRIO VERIFICAR REGIME',
            'verif_for_calc': 'lambda_p'
        })

        sigma_r = Config.FATOR_SIGMA_R * fy
        detalhes['passos_calculo'].append({
            'desc': 'Tensão Residual (σr)',
            'formula': '\\sigma_r = 0,3 \\times f_y',
            'valores': {'f_y': fy},
            'valor': sigma_r,
            'unidade': 'kN/cm²',
            'verif_id': 'sigma_r'
        })

        Mr = (fy - sigma_r) * Wx
        detalhes['passos_calculo'].append({
            'desc': 'Momento de Escoamento Residual (Mr)',
            'formula': 'M_r = (f_y - \\sigma_r) \\times W_x',
            'valores': {'f_y': fy, '\\sigma_r': sigma_r, 'W_x': Wx},
            'valor': Mr,
            'unidade': 'kN.cm',
            'verif_id': 'Mr'
        })

        if tipo_fabricacao == "Laminado":
            lambda_r = Config.FATOR_LAMBDA_R_FLM_LAMINADO * math.sqrt(E / (fy - sigma_r)) if (fy - sigma_r) > 0 else float('inf')
            lambda_r_formula_str = '\\lambda_r = 0,83 \\sqrt{{\\frac{{E}}{{f_y - \\sigma_r}}}}'
            detalhes['passos_calculo'].append({
                'desc': 'Esbeltez Limite Semicompacta (λr) - Laminado',
                'formula': lambda_r_formula_str,
                'valores': {'E': E, 'f_y': fy, '\\sigma_r': sigma_r},
                'valor': lambda_r,
                'ref': 'Tabela F.1',
                'verif_id': 'lambda_r'
            })
        else: # Soldado
            kc_val = 4 / math.sqrt(h/tw) if (h/tw) > 0 else 0.76
            kc = max(0.35, min(kc_val, 0.76))
            detalhes['passos_calculo'].append({
                'desc': 'Parâmetro kc',
                'formula': 'k_c = \\frac{4}{\\sqrt{h/t_w}} \\, (0,35 \\le k_c \\le 0,76)',
                'valores': {'h': h, 't_w': tw},
                'valor': kc,
                'unidade': '',
                'verif_id': 'kc'
            })

            lambda_r = Config.FATOR_LAMBDA_R_FLM_SOLDADO * math.sqrt(E * kc / (fy - sigma_r)) if (fy - sigma_r) > 0 and kc > 0 else float('inf')
            lambda_r_formula_str = '\\lambda_r = 0,95 \\sqrt{{\\frac{E \\times k_c}{{f_y - \\sigma_r}}}}'
            detalhes['passos_calculo'].append({
                'desc': 'Esbeltez Limite Semicompacta (λr) - Soldado',
                'formula': lambda_r_formula_str,
                'valores': {'E': E, 'k_c': kc, 'f_y': fy, '\\sigma_r': sigma_r},
                'valor': lambda_r,
                'ref': 'Tabela F.1',
                'verif_id': 'lambda_r'
            })

        if lambda_val <= lambda_r:
            verificacao_texto = f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}"
            conclusao_texto = "MESA SEMICOMPACTA - O regime é de transição."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME INELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx = (1 / Config.GAMMA_A1) * (Mp - (Mp - Mr) * ((lambda_val - lambda_p) / (lambda_r - lambda_p)))
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Mesa Semicompacta)',
                'formula': 'M_{rd} = \\frac{{1}}{{\\gamma_{{a1}}}} [M_p - (M_p - M_r) (\\frac{{\\lambda - \\lambda_p}}{{\\lambda_r - \\lambda_p}})]',
                'valores': {'\\gamma_{{a1}}': Config.GAMMA_A1, 'M_p': Mp, 'M_r': Mr,  '\\lambda': lambda_val, '\\lambda_p': lambda_p, '\\lambda_r': lambda_r},
                'valor': Mrdx,
                'unidade': 'kN.cm'
            }
        else:
            verificacao_texto = f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}"
            conclusao_texto = "MESA ESBELTA - O regime de flambagem é elástico."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME ELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            if tipo_fabricacao == "Laminado":
                Mcr = (0.69 * E * Wx) / (lambda_val**2) if lambda_val > 0 else 0
                detalhes['passos_calculo'].append({
                    'desc': 'Momento Crítico (Mcr) - Laminado',
                    'formula': 'M_{cr} = \\frac{0,69 \\times E \\times W_x}{\\lambda^2}',
                    'valores': {'E': E, 'W_x': Wx, '\\lambda': lambda_val},
                    'valor': Mcr,
                    'unidade': 'kN.cm',
                    'verif_id': 'Mcr'
                })
            else: # Soldado
                kc = detalhes['passos_calculo'][-2]['valor']
                Mcr = (0.90 * E * kc * Wx) / (lambda_val**2) if lambda_val > 0 else 0
                detalhes['passos_calculo'].append({
                    'desc': 'Momento Crítico (Mcr) - Soldado',
                    'formula': 'M_{cr} = \\frac{0,90 \\times E \\times k_c \\times W_x}{\\lambda^2}',
                    'valores': {'E': E, 'k_c': kc, 'W_x': Wx, '\\lambda': lambda_val},
                    'valor': Mcr,
                    'unidade': 'kN.cm',
                    'verif_id': 'Mcr'
                })

            Mrdx = Mcr / Config.GAMMA_A1
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Mesa Esbelta)',
                'formula': 'M_{rd} = \\frac{M_{cr}}{\\gamma_{a1}}',
                'valores': {'M_{cr}': Mcr, '\\gamma_{a1}': Config.GAMMA_A1},
                'valor': Mrdx,
                'unidade': 'kN.cm'
            }

    detalhes['Mrdx'] = Mrdx
    return detalhes

def _calcular_mrdx_fla(props, fy, E):
    h, tw, Zx, Wx = props['h'], props['tw'], props['Zx'], props['Wx']
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Mp = Zx * fy
    detalhes['passos_calculo'].append({
        'desc': 'Momento de Plastificação',
        'formula': 'M_p = Z_x \\times f_y',
        'valores': {'Z_x': Zx, 'f_y': fy},
        'valor': Mp,
        'unidade': 'kN.cm',
        'verif_id': 'Mp'
    })

    lambda_val = h / tw if tw > 0 else float('inf')
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez da Alma (λ = h/tw)',
        'formula': '\\lambda = \\frac{{h}}{{t_w}}',
        'valores': {'h': h, 't_w': tw},
        'valor': lambda_val,
        'verif_id': 'lambda'
    })

    lambda_p = Config.FATOR_LAMBDA_P_FLA * math.sqrt(E / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)',
        'formula': '\\lambda_p = 3,76 \\sqrt{{\\frac{{E}}{{f_y}}}}',
        'valores': {'E': E, 'f_y': fy},
        'valor': lambda_p,
        'ref': 'Tabela F.1',
        'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        verificacao_texto = f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}"
        conclusao_texto = "ALMA COMPACTA - Não ocorre flambagem local da alma."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'REGIME PLÁSTICO',
            'verif_for_calc': 'lambda_p'
        })

        Mrdx = Mp / Config.GAMMA_A1
        detalhes['Mrdx_calc'] = {
            'desc': 'Momento Resistente (Alma Compacta)',
            'formula': 'M_{rd} = \\frac{{M_p}}{{\\gamma_{{a1}}}}',
            'valores': {'M_p': Mp, '\\gamma_{{a1}}': Config.GAMMA_A1},
            'valor': Mrdx,
            'unidade': 'kN.cm'
        }
    else:
        verificacao_texto = f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}"
        conclusao_texto = "ALMA NÃO COMPACTA - Verificar se é semicompacta ou esbelta."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'NECESSÁRIO VERIFICAR REGIME',
            'verif_for_calc': 'lambda_p'
        })

        lambda_r = Config.FATOR_LAMBDA_R_FLA * math.sqrt(E / fy)
        detalhes['passos_calculo'].append({
            'desc': 'Esbeltez Limite Semicompacta (λr)',
            'formula': '\\lambda_r = 5,70 \\sqrt{{\\frac{{E}}{{f_y}}}}',
            'valores': {'E': E, 'f_y': fy},
            'valor': lambda_r,
            'ref': 'Tabela F.1',
            'verif_id': 'lambda_r'
        })

        Mr = fy * Wx
        detalhes['passos_calculo'].append({
            'desc': 'Momento de Escoamento (Mr)',
            'formula': 'M_r = f_y \\times W_x',
            'valores': {'f_y': fy, 'W_x': Wx},
            'valor': Mr,
            'unidade': 'kN.cm',
            'verif_id': 'Mr'
        })

        if lambda_val <= lambda_r:
            verificacao_texto = f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}"
            conclusao_texto = "ALMA SEMICOMPACTA - O regime é de transição."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME INELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx = (1 / Config.GAMMA_A1) * (Mp - (Mp - Mr) * ((lambda_val - lambda_p) / (lambda_r - lambda_p)))
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Alma Semicompacta)',
                'formula': 'M_{rd} = \\frac{{1}}{{\\gamma_{{a1}}}} [M_p - (M_p - M_r) (\\frac{{\\lambda - \\lambda_p}}{{\\lambda_r - \\lambda_p}})]',
                'valores': {'\\gamma_{{a1}}': Config.GAMMA_A1, 'M_p': Mp, 'M_r': Mr,  '\\lambda': lambda_val, '\\lambda_p': lambda_p, '\\lambda_r': lambda_r},
                'valor': Mrdx,
                'unidade': 'kN.cm'
            }
        else:
            verificacao_texto = f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}"
            conclusao_texto = "ALMA ESBELTA - Regime elástico (Ver Anexo H da NBR 8800)"
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME ELÁSTICO - NÃO IMPLEMENTADO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx = 0
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Alma Esbelta)',
                'formula': 'N/A - Ver Anexo H da NBR 8800',
                'valores': {},
                'valor': Mrdx,
                'unidade': 'kN.cm',
                'ref': 'Perfil com alma esbelta. Ver Anexo H.'
            }

    detalhes['Mrdx'] = Mrdx
    return detalhes

def _calcular_vrd(props, fy, usa_enrijecedores, a_enr, E):
    d, h, tw = props['d'], props['h'], props['tw']
    lambda_val = h / tw if tw > 0 else float('inf')
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Vpl = Config.FATOR_VP * d * tw * fy
    detalhes['passos_calculo'].append({
        'desc': 'Força Cortante de Plastificação', 'formula': 'V_{pl} = 0,60 \\times d \\times t_{w} \\times f_{y}',
        'valores': {'d': d, 't_{w}': tw, 'f_{y}': fy}, 'valor': Vpl, 'unidade': 'kN', 'verif_id': 'Vpl'
    })

    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez da Alma (λ)', 'formula': '\\lambda = \\frac{h}{t_w}',
        'valores': {'h': h, 't_w': tw}, 'valor': lambda_val, 'verif_id': 'lambda'
    })

    if not usa_enrijecedores or a_enr <= 0:
        kv = 5.0
        detalhes['passos_calculo'].append({'desc': 'Fator de Flambagem (kv) - Alma sem enrijecedores', 'formula': 'k_v = 5,0', 'valores': {}, 'valor': kv})
    else:
        a_h_ratio = a_enr / h
        detalhes['passos_calculo'].append({
            'desc': 'Relação de Enrijecedores (a/h)',
            'formula': '\\frac{a}{h}',
            'formula_expandida': f'\\frac{{\\mathbf{{{a_enr:.2f}}}}}{{\\mathbf{{{h:.2f}}}}}',
            'valores': {},
            'valor': a_h_ratio
        })

        check1_pass = a_h_ratio < 3
        detalhes['passos_calculo'].append({
            'type': 'verification', 'desc': 'Verificação do Espaçamento Máximo',
            'lhs_value': a_h_ratio, 'comparator': '<', 'rhs_value': 3, 'passed': check1_pass,
            'conclusion_pass': 'O espaçamento entre enrijecedores atende ao critério inicial.',
            'conclusion_fail': 'O espaçamento é muito grande (a/h ≥ 3), tornando os enrijecedores ineficazes.'
        })

        check2_pass = False
        if check1_pass:
            limite_esbeltez = (260 / lambda_val)**2 if lambda_val > 0 else float('inf')
            detalhes['passos_calculo'].append({
                'desc': 'Limite de Esbeltez da Chapa', 'formula': 'Limite = (\\frac{260}{\\lambda})^2',
                'valores': {'\\lambda': lambda_val}, 'valor': limite_esbeltez
            })
            check2_pass = a_h_ratio < limite_esbeltez
            detalhes['passos_calculo'].append({
                'type': 'verification', 'desc': 'Verificação da Esbeltez da Chapa',
                'lhs_value': a_h_ratio, 'comparator': '<', 'rhs_value': limite_esbeltez, 'passed': check2_pass,
                'conclusion_pass': 'A alma não é esbelta demais para o espaçamento dos enrijecedores.',
                'conclusion_fail': 'A alma é muito esbelta para este espaçamento (a/h ≥ (260/λ)²), tornando os enrijecedores ineficazes.'
            })

        if check1_pass and check2_pass:
            denominador_val = a_h_ratio**2
            kv = 5 + (5 / denominador_val)
            kv_formula_expandida = (f"k_v = 5 + \\frac{{5}}{{(\\frac{{\\mathbf{{{a_enr:.2f}}}}}{{\\mathbf{{{h:.2f}}}}})^2}} = 5 + \\frac{{5}}{{(\\mathbf{{{a_h_ratio:.2f}}})^2}} = 5 + \\frac{{5}}{{\\mathbf{{{denominador_val:.2f}}}}}")
            detalhes['passos_calculo'].append({
                'desc': 'Cálculo Final de kv (Enrijecedores Efetivos)', 'formula': 'k_v = 5 + \\frac{5}{(a/h)^2}', 'formula_expandida': kv_formula_expandida,
                'valores': {'(a/h)': a_h_ratio}, 'valor': kv
            })
        else:
            kv = 5.0
            detalhes['passos_calculo'].append({'desc': 'Cálculo Final de kv (Enrijecedores Ineficazes)', 'formula': 'k_v = 5,0', 'valores': {}, 'valor': kv})

    lambda_p = Config.FATOR_LAMBDA_P_VRD * math.sqrt((kv * E) / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)', 'formula': '\\lambda_p = 1,10 \\sqrt{\\frac{k_v \\times E}{f_y}}',
        'valores': {'k_v': kv, 'E': E, 'f_y': fy}, 'valor': lambda_p, 'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        Vrd = Vpl / Config.GAMMA_A1
        detalhes['passos_verificacao'].append({'titulo': 'Verificação 1: λ ≤ λp?', 'texto': f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}", 'conclusao': "ESCOAMENTO DA ALMA - Resistência governada pelo escoamento.", 'regime': 'REGIME PLÁSTICO', 'verif_for_calc': 'lambda_p'})
        detalhes['Vrd_calc'] = {'desc': 'Cortante Resistente (Escoamento)', 'formula': 'V_{rd} = \\frac{V_{pl}}{\\gamma_{a1}}', 'valores': {'V_{pl}': Vpl, '\\gamma_{a1}': Config.GAMMA_A1}, 'valor': Vrd, 'unidade': 'kN'}
    else:
        detalhes['passos_verificacao'].append({'titulo': 'Verificação 1: λ ≤ λp?', 'texto': f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}", 'conclusao': "FLAMBAGEM POR CISALHAMENTO - O regime é Inelástico ou Elástico.", 'regime': 'NECESSÁRIO VERIFICAR REGIME', 'verif_for_calc': 'lambda_p'})
        lambda_r = Config.FATOR_LAMBDA_R_VRD * math.sqrt((kv * E) / fy)
        detalhes['passos_calculo'].append({'desc': 'Esbeltez Limite Inelástica (λr)', 'formula': '\\lambda_r = 1,37 \\sqrt{\\frac{k_v \\times E}{f_y}}', 'valores': {'k_v': kv, 'E': E, 'f_y': fy}, 'valor': lambda_r, 'verif_id': 'lambda_r'})
        if lambda_val <= lambda_r:
            detalhes['passos_verificacao'].append({'titulo': 'Verificação 2: λ ≤ λr?', 'texto': f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}", 'conclusao': "FLAMBAGEM INELÁSTICA - Regime de transição por cisalhamento.", 'regime': 'REGIME INELÁSTICO', 'verif_for_calc': 'lambda_r'})
            Vrd = (lambda_p / lambda_val) * (Vpl / Config.GAMMA_A1) if lambda_val > 0 else 0
            detalhes['Vrd_calc'] = {'desc': 'Cortante Resistente (Flambagem Inelástica)', 'formula': 'V_{rd} = \\frac{\\lambda_p}{\\lambda} \\times \\frac{V_{pl}}{\\gamma_{a1}}', 'valores': {'\\lambda_p': lambda_p, '\\lambda': lambda_val, 'V_{pl}': Vpl, '\\gamma_{a1}': Config.GAMMA_A1}, 'valor': Vrd, 'unidade': 'kN'}
        else:
            detalhes['passos_verificacao'].append({'titulo': 'Verificação 2: λ ≤ λr?', 'texto': f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}", 'conclusao': "FLAMBAGEM ELÁSTICA - Regime elástico por cisalhamento.", 'regime': 'REGIME ELÁSTICO', 'verif_for_calc': 'lambda_r'})
            Vrd = (Config.FATOR_VRD_ELASTICO * (lambda_p / lambda_val)**2) * (Vpl / Config.GAMMA_A1) if lambda_val > 0 else 0
            detalhes['Vrd_calc'] = {'desc': 'Cortante Resistente (Flambagem Elástica)', 'formula': 'V_{rd} = 1,24 (\\frac{\\lambda_p}{\\lambda})^2 \\times \\frac{V_{pl}}{\\gamma_{a1}}', 'valores': {'\\lambda_p': lambda_p, '\\lambda': lambda_val, 'V_{pl}': Vpl, '\\gamma_{a1}': Config.GAMMA_A1}, 'valor': Vrd, 'unidade': 'kN'}

    detalhes['Vrd'] = Vrd
    return detalhes

# ==============================================================================
# 3. FUNÇÕES DE GERAÇÃO DE INTERFACE E GRÁFICOS
# ==============================================================================

def create_excel_with_colors(df_list, sheet_names):
    """
    Cria um arquivo Excel com múltiplas abas, aplicando formatação de cores
    baseada na eficiência dos perfis.
    """
    output = io.BytesIO()
    workbook = openpyxl.Workbook()

    # Remova a folha padrão criada automaticamente
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    for df, sheet_name in zip(df_list, sheet_names):
        sheet = workbook.create_sheet(title=sheet_name)

        # Escreva os cabeçalhos
        for col_idx, col_name in enumerate(df.columns, 1):
            sheet.cell(row=1, column=col_idx, value=col_name)

        # Defina os estilos de cores
        fill_fail = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') # Vermelho
        fill_warning_high = PatternFill(start_color='FFEBAE', end_color='FFEBAE', fill_type='solid') # Amarelo escuro (95-100%)
        fill_warning_low = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid') # Amarelo claro (80-95%)
        fill_pass = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid') # Verde
        
        # Escreva os dados e aplique as cores
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Regras de formatação para as colunas de eficiência
                if 'Ef.' in df.columns[col_idx-1]:
                    try:
                        efficiency = float(value)
                        if efficiency > 100.1:
                            cell.fill = fill_fail
                        elif efficiency > 95:
                            cell.fill = fill_warning_high
                        elif efficiency > 80:
                            cell.fill = fill_warning_low
                        else:
                            cell.fill = fill_pass
                    except (ValueError, TypeError):
                        pass

    workbook.save(output)
    output.seek(0)
    return output

# Mantenha esta função como está:
def create_professional_header():
    logo_url = "https://lh3.googleusercontent.com/a/ACg8ocKplHKwLPBbUbVCvtwvTPhn0aboS42tEQxuNtiVPVZ7Xboh1pk=s96-c" # Substitua pela URL da sua logo
    
    st.markdown(f"""
    <div class="pro-header">
        <div class="header-content">
            <img src="{logo_url}" alt="Logo HQ Engenharia">
            <h1 class="gradient-text">Calculadora Estrutural de Perfis</h1>
            <p>Análise de Perfis Metálicos | {Config.NOME_NORMA}</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

def create_metrics_dashboard(input_params):
    """Cria um dashboard com métricas principais do projeto e esforços."""
    st.markdown("### 📊 Parâmetros do Projeto")
    
    msd_value = input_params.get('Msd', 0)
    vsd_value = input_params.get('Vsd', 0)
    cb_value = input_params.get('Cb_projeto', 1.0)
    
    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
    
    with col1:
        st.metric(label="📐 Norma", value="NBR 8800")
    with col2:
        st.metric(label="⚡ Módulo E", value=f"{input_params['E_aco']:.0f} kN/cm²")
    with col3:
        st.metric(label="🛡️ γa1", value="1,10")
    with col4:
        st.metric(label="📏 Vão", value=f"{input_params['L_cm']/100:.2f} m")
    with col5:
        st.metric(label="🔥 fy", value=f"{input_params['fy_aco']:.1f} kN/cm²")
    
    with col6:
        st.metric(
            label=" Msd (Momento)",
            value=f"{msd_value/100:.2f} kNm" if msd_value > 0 else "-",
            help="Momento Fletor Solicitante de Cálculo"
        )
    with col7:
        st.metric(
            label=" Vsd (Cortante)",
            value=f"{vsd_value:.2f} kN" if vsd_value > 0 else "-",
            help="Força Cortante Solicitante de Cálculo"
        )
    
    st.markdown(f"**Fator Cb:** {cb_value:.2f} | **Lb:** {input_params['Lb_projeto']:.2f} cm | **Flecha Limite:** L/{input_params['limite_flecha_divisor']:.0f}")

def style_classic_dataframe(df):
    """Aplica estilização clássica com cores sólidas ao DataFrame."""
    def color_efficiency(val):
        if pd.isna(val) or not isinstance(val, (int, float)): return ''
        if val > 100:   color = '#f8d7da'
        elif val > 95:  color = '#ffeeba'
        elif val > 80:  color = '#fff3cd'
        else:           color = '#d4edda'
        return f'background-color: {color}'

    def style_status(val):
        if val == 'APROVADO':
            return 'background-color: #d4edda; font-weight: bold; color: #155724;'
        elif val == 'REPROVADO':
            return 'background-color: #f8d7da; font-weight: bold; color: #721c24;'
        return ''

    efficiency_cols = [col for col in df.columns if '%' in col]
    
    styled_df = df.style.applymap(color_efficiency, subset=efficiency_cols)
    
    if 'Status' in df.columns:
        styled_df = styled_df.applymap(style_status, subset=['Status'])
        
    format_dict = {"Peso (kg/m)": "{:.2f}", **{col: "{:.1f}" for col in efficiency_cols}}
    return styled_df.format(format_dict)

def create_top_profiles_chart(df_approved, top_n=10):
    if df_approved.empty: return None
    df_top = df_approved.head(top_n).sort_values(by='Peso (kg/m)', ascending=False)
    fig = go.Figure(go.Bar(
        y=df_top['Perfil'], x=df_top['Peso (kg/m)'], orientation='h',
        text=[f'{w:.2f} kg/m' for w in df_top['Peso (kg/m)']], textposition='auto',
        marker=dict(color=df_top['Peso (kg/m)'], colorscale='Blues_r', colorbar=dict(title="Peso")),
        hovertemplate='<b>%{y}</b><br>Peso: %{x:.2f} kg/m<extra></extra>'
    ))
    fig.update_layout(
        title={'text': f'🏆 Top {top_n} Perfis Mais Leves (Aprovados)', 'x': 0.5},
        xaxis_title='Peso (kg/m)', yaxis_title='Perfil', template='plotly_white', height=500, margin=dict(l=150)
    )
    return fig

def create_profile_efficiency_chart(perfil_nome, eficiencias):
    """
    Cria um gráfico de barras comparando as eficiências de um perfil.
    """
    labels = list(eficiencias.keys())
    values = [min(v, 150) for v in eficiencias.values()]
    
    colors = ['#1e40af' if v < 90 else '#60a5fa' if v <= 100 else '#ef4444' for v in values]
    
    fig = go.Figure(data=[
        go.Bar(
            x=labels,
            y=values,
            text=[f'{v:.1f}%' for v in eficiencias.values()],
            textposition='auto',
            marker_color=colors
        )
    ])
    
    fig.add_hline(y=100, line_dash="dash", line_color="#10b981",
                    annotation_text="Limite de Aprovação (100%)",
                    annotation_position="bottom right")

    fig.update_layout(
        title=f'Análise de Eficiência para o Perfil: {perfil_nome}',
        yaxis_title='Eficiência (%)',
        xaxis_title='Verificação',
        yaxis_range=[0, max(max(values), 100) + 10],
        template='plotly_white',
    )
    return fig
brazilia_tz = pytz.timezone('America/Sao_Paulo')
def create_professional_memorial_html(perfil_nome, perfil_tipo, resultados, input_details, projeto_info):
    conteudo_memorial = f"""
    <h2>1. Resumo Executivo</h2>
    <div class="result-highlight">{resultados['resumo_html']}</div>
    

    <h2>2. Dados de Entrada e Solicitações</h2>
    <div class="info-card">
        <h3>2.1. Propriedades do Perfil e Materiais</h3>
        {input_details}
    </div>
    {resultados.get('esforcos_html', '')}
    {resultados.get('cb_calc_html', '')}
    <h2>3. Verificações de Resistência (ELU)</h2>
    {resultados['passo_a_passo_html']}
    """
    html_template = f"""
    <!DOCTYPE html><html lang="pt-BR"><head>
        <meta charset="UTF-8"><title>Memorial de Cálculo - {perfil_nome}</title>
        {HTML_TEMPLATE_CSS_PRO}
        <script type="text/javascript" async src="https://cdnjs.cloudflare.com/ajax/libs/mathjax/2.7.7/MathJax.js?config=TeX-MML-AM_CHTML"></script>
    </head><body><div class="container">
        <div class="pro-header">
            <h1>Memorial de Cálculo Estrutural</h1>
            <p><strong>{perfil_nome}</strong> ({perfil_tipo})</p>
        </div>
        <div class="info-card">
            <h3>📋 Identificação do Projeto</h3>
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 1rem;">
                <div><strong>Projeto:</strong> {projeto_info['nome']}</div>
                <div><strong>Engenheiro:</strong> {projeto_info['engenheiro']}</div>
                <div><strong>Data:</strong> {projeto_info['data']}</div>
                <div><strong>Revisão:</strong> {projeto_info['revisao']}</div>
            </div>
        </div>
        {conteudo_memorial}
        <div style="text-align: center; margin-top: 3rem; padding-top: 2rem; border-top: 1px solid var(--border); color: var(--text-secondary);">
            <p>Memorial gerado em {datetime.now(brazilia_tz).strftime('%d/%m/%Y às %H:%M')}</p>
        </div>
    </div></body></html>
    """
    return html_template

def _build_verification_block_html(title, solicitante, s_symbol, resistente, r_symbol, eficiencia, status, unit):
    status_class = "pass" if status == "APROVADO" else "fail"
    comp_symbol = "\\le" if status == "APROVADO" else ">"
    return f"""<h4>{title}</h4><div class="formula-block"><p class="formula">$${s_symbol} = {solicitante:.2f} \\, {unit}$$</p><p class="formula">$${r_symbol} = {resistente:.2f} \\, {unit}$$</p><p class="formula">$$\\text{{Verificação: }} {s_symbol} {comp_symbol} {r_symbol}$$</p><p class="formula">$$\\text{{Eficiência}} = \\frac{{{s_symbol}}}{{{r_symbol}}} = \\frac{{{solicitante:.2f}}}{{{resistente:.2f}}} = {eficiencia:.1f}\\%$$</p><div class="final-status {status_class}">{status}</div></div>"""

def _render_cb_calc_section(cb_details, Cb_final, input_mode):
    """Renderiza a seção de cálculo de Cb no memorial com passo a passo detalhado."""
    html = "<h3>2.3. Cálculo do Fator de Modificação Cb</h3>"

    if input_mode == 'Inserir Esforços Manualmente' or not cb_details or 'N/A' in cb_details.get('formula_simbolica', ''):
        html += "<div class='formula-block'>"
        html += f"<h5>Descrição:</h5><p>Fator Cb = **{Cb_final:.2f}** (Valor inserido manualmente ou não aplicável)</p>"
        html += "</div>"
        return html

    # Renderização Detalhada
    html += "<h4>Passo a Passo do Cálculo dos Momentos ($kNm$)</h4>"
    momentos_detalhes = cb_details.get('momentos', {})

    for m_key, m_data in momentos_detalhes.items():
        html += f"<div class='formula-block'>"
        html += f"<h5>Cálculo de ${m_key}$</h5>"
        
        # Renderiza cada componente de carga
        for component in m_data.get('components', []):
            html += _render_calculation_step(component)
        
        # Se houver mais de uma componente, mostra a soma
        if len(m_data.get('components', [])) > 1:
            comp_vals = [c['valor'] for c in m_data['components']]
            sum_str = ' + '.join([f'{abs(v):.2f}' for v in comp_vals])
            html += f"""<h6>Soma das Componentes para ${m_key}$</h6>
                        <p class="formula">$${m_key} = {sum_str} = \\mathbf{{{m_data['final_value']:.2f}}} \\, kNm$$</p>"""
        
        # Mostra o resultado final do momento
        html += f"<div class='final-status pass' style='font-size: 1.1em; padding: 0.75rem;'>Valor Final: ${m_key} = {m_data['final_value']:.2f} \\, kNm$</div>"
        html += "</div>"

    # Renderização do cálculo final do Cb
    html += "<div class='formula-block'>"
    html += f"<h4>Cálculo Final do Fator Cb</h4>"
    html += f"<p class='formula'>$${cb_details['formula_simbolica']}$$</p>"
    
    M_max_val = momentos_detalhes['M_max']['final_value']
    M_A_val = momentos_detalhes['M_A']['final_value']
    M_B_val = momentos_detalhes['M_B']['final_value']
    M_C_val = momentos_detalhes['M_C']['final_value']
    
    numerador_val = 12.5 * M_max_val
    denominador_val = 2.5 * M_max_val + 3 * M_A_val + 4 * M_B_val + 3 * M_C_val
        
    html += f"<h5>Cálculo numérico:</h5>"
    html += f"<p class='formula'>$$C_b = \\frac{{12,5 \\times \\mathbf{{{M_max_val:.2f}}}}}{{|2,5 \\times \\mathbf{{{M_max_val:.2f}}} + 3 \\times \\mathbf{{{M_A_val:.2f}}} + 4 \\times \\mathbf{{{M_B_val:.2f}}} + 3 \\times \\mathbf{{{M_C_val:.2f}}}|}} = \\frac{{{numerador_val:.2f}}}{{{abs(denominador_val):.2f}}}$$</p>"
    html += f"<p class='formula'>$$C_b = \\mathbf{{{Cb_final:.2f}}}$$</p>"
    html += "</div>"
    return html

def _render_esforcos_viga_section(params):
    """Gera a seção do memorial com o cálculo dos esforços solicitantes."""
    # --- AJUSTE: Nova apresentação visual para o modo de entrada manual ---
    if params['input_mode'] == 'Inserir Esforços Manualmente':
        msd_knm = params['Msd'] / 100
        vsd_kn = params['Vsd']
        return f"""
        <h3>2.2. Esforços Solicitantes de Cálculo (ELU)</h3>
        <div class="info-card">
            <div style="display: flex; justify-content: space-around; align-items: stretch; flex-wrap: wrap; gap: 1rem; padding: 1rem 0;">

                <div style="text-align: center; border: 1px solid #e5e7eb; border-radius: 12px; padding: 1.5rem; width: 45%; min-width: 200px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                    <span style="font-size: 2em; line-height: 1;">🔄</span>
                    <h5 style="margin: 0.5rem 0; color: var(--text-secondary); font-weight: 500;">Momento Fletor ($M_{{sd}}$)</h5>
                    <p style="font-size: 2.2em; font-weight: 700; color: var(--primary-color); margin: 0; line-height: 1.2;">
                        {msd_knm:.2f}
                        <span style="font-size: 0.5em; font-weight: 500; color: var(--text-primary);">kNm</span>
                    </p>
                </div>

                <div style="text-align: center; border: 1px solid #e5e7eb; border-radius: 12px; padding: 1.5rem; width: 45%; min-width: 200px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                    <span style="font-size: 2em; line-height: 1;">✂️</span>
                    <h5 style="margin: 0.5rem 0; color: var(--text-secondary); font-weight: 500;">Força Cortante ($V_{{sd}}$)</h5>
                    <p style="font-size: 2.2em; font-weight: 700; color: var(--primary-color); margin: 0; line-height: 1.2;">
                        {vsd_kn:.2f}
                        <span style="font-size: 0.5em; font-weight: 500; color: var(--text-primary);">kN</span>
                    </p>
                </div>

            </div>
            <p style="text-align: center; font-size: 0.9em; color: var(--text-secondary); margin-top: 1rem;"><i>(Valores inseridos manualmente)</i></p>
        </div>
        """

    # Renderização para o modo de cálculo automático (permanece inalterada)
    html = f"""
    <h3>2.2. Esforços Solicitantes de Cálculo (ELU)</h3>
    <div class="info-card">
        <h4>Cálculo das Cargas de Projeto (Últimas)</h4>
        <div class="formula-block">
            <h5>Largura de Influência (B)</h5>
            <p class="formula">$$B = \\frac{{b_{{esq}}}}{2} + \\frac{{b_{{dir}}}}{2} = \\frac{{ {params['larg_esq_cm']:.2f} }}{2} + \\frac{{ {params['larg_dir_cm']:.2f} }}{2} = \\mathbf{{ {params['larg_inf_total_m']:.2f} \\, m}}$$</p>

            <h5>Carga Distribuída de Serviço (q_serv)</h5>
            <p class="formula">$$q_{{serv}} = Carga_{{area}} \\times B = {params['carga_area']:.2f} \\, kN/m^2 \\times {params['larg_inf_total_m']:.2f} \\, m = \\mathbf{{ {params['q_servico_kn_m']:.2f} \\, kN/m}}$$</p>

            <h5>Carga Distribuída de Cálculo (q_ultima)</h5>
            <p class="formula">$$q_{{ultima}} = q_{{serv}} \\times \\gamma_f = {params['q_servico_kn_m']:.2f} \\times {params['gamma_f']:.2f} = \\mathbf{{ {params['q_ult_kn_cm']*100:.2f} \\, kN/m}}$$</p>
    """
    if params.get('p_load_serv'):
        html += f"""
            <h5>Carga Pontual de Cálculo (P_ultima)</h5>
            <p class="formula">$$P_{{ultima}} = P_{{serv}} \\times \\gamma_f = {params['p_load_serv'][0]:.2f} \\times {params['gamma_f']:.2f} = \\mathbf{{ {params['p_load_ult'][0]:.2f} \\, kN}}$$</p>
        """
    html += """
        </div>
    </div>
    """

    html += f"<h4>Cálculo dos Esforços para Viga {params['tipo_viga']}</h4>"
    html += "<div class='formula-block'>"

    detalhes = params['detalhes_esforcos']

    # Momento
    html += f"<h5>Momento Fletor de Cálculo (Msd)</h5>"
    if detalhes['Msd_q']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Momento devido à carga distribuída (M_q)',
            'formula': detalhes['Msd_q']['formula_simbolica'],
            'valores': {'q_{ultima}': params['q_ult_kn_cm']*100, 'L': params['L_cm']/100},
            'valor': detalhes['Msd_q']['valor'],
            'unidade': 'kN.cm'
        })
    if detalhes['Msd_p']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Momento devido à carga pontual (M_p)',
            'formula': detalhes['Msd_p']['formula_simbolica'],
            'valores': {'P_{ultima}': params['p_load_ult'][0], 'a': params['p_load_ult'][1]/100, 'b': (params['L_cm']-params['p_load_ult'][1])/100, 'L': params['L_cm']/100},
            'valor': detalhes['Msd_p']['valor'],
            'unidade': 'kN.cm'
        })

    # Soma dos momentos
    if (detalhes['Msd_q']['valor'] > 0 or detalhes['Msd_p']['valor'] > 0):
        mom_q = detalhes['Msd_q']['valor'] if detalhes['Msd_q']['valor'] > 0 else 0
        mom_p = detalhes['Msd_p']['valor'] if detalhes['Msd_p']['valor'] > 0 else 0
        if mom_q > 0 and mom_p > 0:
            html += f"<p class='formula'>$$M_{{sd, total}} = {mom_q:.2f} + {mom_p:.2f} = \\mathbf{{{params['Msd']:.2f}}} \\, kN.cm$$</p>"
        else:
            html += f"<p class='formula'>$$M_{{sd}} = \\mathbf{{{params['Msd']:.2f}}} \\, kN.cm$$</p>"

    # Cortante
    html += f"<h5>Força Cortante de Cálculo (Vsd)</h5>"
    if detalhes['Vsd_q']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Cortante devido à carga distribuída (V_q)',
            'formula': detalhes['Vsd_q']['formula_simbolica'],
            'valores': {'q_{ultima}': params['q_ult_kn_cm']*100, 'L': params['L_cm']/100},
            'valor': detalhes['Vsd_q']['valor'],
            'unidade': 'kN'
        })
    if detalhes['Vsd_p']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Cortante devido à carga pontual (V_p)',
            'formula': detalhes['Vsd_p']['formula_simbolica'],
            'valores': {'P_{ultima}': params['p_load_ult'][0], 'a': params['p_load_ult'][1]/100, 'b': (params['L_cm']-params['p_load_ult'][1])/100, 'L': params['L_cm']/100},
            'valor': detalhes['Vsd_p']['valor'],
            'unidade': 'kN'
        })

    # Soma dos cortantes
    if (detalhes['Vsd_q']['valor'] > 0 or detalhes['Vsd_p']['valor'] > 0):
        v_q = detalhes['Vsd_q']['valor'] if detalhes['Vsd_q']['valor'] > 0 else 0
        v_p = detalhes['Vsd_p']['valor'] if detalhes['Vsd_p']['valor'] > 0 else 0
        if v_q > 0 and v_p > 0:
            html += f"<p class='formula'>$$V_{{sd, total}} = {v_q:.2f} + {v_p:.2f} = \\mathbf{{{params['Vsd']:.2f}}} \\, kN$$</p>"
        else:
            html += f"<p class='formula'>$$V_{{sd}} = \\mathbf{{{params['Vsd']:.2f}}} \\, kN$$</p>"

    html += "</div>"

    return html

def perform_all_checks(props, fy_aco, Lb_projeto, Cb_projeto, L_cm, Msd, Vsd, q_serv_kn_cm, p_load_serv, tipo_viga, input_mode, tipo_fabricacao, usa_enrijecedores, a_enr, limite_flecha_divisor, projeto_info, E_aco, detalhado=False, **kwargs):
    res_flt = _calcular_mrdx_flt(props, Lb_projeto, Cb_projeto, fy_aco, E_aco)
    res_flm = _calcular_mrdx_flm(props, fy_aco, tipo_fabricacao, E_aco)
    res_fla = _calcular_mrdx_fla(props, fy_aco, E_aco)
    res_vrd = _calcular_vrd(props, fy_aco, usa_enrijecedores, a_enr, E_aco)

    if res_flt['Mrdx'] > 0: res_flt['eficiencia'] = (Msd / res_flt['Mrdx']) * 100
    else: res_flt['eficiencia'] = float('inf')
    if res_flm['Mrdx'] > 0: res_flm['eficiencia'] = (Msd / res_flm['Mrdx']) * 100
    else: res_flm['eficiencia'] = float('inf')
    if res_fla['Mrdx'] > 0: res_fla['eficiencia'] = (Msd / res_fla['Mrdx']) * 100
    else: res_fla['eficiencia'] = float('inf')

    Vrd = res_vrd['Vrd']
    if Vrd > 0: res_cis = {'Vrd': Vrd, 'eficiencia': (Vsd / Vrd) * 100, 'status': "APROVADO" if (Vsd / Vrd) * 100 <= 100.1 else "REPROVADO"}
    else: res_cis = {'Vrd': Vrd, 'eficiencia': float('inf'), 'status': "REPROVADO"}

    flecha_max, flecha_limite, eficiencia_flecha, status_flecha = 0, 0, 0, "N/A"
    detalhes_flecha = {}
    if input_mode == "Calcular a partir de Cargas na Viga":
        detalhes_flecha = calcular_flecha_maxima(tipo_viga, L_cm, E_aco, props['Ix'], q_serv_kn_cm, p_load_serv)
        flecha_max = detalhes_flecha['delta_total']
        flecha_limite = L_cm / limite_flecha_divisor if L_cm > 0 else 0
        eficiencia_flecha = (flecha_max / flecha_limite) * 100 if flecha_limite > 0 else float('inf')
        status_flecha = "APROVADO" if eficiencia_flecha <= 100.1 else "REPROVADO"
    else:
        status_flecha = "N/A"

    res_flecha = {'flecha_max': flecha_max, 'flecha_limite': flecha_limite, 'eficiencia': eficiencia_flecha, 'status': status_flecha, 'Ix': props['Ix'], 'detalhes': detalhes_flecha, 'divisor': limite_flecha_divisor}

    Mrd_final = min(res_flt['Mrdx'], res_flm['Mrdx'], res_fla['Mrdx'])
    ef_geral = (Msd / Mrd_final) * 100 if Mrd_final > 0 else float('inf')
    status_flexao = "APROVADO" if ef_geral <= 100.1 else "REPROVADO"
    res_flexao = {'Mrd': Mrd_final, 'eficiencia': ef_geral, 'status': status_flexao}

    passo_a_passo_html = ""
    if detalhado:
        passo_a_passo_html = build_step_by_step_html(L_cm, Msd, Vsd, res_flexao, res_cis, res_flecha, res_flt, res_flm, res_fla, res_vrd, input_mode)

    return res_flt, res_flm, res_fla, res_cis, res_flecha, passo_a_passo_html

def build_summary_html(Msd, Vsd, res_flt, res_flm, res_fla, res_cisalhamento, res_flecha):
    verificacoes = [
        ('Flexão (FLT)', f"{Msd/100:.2f} kNm", f"{res_flt['Mrdx']/100:.2f} kNm", res_flt['eficiencia'], "APROVADO" if res_flt['eficiencia'] <= 100.1 else "REPROVADO"),
        ('Flexão (FLM)', f"{Msd/100:.2f} kNm", f"{res_flm['Mrdx']/100:.2f} kNm", res_flm['eficiencia'], "APROVADO" if res_flm['eficiencia'] <= 100.1 else "REPROVADO"),
        ('Flexão (FLA)', f"{Msd/100:.2f} kNm", f"{res_fla['Mrdx']/100:.2f} kNm", res_fla['eficiencia'], "APROVADO" if res_fla['eficiencia'] <= 100.1 else "REPROVADO"),
        ('Cisalhamento', f"{Vsd:.2f} kN", f"{res_cisalhamento['Vrd']:.2f} kN", res_cisalhamento['eficiencia'], res_cisalhamento['status']),
        ('Flecha (ELS)', f"{res_flecha['flecha_max']:.2f} cm" if res_flecha['status'] != "N/A" else "N/A", f"≤ {res_flecha['flecha_limite']:.2f} cm" if res_flecha['status'] != "N/A" else "N/A", res_flecha['eficiencia'], res_flecha['status'])
    ]
    rows_html = ""
    for nome, sol, res, efic, status in verificacoes:
        status_class = "pass" if status == "APROVADO" else "fail"
        rows_html += f"""<tr><td>{nome}</td><td>{sol}</td><td>{res}</td><td>{efic:.1f}%</td><td class="{status_class}">{status}</td></tr>"""

    return f"""<table class="summary-table">
        <thead><tr><th>Verificação</th><th>Solicitante</th><th>Resistente</th><th>Eficiência</th><th>Status</th></tr></thead>
        <tbody>{rows_html}</tbody>
    </table>"""

def build_step_by_step_html(L, Msd, Vsd, res_flexao, res_cisalhamento, res_flecha, res_flt, res_flm, res_fla, res_vrd, input_mode):
    html = ""
    html += _render_resistance_calc_section(
        "Flambagem Lateral com Torção (FLT)", Msd, "M_{sd}", "kNm", res_flt, 'Mrdx', "M_{rd}"
    )
    html += _render_resistance_calc_section(
        "Flambagem Local da Mesa (FLM)", Msd, "M_{sd}", "kNm", res_flm, 'Mrdx', "M_{rd}"
    )
    html += _render_resistance_calc_section(
        "Flambagem Local da Alma (FLA)", Msd, "M_{sd}", "kNm", res_fla, 'Mrdx', "M_{rd}"
    )
    html += _build_verification_block_html("Verificação Final à Flexão", Msd/100, "M_{{sd}}", res_flexao['Mrd']/100, "M_{{rd}}", res_flexao['eficiencia'], res_flexao['status'], "kNm")

    html += f"<h3>3.2 Cálculo da Resistência ao Cisalhamento (Vrd)</h3>"
    html += _render_resistance_calc_section(
        "Resistência ao Cisalhamento (VRd)", Vsd, "V_{sd}", "kN", res_vrd, 'Vrd', "V_{rd}"
    )

    if input_mode == "Calcular a partir de Cargas na Viga":
        html += """<h2>4. Verificação de Serviço (ELS)</h2>"""
        html += """<h3>4.1. Cálculo da Flecha Máxima Atuante (δ_max)</h3>"""
        html += "<div class='formula-block'>"

        detalhes_flecha = res_flecha.get('detalhes', {})
        delta_q_details = detalhes_flecha.get('delta_q', {})
        if delta_q_details.get('valor', 0) > 0:
            html += f"<h5>Flecha devido à Carga Distribuída (δ_q)</h5>"
            html += f"<p class='formula'>$${delta_q_details['formula_simbolica']}$$</p>"
            html += f"<p class='formula'>$${delta_q_details['formula_numerica']} = \\mathbf{{{delta_q_details['valor']:.4f}}} \\, cm$$</p>"

        delta_p_details = detalhes_flecha.get('delta_p', {})
        if delta_p_details.get('valor', 0) > 0:
            html += f"<h5>Flecha devido à Carga Pontual (δ_p)</h5>"
            html += f"<p class='formula'>$${delta_p_details['formula_simbolica']}$$</p>"
            html += f"<p class='formula'>$${delta_p_details['formula_numerica']} = \\mathbf{{{delta_p_details['valor']:.4f}}} \\, cm$$</p>"

        html += f"<h5>Flecha Total</h5>"
        q_val = detalhes_flecha.get('delta_q', {}).get('valor', 0)
        p_val = detalhes_flecha.get('delta_p', {}).get('valor', 0)
        html += f"<p class='formula'>$$\\delta_{{max}} = {q_val:.4f} + {p_val:.4f} = \\mathbf{{{res_flecha['flecha_max']:.4f}}} \\, cm$$</p>"
        html += "</div>"

        html += "<h3>4.2. Cálculo da Flecha Limite (δ_lim)</h3>"
        html += f"""<div class="formula-block">
            <p class="formula">$$\\delta_{{lim}} = \\frac{{L}}{{{res_flecha['divisor']}}} = \\frac{{{L:.2f}}}{{{res_flecha['divisor']}}} = \\mathbf{{{res_flecha['flecha_limite']:.2f}}} \\, cm$$</p>
        </div>"""
        
        html += "<h3>4.3. Verificação Final da Flecha</h3>"
        html += _build_verification_block_html("Verificação da Flecha", res_flecha['flecha_max'], "\\delta_{{max}}", res_flecha['flecha_limite'], "\\delta_{{lim}}", res_flecha['eficiencia'], res_flecha['status'], "cm")

    return html

def _render_resistance_calc_section(title, solicitante_val, solicitante_sym, solicitante_unit, details_dict, res_key, res_sym):
    """Renderiza uma seção completa de cálculo de resistência com verificações sequenciais."""
    html = f"<h4>{title}</h4><div class='formula-block'>"

    verificacoes_map = {v.get('verif_for_calc'): v for v in details_dict.get('passos_verificacao', [])}

    for step in details_dict.get('passos_calculo', []):
        if step.get('type') == 'verification':
            html += _render_verification_step(step)
        else:
            html += _render_calculation_step(step)
            
            if step.get('verif_id') in verificacoes_map:
                verificacao_step = verificacoes_map.get(step['verif_id'])
                is_pass = any(s in verificacao_step['conclusao'] for s in ["COMPACTA", "PLÁSTICO", "ESCOAMENTO"])
                status_class = "pass" if is_pass else "fail"
                
                html += f"""
                <div class='verification-step'>
                    <h5>🔍 {verificacao_step['titulo']}</h5>
                    <p><strong>Comparação:</strong> {verificacao_step['texto']}</p>
                    <p class='conclusion {status_class}'>{verificacao_step['conclusao']}</p>
                    <p><strong>Classificação:</strong> {verificacao_step['regime']}</p>
                </div>
                """

    if details_dict.get('Mrdx_calc') or details_dict.get('Vrd_calc'):
        final_calc_key = 'Mrdx_calc' if 'Mrdx_calc' in details_dict else 'Vrd_calc'
        final_calc_info = details_dict[final_calc_key]
        html += _render_calculation_step(final_calc_info)

    if 'verificacao_limite' in details_dict:
        limite_info = details_dict['verificacao_limite']
        html += f"<h5>⚖️ {limite_info['desc']}</h5><div class='verification-step'>{limite_info['texto']}</div>"

    html += "</div>"

    res_val = details_dict.get(res_key, 0)
    solicitante_display = solicitante_val / 100 if solicitante_unit == "kNm" else solicitante_val

    if res_val > 0:
        res_display = res_val / 100 if solicitante_unit == "kNm" else res_val
        eficiencia = (solicitante_val / res_val) * 100
        status = "APROVADO" if eficiencia <= 100.1 else "REPROVADO"
        html += _build_verification_block_html(f"✅ Verificação Final - {title}", solicitante_display, solicitante_sym, res_display, res_sym, eficiencia, status, solicitante_unit)
    else:
        html += "<div class='final-status fail'>❌ REPROVADO (Resistência nula ou não implementada)</div>"

    return html

def _render_verification_step(step_dict):
    """Renderiza um bloco de verificação (comparação) no memorial."""
    desc = step_dict.get('desc', 'Verificação')
    lhs_val = step_dict.get('lhs_value', 0)
    rhs_val = step_dict.get('rhs_value', 0)
    comparator = step_dict.get('comparator', '<')
    passed = step_dict.get('passed', False)

    status_text = "ATENDE" if passed else "NÃO ATENDE"
    status_class = "pass" if passed else "fail"
    conclusion = step_dict.get('conclusion_pass', '') if passed else step_dict.get('conclusion_fail', '')

    html = f"""
    <h5><span style="font-size: 0.8em; color: #6b7280;">Passo de Verificação</span><br>📊 {desc}</h5>
    <div class="verification-step">
        <p class="formula" style="font-size: 1.2em;">
            $$ {lhs_val:.2f} \\, {comparator} \\, {rhs_val:.2f} $$
        </p>
        <div class="final-status {status_class}" style="font-size: 1.1em; padding: 0.75rem; text-transform: none;">Resultado: {status_text}</div>
        <p style="text-align: center; margin-top: 1rem;"><b>Conclusão:</b> {conclusion}</p>
    </div>
    """
    return html

def _render_calculation_step(step_dict):
    """
    Renderiza um passo de cálculo de forma padronizada e explícita:
    1. Fórmula Simbólica
    2. Fórmula Numérica com substituições = Resultado
    """
    desc = step_dict.get('desc', 'Cálculo')
    formula_simbolica = step_dict.get('formula', '')
    valores = step_dict.get('valores', {})
    formula_expandida = step_dict.get('formula_expandida')
    valor_final = step_dict.get('valor')
    unidade = step_dict.get('unidade', '')
    ref = f"<p class='ref-norma'>{step_dict.get('ref', '')}</p>" if step_dict.get('ref') else ""
    nota = step_dict.get('nota', '')

    valor_final_str = ""
    if isinstance(valor_final, (int, float)):
        if valor_final == float('inf'):
            valor_final_str = "\\infty"
        else:
            if abs(valor_final) > 0.01 or valor_final == 0:
                valor_final_str = f"{valor_final:.2f}"
            else:
                valor_final_str = f"{valor_final:.4f}"
    elif valor_final is not None:
        valor_final_str = str(valor_final)

    formula_numerica_final = None
    if formula_expandida:
        formula_numerica_final = formula_expandida
    elif valores:
        formula_numerica = formula_simbolica
        for var, val_num in valores.items():
            if isinstance(val_num, (int, float)):
                if abs(val_num) > 0.01 or val_num == 0: val_str = f"{val_num:.2f}"
                else: val_str = f"{val_num:.3f}"
            else:
                val_str = str(val_num)
            formula_numerica = formula_numerica.replace(var, f"\\mathbf{{{val_str}}}")
        formula_numerica_final = formula_numerica
    
    titulo = f"<h6>{desc}</h6>" if '(' in desc else f"<h5>📏 {desc}</h5>"
    
    html_output = f"""{titulo}
                    <p class="formula">$${formula_simbolica}$$</p>"""
    
    if formula_numerica_final:
        html_output += f"""<p class="formula">$${formula_numerica_final} = \\mathbf{{{valor_final_str}}} \\, {unidade}$$</p>"""
    elif valor_final is not None:
          html_output += f"""<p class="formula">$${formula_simbolica} = \\mathbf{{{valor_final_str}}} \\, {unidade}$$</p>"""

    html_output += f"{ref}{nota}"
    
    return html_output

# ==============================================================================
# 5. APLICAÇÃO PRINCIPAL STREAMLIT (REESTRUTURADA)
# ==============================================================================

def main():
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'detailed_analysis_html' not in st.session_state:
        st.session_state.detailed_analysis_html = None
    if 'analysis_mode' not in st.session_state:
        st.session_state.analysis_mode = "batch"
    if 'profile_efficiency_chart' not in st.session_state:
        st.session_state.profile_efficiency_chart = None

    all_sheets = load_data_from_local_file()
    if not all_sheets:
        st.stop()
    
    st.markdown(HTML_TEMPLATE_CSS_PRO, unsafe_allow_html=True)
    create_professional_header()

    with st.sidebar:
        st.markdown("## ⚙️ Configuração do Projeto")
        with st.expander("📋 Identificação do Projeto", expanded=False):
            projeto_nome = st.text_input("Nome do Projeto", "Análise Estrutural")
            engenheiro = st.text_input("Engenheiro Responsável", " ")
            data_projeto = st.date_input("Data", datetime.now())
            revisao = st.text_input("Revisão", "00")
        
        st.markdown("---")
        st.markdown("### 🏗️ Modelo Estrutural")
        tipo_viga = st.selectbox("🔗 Tipo de Viga:", ('Bi-apoiada', 'Engastada e Livre (Balanço)', 'Bi-engastada', 'Engastada e Apoiada'), key='tipo_viga')
        L_cm = st.number_input("📏 Comprimento (L, cm)", 10.0, value=500.0, step=10.0, key='L_cm')
        
        st.markdown("---")
        st.markdown("### ⚖️ Carregamento")
        input_mode = st.radio("Método de entrada:", ("Calcular a partir de Cargas na Viga", "Inserir Esforços Manualmente"), key='input_mode')

        Msd, Vsd, q_serv_kn_cm, p_load_serv = 0, 0, 0, None
        q_ult_kn_cm, p_load_ult = 0, None
        detalhes_esforcos_memorial = {}
        
        if input_mode == "Calcular a partir de Cargas na Viga":
            with st.container(border=True):
                st.subheader("Carga Distribuída (q)")
                larg_esq_cm = st.number_input("Largura da laje à esquerda (cm)", 0.0, value=200.0, step=10.0, key='larg_esq_cm')
                larg_dir_cm = st.number_input("Largura da laje à direita (cm)", 0.0, value=200.0, step=10.0, key='larg_dir_cm')
                larg_inf_total_m = (larg_esq_cm / 100.0 / 2) + (larg_dir_cm / 100.0 / 2)
                st.info(f"Largura de Influência Total (B) = {larg_inf_total_m:.2f} m")
                carga_area = st.number_input("Carga Distribuída (serviço, kN/m²)", 0.0, value=4.0, step=0.5, key='carga_area')
                st.subheader("Carga Pontual (P)")
                add_p_load = st.checkbox("Adicionar Carga Pontual (ex: parede)", key='add_p_load')
                if add_p_load:
                    p_serv_kn = st.number_input("Valor da Carga P (serviço, kN)", min_value=0.0, value=10.0, key='p_serv_kn')
                    p_pos_cm = st.number_input("Posição da Carga P (x, cm do apoio esquerdo)", min_value=0.0, max_value=L_cm, value=L_cm/2, key='p_pos_cm')
                    p_load_serv = (p_serv_kn, p_pos_cm)
                gamma_f = st.number_input("Coeficiente de Majoração de Cargas (γf)", 1.0, value=1.4, step=0.1, key='gamma_f')
                q_servico_kn_m = (carga_area * larg_inf_total_m)
                q_serv_kn_cm = q_servico_kn_m / 100.0
                q_ult_kn_cm = q_serv_kn_cm * gamma_f
                p_load_ult = (p_load_serv[0] * gamma_f, p_load_serv[1]) if p_load_serv else None
                Msd, Vsd, detalhes_esforcos_internos, _ = calcular_esforcos_viga(tipo_viga, L_cm, q_ult_kn_cm, p_load_ult)
                detalhes_esforcos_memorial = {
                    'input_mode': input_mode, 'tipo_viga': tipo_viga,
                    'larg_esq_cm': larg_esq_cm, 'larg_dir_cm': larg_dir_cm, 'larg_inf_total_m': larg_inf_total_m,
                    'carga_area': carga_area, 'q_servico_kn_m': q_servico_kn_m, 'gamma_f': gamma_f,
                    'q_ult_kn_cm': q_ult_kn_cm, 'p_load_serv': p_load_serv, 'p_load_ult': p_load_ult,
                    'detalhes_esforcos': detalhes_esforcos_internos, 'Msd': Msd, 'Vsd': Vsd, 'L_cm': L_cm
                }
        else:
            with st.container(border=True):
                st.warning("No modo manual, a verificação de flecha (ELS) e o cálculo automático de Cb não são realizados.")
                msd_input = st.number_input("Momento Solicitante de Cálculo (Msd, kNm)", min_value=0.0, value=100.0, key='msd_input')
                Msd = msd_input * 100
                Vsd = st.number_input("Força Cortante Solicitante de Cálculo (Vsd, kN)", min_value=0.0, value=50.0, key='vsd_input')
            detalhes_esforcos_memorial = {
                'input_mode': input_mode, 'Msd': Msd, 'Vsd': Vsd, 'L_cm': L_cm
            }
        
        st.markdown("---")
        st.markdown("### 🔩 Parâmetros do Aço e Viga")
        E_aco_input = st.number_input("Módulo de Elasticidade (E, kN/cm²)", value=20000.0, step=100.0, key='E_aco_input')
        fy_aco = st.number_input("Tensão de Escoamento (fy, kN/cm²)", 20.0, 50.0, 34.5, 0.5, key='fy_aco')
        Lb_projeto = st.number_input(
            "Comprimento Destravado (Lb, cm)",
            min_value=1.0,
            max_value=L_cm,
            value=L_cm,
            step=1.0,
            key='Lb_projeto'
        )
        
        cb_modo_auto = st.checkbox("Calcular Cb automaticamente?", value=False, disabled=(input_mode == "Inserir Esforços Manualmente"))
        Cb_projeto = 0
        detalhes_cb_memorial = None
        
        if cb_modo_auto: # Se o usuário marcar a caixa
            Cb_projeto, detalhes_cb_memorial = calcular_cb(tipo_viga, L_cm, q_ult_kn_cm, p_load_ult)
            st.info(f"Fator Cb calculado: **{Cb_projeto:.2f}**")
        else: # Se a caixa não estiver marcada (padrão)
            Cb_projeto = st.number_input("Fator de Modificação (Cb)", 1.0, 3.0, 1.10, key='Cb_projeto')
            detalhes_cb_memorial = {'formula_simbolica': 'N/A', 'momentos': {}}
            if input_mode == "Inserir Esforços Manualmente":
                st.caption("Cálculo automático de Cb indisponível no modo manual.")
        
        with st.container(border=True):
            st.subheader("Enrijecedores de Alma")
            usa_enrijecedores = st.checkbox("Utilizar enrijecedores transversais?", key='usa_enrijecedores')
            a_enr = 0
            if usa_enrijecedores:
                a_enr = st.number_input("Distância entre enrijecedores (a, cm)", min_value=1.0, value=100.0, step=1.0, key='a_enr')

        st.markdown("---")
        st.markdown("### 📐 Verificação de Serviço (ELS)")
        limite_flecha_divisor = st.selectbox("Limite de Flecha (L/x)", (180, 250, 350, 500), index=2, key='limite_flecha_divisor')

    projeto_info = {'nome': projeto_nome, 'engenheiro': engenheiro, 'data': data_projeto.strftime('%d/%m/%Y'), 'revisao': revisao}
    input_params = {
        'tipo_viga': tipo_viga, 'L_cm': L_cm, 'input_mode': input_mode, 'Msd': Msd, 'Vsd': Vsd,
        'q_serv_kn_cm': q_serv_kn_cm, 'p_load_serv': p_load_serv, 'fy_aco': fy_aco,
        'Lb_projeto': Lb_projeto, 'Cb_projeto': Cb_projeto, 'detalhes_esforcos_memorial': detalhes_esforcos_memorial,
        'usa_enrijecedores': usa_enrijecedores, 'a_enr': a_enr, 'limite_flecha_divisor': limite_flecha_divisor,
        'projeto_info': projeto_info, 'cb_modo_auto': cb_modo_auto, 'detalhes_cb_memorial': detalhes_cb_memorial,
        'E_aco': E_aco_input # Novo parâmetro adicionado
    }
    
    if input_mode == "Calcular a partir de Cargas na Viga":
        input_params.update({
            'larg_esq_cm': larg_esq_cm, 'larg_dir_cm': larg_dir_cm, 'larg_inf_total_m': larg_inf_total_m,
            'carga_area': carga_area, 'q_servico_kn_m': q_servico_kn_m, 'gamma_f': gamma_f,
            'q_ult_kn_cm': q_ult_kn_cm, 'p_load_ult': p_load_ult,
        })
    else:
        input_params.update({'q_ult_kn_cm': 0, 'p_load_ult': None})

    create_metrics_dashboard(input_params)

    st.markdown("### 🎯 Modo de Análise")
    col1, col2 = st.columns(2)
    if col1.button("📊 Análise em Lote e Otimização", use_container_width=True):
        st.session_state.analysis_mode = "batch"
    if col2.button("📋 Memorial Detalhado de Perfil", use_container_width=True):
        st.session_state.analysis_mode = "detailed"

    if st.session_state.analysis_mode == "batch":
        st.header("📊 Análise em Lote")
        if st.button("🚀 Iniciar Análise Otimizada", type="primary", use_container_width=True):
            run_batch_analysis(all_sheets, input_params)
        
        if st.session_state.analysis_results is not None:
            df_all_results = st.session_state.analysis_results
            
            tabs = st.tabs([PROFILE_TYPE_MAP.get(name, name) for name in all_sheets.keys()])
            for i, sheet_name in enumerate(all_sheets.keys()):
                with tabs[i]:
                    df_type = df_all_results[df_all_results['Tipo'] == sheet_name].drop(columns=['Tipo'])
                    df_aprovados_cat = df_type[df_type['Status'] == 'APROVADO'].copy().sort_values(by='Peso (kg/m)')
                    df_reprovados_cat = df_type[df_type['Status'] == 'REPROVADO'].copy().sort_values(by='Peso (kg/m)')

                    # Botão de download para todos os resultados em uma aba
                    if not df_type.empty:
                        df_total = pd.concat([df_aprovados_cat, df_reprovados_cat])
                        excel_data = create_excel_with_colors([df_total], [f"{sheet_name}_Resultados"])
                        st.download_button(
                            label=f"📥 Baixar todos os resultados ({sheet_name}) em XLSX",
                            data=excel_data,
                            file_name=f"resultados_{sheet_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                    if not df_aprovados_cat.empty:
                        st.plotly_chart(create_top_profiles_chart(df_aprovados_cat), use_container_width=True)
                        with st.expander(f"Ver todos os {len(df_aprovados_cat)} perfis aprovados"):
                            st.dataframe(style_classic_dataframe(df_aprovados_cat), use_container_width=True)
                    else:
                        st.info("Nenhum perfil aprovado nesta categoria.")

                    if not df_reprovados_cat.empty:
                        with st.expander(f"Ver os {len(df_reprovados_cat)} perfis reprovados"):
                            st.dataframe(style_classic_dataframe(df_reprovados_cat), use_container_width=True)

    elif st.session_state.analysis_mode == "detailed":
        st.header("📋 Memorial Detalhado")
        display_names = [PROFILE_TYPE_MAP.get(name, name) for name in all_sheets.keys()]
        reverse_name_map = {v: k for k, v in PROFILE_TYPE_MAP.items()}

        col1, col2 = st.columns(2)
        selected_display_name = col1.selectbox("Selecione o Tipo de Perfil:", display_names)
        sheet_name = reverse_name_map.get(selected_display_name, selected_display_name)
        df_selecionado = all_sheets[sheet_name]
        perfil_selecionado_nome = col2.selectbox("Selecione o Perfil Específico:", df_selecionado['Bitola (mm x kg/m)'])

        if st.button("📄 Gerar Memorial Completo", type="primary", use_container_width=True):
            run_detailed_analysis(df_selecionado, perfil_selecionado_nome, selected_display_name, input_params)

        if st.session_state.detailed_analysis_html:
            with st.expander("📊 Resumo Visual da Análise", expanded=True):
                st.plotly_chart(st.session_state.profile_efficiency_chart, use_container_width=True)

            with st.expander("📄 Visualização do Memorial", expanded=True):
                st.components.v1.html(st.session_state.detailed_analysis_html, height=3000, width=2500, scrolling=True)
            
            st.download_button(
                label="📥 Baixar Memorial em HTML",
                data=st.session_state.detailed_analysis_html.encode('utf-8'),
                file_name=f"Memorial_{perfil_selecionado_nome.replace(' ', '_')}.html",
                mime="text/html",
                use_container_width=True
            )

def run_detailed_analysis(df, perfil_nome, perfil_tipo_display, input_params):
    with st.spinner(f"Gerando análise completa para {perfil_nome}..."):
        try:
            perfil_series = df[df['Bitola (mm x kg/m)'] == perfil_nome].iloc[0]
            props = get_profile_properties(perfil_series)

            tipo_fabricacao = "Soldado" if "Soldado" in perfil_tipo_display else "Laminado"

            esforcos_html = _render_esforcos_viga_section(input_params['detalhes_esforcos_memorial'])
            
            cb_calc_html = ""
            if input_params.get('cb_modo_auto'):
                cb_calc_html = _render_cb_calc_section(input_params['detalhes_cb_memorial'], input_params['Cb_projeto'], input_params['input_mode'])

            res_flt, res_flm, res_fla, res_cis, res_flecha, passo_a_passo = perform_all_checks(
                props=props, detalhado=True, tipo_fabricacao=tipo_fabricacao, **input_params
            )
            
            eficiencias = {
                "FLT": res_flt['eficiencia'],
                "FLM": res_flm['eficiencia'],
                "FLA": res_fla['eficiencia'],
                "Cisalhamento": res_cis['eficiencia'],
                "Flecha": res_flecha['eficiencia'],
            }
            st.session_state.profile_efficiency_chart = create_profile_efficiency_chart(perfil_nome, eficiencias)
            
            resumo_html = build_summary_html(input_params['Msd'], input_params['Vsd'], res_flt, res_flm, res_fla, res_cis, res_flecha)
            resultados = {'resumo_html': resumo_html, 'passo_a_passo_html': passo_a_passo, 'esforcos_html': esforcos_html, 'cb_calc_html': cb_calc_html}
            
            html_content = create_professional_memorial_html(
                perfil_nome, perfil_tipo_display, resultados,
                f"""
                <div style="text-align: left;">
                    <p><strong>Módulo de Elasticidade (E):</strong> {input_params['E_aco']:.2f} kN/cm²</p>
                    <p><strong>Tensão de Escoamento (fy):</strong> {input_params['fy_aco']:.2f} kN/cm²</p>
                    <p><strong>Altura total (d):</strong> {perfil_series.get('d (mm)'):.2f} mm</p>
                    <p><strong>Largura da Mesa (bf):</strong> {perfil_series.get('bf (mm)'):.2f} mm</p>
                    <p><strong>Espessura da Alma (tw):</strong> {perfil_series.get('tw (mm)'):.2f} mm</p>
                    <p><strong>Espessura da Mesa (tf):</strong> {perfil_series.get('tf (mm)'):.2f} mm</p>
                    <p><strong>Altura da Alma (h):</strong> {perfil_series.get('h (mm)'):.2f} mm</p>
                    <p><strong>Área (A):</strong> {perfil_series.get('Área (cm2)'):.2f} cm²</p>
                    <p><strong>Inércia Ix:</strong> {perfil_series.get('Ix (cm4)'):.2f} cm⁴</p>
                    <p><strong>Módulo de Seção Elástico (Wx):</strong> {perfil_series.get('Wx (cm3)'):.2f} cm³</p>
                    <p><strong>Raio de Giração (rx):</strong> {props.get('rx'):.2f} cm</p>
                    <p><strong>Módulo de Seção Plástico (Zx):</strong> {props.get('Zx'):.2f} cm³</p>
                    <p><strong>Inércia Iy:</strong> {props.get('Iy'):.2f} cm⁴</p>
                    <p><strong>Raio de Giração (ry):</strong> {props.get('ry'):.2f} cm</p>
                    <p><strong>Constante de Torção (J):</strong> {props.get('J'):.2f} cm⁴</p>
                    <p><strong>Constante de Empenamento (Cw):</strong> {props.get('Cw'):.2f} cm⁶</p>
                </div>
                """, input_params['projeto_info']
            )
            st.session_state.detailed_analysis_html = html_content
        except Exception as e:
            st.error(f"❌ Ocorreu um erro: {e}")

def run_batch_analysis(all_sheets, input_params):
    all_results = []
    progress_bar = st.progress(0, text="Analisando perfis...")
    total_perfis = sum(len(df) for df in all_sheets.values())
    perfis_processados = 0
    
    for sheet_name, df in all_sheets.items():
        display_name = PROFILE_TYPE_MAP.get(sheet_name, sheet_name)
        tipo_fabricacao_auto = "Soldado" if "Soldado" in display_name else "Laminado"
        
        for _, row in df.iterrows():
            perfis_processados += 1
            progress_bar.progress(perfis_processados / total_perfis, text=f"Analisando: {row['Bitola (mm x kg/m)']}")
            try:
                props = get_profile_properties(row)
                res_flt, res_flm, res_fla, res_cis, res_flecha, _ = perform_all_checks(
                    props=props, tipo_fabricacao=tipo_fabricacao_auto, **input_params
                )
                
                status_geral = "APROVADO"
                if any(res['eficiencia'] > 100.1 for res in [res_flt, res_flm, res_fla, res_cis, res_flecha]):
                    status_geral = "REPROVADO"
                
                all_results.append({
                    'Tipo': sheet_name, 'Perfil': row['Bitola (mm x kg/m)'],
                    'Peso (kg/m)': props.get('Peso', 0), 'Status': status_geral,
                    'Ef. FLT (%)': res_flt['eficiencia'], 'Ef. FLM (%)': res_flm['eficiencia'],
                    'Ef. FLA (%)': res_fla['eficiencia'], 'Ef. Cisalhamento (%)': res_cis['eficiencia'],
                    'Ef. Flecha (%)': res_flecha['eficiencia']
                })
            except (ValueError, KeyError):
                continue
    progress_bar.empty()
    st.session_state.analysis_results = pd.DataFrame(all_results) if all_results else pd.DataFrame()

if __name__ == '__main__':
    main()
