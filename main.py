import streamlit as st
import pandas as pd
import math
import plotly.graph_objects as go
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import PatternFill
import base64
import pytz
from pathlib import Path
from calculos_nbr8800_2024 import (
    ERRATA,
    NORMA,
    analyze_beam,
    calculate_cb as calculate_cb_nbr2024,
    combine_els,
    combine_elu_normal,
    deflection_limit,
    flexural_strength_i,
    local_compression_strength,
    overall_status,
    shear_strength_i,
    validate_material,
)
from memorial_nbr8800_2024 import build_memorial_details
# ==============================================================================
# 1. CONFIGURAÇÕES E CONSTANTES GLOBAIS APRIMORADAS
# ==============================================================================

st.set_page_config(
    page_title="🏗️ Calculadora Estrutural - Perfis Metálicos",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.abnt.org.br',
        'Report a bug': None,
        'About': f"# Calculadora Estrutural\nCálculos baseados na {NORMA} + Errata 1:2025"
    }
)

class Config:
    NOME_NORMA = f'{NORMA} + Errata 1:2025'
    GAMMA_A1 = 1.10
    FATOR_SIGMA_R = 0.3
    FATOR_LAMBDA_P_FLT = 1.76
    FATOR_LAMBDA_P_FLM = 0.38
    FATOR_LAMBDA_R_FLM_LAMINADO = 0.83
    FATOR_LAMBDA_R_FLM_SOLDADO = 0.95
    FATOR_LAMBDA_P_FLA = 3.76
    FATOR_LAMBDA_R_FLA = 5.70
    KV_ALMA_SEM_ENRIJECEDORES = 5.34
    FATOR_VP = 0.60
    FATOR_LAMBDA_P_VRD = 1.10
    FATOR_LAMBDA_R_VRD = 1.37
    FATOR_VRD_ELASTICO = 1.24

PROFILE_TYPE_MAP = {
    "Laminados": "Perfis Laminados",
    "CS": "Perfis Compactos Soldados",
    "CVS": "Perfis CVS Soldados",
    "VS": "Perfis Soldados"
}

PROFILE_FABRICATION_MAP = {
    "Laminados": "Laminado",
    "CS": "Soldado",
    "CVS": "Soldado",
    "VS": "Soldado",
}


HTML_TEMPLATE_CSS_PRO = """
<style>
    /* Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Poppins:wght@600;700;800&display=swap');
    @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500&display=swap');

    /* --- Paleta de Cores Definitiva HQ Engenharia --- */
    :root {
        --background: #0f172a;
        --surface: #1e2b3b;
        --border: #334155;
        --text-display: #ffffff;
        --text-primary: #e2e8f0;
        --text-secondary: #94a3b8;
        --accent-gold: #fbbd24;
        --accent-amber: #f59e0b;
        --button-text: #1e293b;
    }

    /* --- Base e Overrides Globais do Streamlit --- */
    body {
        font-family: 'Inter', sans-serif;
        background-color: var(--background);
        color: var(--text-primary);
    }
    .stApp {
        background-color: var(--background);
    }
    .block-container {
        padding: 2rem 3rem 3rem 3rem;
    }

    /* --- Títulos e Textos Genéricos --- */
    h1, h2 {
        font-family: 'Poppins', sans-serif;
        color: var(--text-display);
        font-weight: 700;
    }
    /* Força a cor dos títulos de seção (H3) para dourado */
    h3 {
        font-family: 'Poppins', sans-serif;
        color: var(--accent-gold) !important;
        font-weight: 700;
    }
    /* Garante que os títulos H3 dentro do markdown também sejam dourados */
    [data-testid="stMarkdownContainer"] h3 {
        color: var(--accent-gold) !important;
    }
    /* Estiliza o texto "Fator Cb..." para ser branco e legível */
    .metric-footer {
        color: var(--text-primary) !important;
        font-family: 'Inter', sans-serif;
        font-size: 0.9rem;
        text-align: center;
        margin-top: -0.5rem;
        margin-bottom: 2rem;
    }

    /* --- BARRA LATERAL (SIDEBAR) --- */
    [data-testid="stSidebar"] {
        background-color: var(--surface) !important;
        border-right: 1px solid var(--border);
    }
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        color: var(--accent-gold);
    }
    
    /* RÓTULOS (LABELS) */
    label, [data-testid="stMetricLabel"] {
        color: var(--text-secondary) !important;
        font-weight: 600;
        font-family: 'Inter', sans-serif;
    }
    .stRadio > label, .stCheckbox > label {
        color: var(--text-secondary) !important;
        font-weight: 600;
        margin-bottom: 0.5rem;
    }
    .stRadio [data-testid="stMarkdownContainer"] p, .stCheckbox [data-testid="stMarkdownContainer"] p {
        color: var(--text-primary);
    }

    /* INPUTS, SELECTBOX, TEXTAREA */
    input, 
    div[data-baseweb="select"] > div,
    textarea {
        background-color: var(--background) !important;
        color: var(--text-primary) !important;
        border: 1px solid var(--border) !important;
        border-radius: 8px !important;
    }
    .stDateInput input {
        color: var(--text-primary) !important;
    }
    input:focus, 
    div[data-baseweb="select"] > div:focus-within,
    textarea:focus {
        border-color: var(--accent-gold) !important;
        box-shadow: 0 0 0 2px rgba(251, 191, 36, 0.3) !important;
    }
    
    .stSelectbox div[data-baseweb="select"] div,
    .stSelectbox div[data-baseweb="select"] span {
        color: var(--text-primary) !important;
    }
    
    /* --- ST.EXPANDER --- */
    [data-testid="stExpander"] {
        border: none !important;
        box-shadow: none !important;
        background: none !important;
    }
    [data-testid="stExpander"] summary {
        background-color: var(--surface) !important;
        color: var(--text-primary) !important;
        border: 1px solid var(--border) !important;
        border-radius: 8px !important;
        margin-bottom: 0.5rem;
        padding: 0.75rem 1rem !important;
    }
    [data-testid="stExpander"] summary p {
        color: var(--text-primary) !important;
        font-weight: 600;
        font-size: 1.1rem;
    }
    [data-testid="stExpander"] summary svg {
        fill: var(--text-primary) !important;
    }
    [data-testid="stExpander"] .st-emotion-cache-16txtl3 {
        background-color: var(--surface) !important;
        border: 1px solid var(--border) !important;
        border-top: none !important;
        border-radius: 0 0 8px 8px !important;
        padding: 1rem !important;
        margin-top: -0.5rem;
    }

    /* --- Cabeçalho e Métricas --- */
    .pro-header {
        background: var(--surface); padding: 2.5rem; border-radius: 12px;
        border: 1px solid var(--border); text-align: center; margin-bottom: 2rem;
        box-shadow: 0 8px 30px rgba(0,0,0,0.3);
    }
    .pro-header h1 { font-size: 2.8rem; }
    .pro-header p { color: var(--text-secondary); }
    .gradient-text {
        background: linear-gradient(135deg, #FBBF24 0%, #FDE68A 50%, #D4AF37 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text;
    }
    /* --- Painel responsivo de parâmetros do projeto --- */
    .project-metrics-shell {
        container-type: inline-size;
        margin: .4rem 0 1.8rem;
    }
    .project-metrics-grid {
        display: grid;
        grid-template-columns: repeat(12, minmax(0, 1fr));
        gap: .85rem;
    }
    .project-metric-card {
        grid-column: span 3;
        min-width: 0;
        min-height: 108px;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        background: linear-gradient(145deg, #1e2b3b 0%, #172438 100%);
        border: 1px solid var(--border);
        border-top: 3px solid var(--accent-amber);
        border-radius: 10px;
        padding: .9rem 1rem;
        box-shadow: 0 8px 20px rgba(2, 6, 23, .18);
    }
    .project-metric-label {
        color: var(--text-secondary);
        font-family: 'Inter', sans-serif;
        font-size: .72rem;
        font-weight: 700;
        letter-spacing: .075em;
        line-height: 1.25;
        text-transform: uppercase;
    }
    .project-metric-value {
        color: var(--text-display);
        font-family: 'Poppins', sans-serif;
        font-size: clamp(1.2rem, 2vw, 1.6rem);
        font-weight: 700;
        line-height: 1.15;
        overflow-wrap: anywhere;
        font-variant-numeric: tabular-nums;
    }
    .project-metric-card:nth-child(n + 5) { grid-column: span 4; }
    .project-metric-unit {
        min-height: 1.15em;
        color: var(--text-secondary);
        font-size: .76rem;
        line-height: 1.2;
    }
    .project-metrics-context {
        display: flex;
        flex-wrap: wrap;
        gap: .55rem;
        margin-top: .85rem;
    }
    .project-metrics-context span {
        flex: 1 1 180px;
        color: var(--text-primary);
        background: rgba(30, 43, 59, .72);
        border: 1px solid var(--border);
        border-radius: 999px;
        padding: .48rem .8rem;
        font-size: .78rem;
        line-height: 1.25;
        text-align: center;
    }
    .project-metrics-context strong {
        color: var(--accent-gold);
        margin-right: .3rem;
    }
    @container (max-width: 760px) {
        .project-metrics-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
        .project-metric-card,
        .project-metric-card:nth-child(n + 5) { grid-column: span 1; }
        .project-metric-card:last-child { grid-column: 1 / -1; }
    }
    @container (max-width: 390px) {
        .project-metrics-grid { grid-template-columns: 1fr; }
        .project-metrics-context span { flex-basis: 100%; }
    }

    /* --- Botões --- */
    .stButton > button, .stDownloadButton > button {
        padding: 12px; border-radius: 8px; font-weight: bold; font-size: 1rem;
        font-family: 'Poppins', sans-serif; border: 2px solid var(--accent-gold);
        transition: all 0.2s ease-in-out;
    }
    .stButton > button[kind="primary"], .stDownloadButton > button {
        background-color: var(--accent-gold); color: var(--button-text) !important;
    }
    .stButton > button[kind="primary"]:hover, .stDownloadButton > button:hover {
        background-color: #ffd042; border-color: #ffd042; transform: translateY(-2px);
    }
    .stButton > button[kind="secondary"] {
        background-color: transparent; color: var(--accent-gold) !important;
    }
    .stButton > button[kind="secondary"]:hover {
        background-color: var(--accent-gold); color: var(--button-text) !important;
    }

    /* --- Abas (Tabs) --- */
    .stTabs [data-baseweb="tab-list"] button {
        color: var(--text-secondary); padding: 1rem;
    }
    .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] {
        color: var(--accent-gold); border-bottom-color: var(--accent-gold);
    }

    /* --- Estilos para o MEMORIAL HTML --- */
    .container { font-family: 'Inter', sans-serif; color: var(--text-primary); }
    .container .pro-header { background: var(--surface); }
    .container h1, .container h2, .container h3, .container h4 { color: var(--text-display); font-family: 'Poppins', sans-serif; }
    .container h2 { border-bottom: 1px solid var(--border); padding-bottom: 10px; }
    .container h3 { color: var(--accent-gold); }
    .container strong { color: var(--text-display); }
    .container .info-card { background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 1.5rem; }
    .container table { width: 100%; border-collapse: collapse; margin-top: 1rem; font-family: 'JetBrains Mono', monospace; }
    .container th, .container td { text-align: left; padding: 0.75rem; border-bottom: 1px solid var(--border); }
    .container th { color: var(--accent-gold); }
    .container td.fail, .container span.fail { color: #FF6B6B !important; font-weight: bold; }
    .container td.pass, .container span.pass { color: #63E6BE !important; font-weight: bold; }
    .container .formula-block { background-color: var(--background); border-left: 4px solid var(--accent-amber); border-radius: 4px; padding: 1px 15px 15px; margin: 15px 0;}

    /* Memorial auditável — hierarquia visual para fórmula, substituição e decisão */
    .container { max-width: 1180px; margin: 0 auto; line-height: 1.62; }
    .container .chapter-intro { color: var(--text-secondary); margin: -0.45rem 0 1.35rem; font-size: 0.98rem; }
    .container .audit-banner {
        display: flex; align-items: center; justify-content: space-between; gap: 1rem;
        background: linear-gradient(135deg, rgba(245,158,11,.18), rgba(251,191,36,.05));
        border: 1px solid rgba(251,191,36,.45); border-radius: 10px; padding: 1rem 1.25rem; margin: 1.5rem 0 2rem;
    }
    .container .audit-banner strong { color: var(--accent-gold); letter-spacing: .09em; font-size: .82rem; }
    .container .audit-banner span { color: var(--text-secondary); font-size: .9rem; }
    .container .calc-step {
        position: relative; background: #111c2e; border: 1px solid var(--border); border-radius: 12px;
        padding: 1.3rem 1.4rem 1.15rem; margin: 0 0 1.2rem; box-shadow: 0 6px 18px rgba(0,0,0,.16);
        break-inside: avoid-page;
    }
    .container .calc-step-head { display: flex; align-items: center; gap: .75rem; margin-bottom: .65rem; }
    .container .calc-step-head h4 { margin: 0; font-size: 1.08rem; }
    .container .calc-step-number {
        display: inline-flex; align-items: center; justify-content: center; min-width: 2.6rem; height: 2rem;
        padding: 0 .45rem; border-radius: 999px; background: var(--accent-gold); color: #172033;
        font: 700 .78rem 'JetBrains Mono', monospace;
    }
    .container .calc-explanation { margin: .3rem 0 1rem; color: var(--text-secondary); }
    .container .equation-label { margin: .8rem 0 .3rem; color: #fcd978; font-size: .72rem; font-weight: 700; letter-spacing: .09em; text-transform: uppercase; }
    .container .formula-symbolic, .container .formula-numeric {
        overflow-x: auto; background: #0b1322; border: 1px solid #27364a; border-radius: 7px;
        padding: .65rem 1rem; margin: .25rem 0; color: #f8fafc;
    }
    .container .formula-numeric { background: #0d1828; border-left: 3px solid #60a5fa; }
    .container .equation-heading { display: flex; align-items: center; margin: 1rem 0 .25rem; }
    .container .equation-heading h5 { margin: 0; color: #f8fafc; font-size: .94rem; }
    .container .equation-caption { color: #8fa4bf; font-size: .76rem; margin: 0 0 .45rem; }
    .container .equation-caption strong { color: var(--accent-gold); padding: 0 .18rem; }
    .container .formula-chain {
        overflow-x: auto; background: #0b1322; border: 1px solid #31445e;
        border-left: 4px solid #60a5fa; border-radius: 8px; padding: 1rem 1.15rem;
        margin: .25rem 0; color: #f8fafc;
    }
    .container .verification-chain {
        overflow-x: auto; background: #0b1322; border: 1px solid #31445e;
        border-left: 4px solid #94a3b8; border-radius: 8px; padding: .8rem 1rem;
        margin: .45rem 0 0; color: #f8fafc;
    }
    .container .verification-card.pass .verification-chain { border-left-color: #34d399; }
    .container .verification-card.fail .verification-chain { border-left-color: #fb7185; }
    .container .step-theory-panel {
        background: #0d1828; border: 1px solid #2e4159; border-radius: 9px;
        margin: .85rem 0 1rem; overflow: hidden;
    }
    .container .step-theory-panel summary {
        display: flex; align-items: center; gap: .7rem; cursor: pointer; list-style: none;
        padding: .8rem .95rem; background: #132137; color: var(--text-display);
    }
    .container .step-theory-panel summary::-webkit-details-marker { display: none; }
    .container .step-theory-panel summary::before {
        content: '+'; display: inline-flex; align-items: center; justify-content: center;
        width: 1.45rem; height: 1.45rem; flex: 0 0 auto; border-radius: 50%;
        background: rgba(96,165,250,.14); color: #93c5fd; font-weight: 700;
    }
    .container .step-theory-panel[open] summary::before { content: '−'; }
    .container .step-theory-panel summary span {
        color: #93c5fd; font-size: .68rem; font-weight: 700; letter-spacing: .08em; text-transform: uppercase;
    }
    .container .step-theory-panel summary strong { margin-left: auto; text-align: right; font-size: .84rem; }
    .container .step-theory-content { padding: .95rem 1rem 1rem; }
    .container .step-theory-content section + section { margin-top: .9rem; padding-top: .85rem; border-top: 1px solid #25374e; }
    .container .step-theory-content h5 { margin: 0 0 .35rem; color: #fcd978; font-size: .86rem; }
    .container .step-theory-content p { margin: 0; color: var(--text-secondary); font-size: .87rem; line-height: 1.7; }
    .container .calc-result {
        margin-top: .9rem; padding: .8rem 1rem; border-radius: 7px;
        background: rgba(96,165,250,.09); border: 1px solid rgba(96,165,250,.26); color: #dbeafe;
    }
    .container .calc-decision { margin-top: .7rem; padding: .65rem .85rem; border-left: 3px solid var(--accent-gold); background: rgba(251,191,36,.07); color: #fde7a7; }
    .container .norm-ref { color: #7f91aa; font-size: .78rem; margin-top: .8rem; }
    .container .theory-panel {
        background: #111c2e; border: 1px solid #3d4f67; border-radius: 12px;
        margin: 0 0 1.35rem; overflow: hidden; break-inside: avoid-page;
    }
    .container .theory-panel summary {
        display: flex; align-items: center; justify-content: space-between; gap: 1rem;
        cursor: pointer; list-style: none; padding: 1rem 1.2rem; background: #162338;
        color: var(--text-display); font-weight: 700;
    }
    .container .theory-panel summary::-webkit-details-marker { display: none; }
    .container .theory-panel summary::before {
        content: '+'; display: inline-flex; align-items: center; justify-content: center;
        width: 1.65rem; height: 1.65rem; flex: 0 0 auto; border-radius: 50%;
        background: rgba(251,191,36,.16); color: var(--accent-gold); font-size: 1.15rem;
    }
    .container .theory-panel[open] summary::before { content: '−'; }
    .container .theory-panel summary span { color: var(--accent-gold); font-size: .72rem; letter-spacing: .08em; text-transform: uppercase; }
    .container .theory-panel summary strong { margin-left: auto; text-align: right; }
    .container .theory-content { padding: 1.2rem 1.3rem 1.35rem; }
    .container .theory-content > h4 { margin: 1.1rem 0 .55rem; color: #f8fafc; }
    .container .theory-content > h4:first-child { margin-top: 0; }
    .container .variable-table-wrap { overflow-x: auto; border: 1px solid #2c3b50; border-radius: 8px; }
    .container .variable-table { margin: 0; min-width: 650px; }
    .container .variable-table th, .container .variable-table td { padding: .65rem .8rem; vertical-align: top; }
    .container .variable-table td:first-child { color: #bfdbfe; white-space: nowrap; }
    .container .theory-grid { display: grid; grid-template-columns: repeat(2,minmax(0,1fr)); gap: .75rem; }
    .container .theory-concept { background: #0d1828; border: 1px solid #293a50; border-radius: 8px; padding: .8rem .9rem; }
    .container .theory-concept h5 { color: #fcd978; margin: 0 0 .35rem; font-size: .88rem; }
    .container .theory-concept p { color: var(--text-secondary); margin: 0; font-size: .85rem; }
    .container .verification-card {
        border: 1px solid var(--border); border-left: 6px solid #94a3b8; border-radius: 10px;
        background: #111c2e; padding: 1.15rem 1.35rem; margin: 1.2rem 0 1.8rem; break-inside: avoid-page;
    }
    .container .verification-card.pass { border-left-color: #34d399; }
    .container .verification-card.fail { border-left-color: #fb7185; }
    .container .verification-card.pending { border-left-color: #fbbf24; }
    .container .verification-card h4 { margin: .2rem 0 .5rem; }
    .container .verification-kicker { color: var(--text-secondary); font-size: .7rem; font-weight: 700; letter-spacing: .12em; }
    .container .verification-metrics { display: flex; justify-content: space-between; align-items: center; margin-top: .75rem; }
    .container .pending { color: #fbbf24 !important; font-weight: 700; }
    .container .scope-grid { display: grid; grid-template-columns: repeat(2, minmax(0,1fr)); gap: 1rem; }
    .container .global-status { display: flex; justify-content: space-between; align-items: center; border-radius: 10px; padding: 1.1rem 1.3rem; margin: 1.2rem 0; border: 1px solid var(--border); background: #111c2e; }
    .container .global-status span { color: var(--text-secondary); font-size: .75rem; letter-spacing: .12em; }
    .container .global-status.pass strong { color: #63e6be; }
    .container .global-status.fail strong { color: #ff6b6b; }
    .container .global-status.pending strong { color: #fbbf24; }
    .container .notice { background: #162236; border: 1px solid var(--border); border-radius: 8px; padding: .9rem 1rem; margin: 1rem 0; color: var(--text-secondary); }
    @media (max-width: 780px) {
        .container .scope-grid, .container .theory-grid { grid-template-columns: 1fr; }
        .container .audit-banner, .container .verification-metrics { align-items: flex-start; flex-direction: column; }
        .container .theory-panel summary { align-items: flex-start; flex-wrap: wrap; }
        .container .theory-panel summary strong { margin-left: 0; text-align: left; width: calc(100% - 2.5rem); }
        .container .step-theory-panel summary { align-items: flex-start; flex-wrap: wrap; }
        .container .step-theory-panel summary strong { margin-left: 2.15rem; text-align: left; width: 100%; }
    }
    @media print {
        body { background: white !important; color: #172033 !important; }
        .container { color: #172033 !important; }
        .container .calc-step, .container .verification-card, .container .info-card, .container .notice, .container .theory-panel, .container .theory-panel summary, .container .theory-concept, .container .step-theory-panel, .container .step-theory-panel summary { background: white !important; color: #172033 !important; box-shadow: none; border-color: #cbd5e1; }
        .container .formula-symbolic, .container .formula-numeric, .container .formula-chain, .container .verification-chain { background: #f8fafc !important; color: #0f172a !important; border-color: #cbd5e1; }
        .container .calc-explanation, .container .chapter-intro, .container .norm-ref { color: #475569 !important; }
    }

</style>
"""




def create_navigation_buttons():
    """Cria os botões de navegação para voltar ao site principal ou à página de ferramentas."""
    
    # URLs completas para os links
    ferramentas_url = "https://hqengenharia.eng.br/ferramentas.html"
    site_url = "https://hqengenharia.eng.br/index.html"
    
    # Estilos CSS para os botões
    button_style = """
        display: inline-flex;
        align-items: center;
        padding: 0.6rem 1.2rem;
        border-radius: 0.375rem; /* rounded-md */
        font-weight: bold;
        text-decoration: none;
        transition: all 0.2s ease-in-out;
        margin-left: 1rem;
        font-family: 'Poppins', sans-serif;
        font-size: 0.9rem;
    """
    
    secondary_style = "border: 2px solid #fbbd24; color: #fbbd24;"
    secondary_hover = ":hover { background-color: #fbbd24; color: #1e2b3b; }"
    
    primary_style = "background-color: #fbbd24; color: #1e2b3b; border: 2px solid #fbbd24;"
    primary_hover = ":hover { background-color: #f59e0b; border-color: #f59e0b; }"

    # HTML com os botões
    st.markdown(f"""
        <style>
            .nav-button-secondary {{ {button_style} {secondary_style} }}
            .nav-button-secondary:hover {{ {secondary_hover} }}
            .nav-button-primary {{ {button_style} {primary_style} }}
            .nav-button-primary:hover {{ {primary_hover} }}
        </style>
        
        <div style="display: flex; justify-content: flex-end; width: 100%; align-items: center;">
            <a href="{ferramentas_url}" target="_self" class="nav-button-secondary">
                Ferramentas
            </a>
            <a href="{site_url}" target="_self" class="nav-button-primary">
                Voltar ao Site
            </a>
        </div>
        <hr style="border-color: var(--border); margin-top: 1.5rem; margin-bottom: 1.5rem;">
    """, unsafe_allow_html=True)



# ==============================================================================
# 2. FUNÇÕES DE CÁLCULO E UTILITÁRIAS
# ==============================================================================

@st.cache_data
def load_data_from_local_file():
    """Carrega os dados da planilha de perfis."""
    try:
        caminho_arquivo_excel = Path(__file__).resolve().with_name('perfis.xlsx')
        return pd.read_excel(caminho_arquivo_excel, sheet_name=None)
    except FileNotFoundError:
        st.error(f"Erro: Arquivo '{caminho_arquivo_excel}' não foi encontrado. Verifique se ele está na mesma pasta que o seu script Python.")
        return None
    except Exception as e:
        st.error(f"Erro ao ler o arquivo Excel: {e}")
        return None

def calcular_esforcos_viga(tipo_viga, L_cm, q_kn_cm=0, p_load=None):
    msd_q, vsd_q, msd_p, vsd_p = 0, 0, 0, 0
    L = L_cm
    detalhes_esforcos = {
        'Msd_q': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN.cm'},
        'Vsd_q': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN'},
        'Msd_p': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN.cm'},
        'Vsd_p': {'valor': 0, 'formula_simbolica': "", 'unidade': 'kN'}
    }

    detalhes_carga_memorial = {'q_ult': {'valor': 0, 'unidade': 'kN/cm'}, 'p_ult': {'valor': 0, 'unidade': 'kN'}}

    if q_kn_cm > 0:
        detalhes_carga_memorial['q_ult']['valor'] = q_kn_cm
        if tipo_viga == 'Bi-apoiada':
            msd_q = (q_kn_cm * L**2) / 8
            vsd_q = (q_kn_cm * L) / 2
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{8}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = \\frac{q_{ultima} \\times L}{2}"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            msd_q = (q_kn_cm * L**2) / 2
            vsd_q = q_kn_cm * L
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{2}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = q_{ultima} \\times L"
        elif tipo_viga == 'Bi-engastada':
            msd_q = (q_kn_cm * L**2) / 12
            vsd_q = (q_kn_cm * L) / 2
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{12}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = \\frac{q_{ultima} \\times L}{2}"
        elif tipo_viga == 'Engastada e Apoiada':
            msd_q = (q_kn_cm * L**2) / 8
            vsd_q = (5 * q_kn_cm * L) / 8
            detalhes_esforcos['Msd_q']['formula_simbolica'] = "M_{sd, q} = \\frac{q_{ultima} \\times L^2}{8}"
            detalhes_esforcos['Vsd_q']['formula_simbolica'] = "V_{sd, q} = \\frac{5 \\times q_{ultima} \\times L}{8}"

    if p_load:
        P, x = p_load
        detalhes_carga_memorial['p_ult']['valor'] = P
        P_kn = P
        L_m = L / 100.0
        x_m = x / 100.0
        a_val = x_m
        b_val = L_m - a_val
        if tipo_viga == 'Bi-apoiada':
            msd_p = (P_kn * a_val * b_val) / L_m
            vsd_p = max((P_kn * b_val) / L_m, (P_kn * a_val) / L_m)
            detalhes_esforcos['Msd_p']['formula_simbolica'] = "M_{sd, P} = \\frac{P_{ultima} \\times a \\times b}{L}"
            detalhes_esforcos['Vsd_p']['formula_simbolica'] = "V_{sd, P} = \\max(\\frac{P_{ultima} \\times b}{L}, \\frac{P_{ultima} \\times a}{L})"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            msd_p = P_kn * a_val
            vsd_p = P_kn
            detalhes_esforcos['Msd_p']['formula_simbolica'] = "M_{sd, P} = P_{ultima} \\times a"
            detalhes_esforcos['Vsd_p']['formula_simbolica'] = "V_{sd, P} = P_{ultima}"
        msd_p *= 100

    detalhes_esforcos['Msd_q']['valor'] = msd_q
    detalhes_esforcos['Vsd_q']['valor'] = vsd_q
    detalhes_esforcos['Msd_p']['valor'] = msd_p
    detalhes_esforcos['Vsd_p']['valor'] = vsd_p

    msd_total = msd_q + msd_p
    vsd_total = vsd_q + vsd_p

    return msd_total, vsd_total, detalhes_esforcos, detalhes_carga_memorial

def calcular_cb(tipo_viga, L_cm, q_kn_cm=0, p_load=None):
    L_m = L_cm / 100
    q_ult_kn_m = q_kn_cm * 100

    detalhes_cb = {
        'formula_simbolica': 'C_b = \\frac{12,5 M_{máx}}{|2,5 M_{máx} + 3 M_A + 4 M_B + 3 M_C|}',
        'momentos': {
            'M_max': {'final_value': 0, 'components': []},
            'M_A': {'final_value': 0, 'components': []},
            'M_B': {'final_value': 0, 'components': []},
            'M_C': {'final_value': 0, 'components': []}
        }
    }

    momentos_q = {'M_max': 0, 'M_A': 0, 'M_B': 0, 'M_C': 0}
    momentos_p = {'M_max': 0, 'M_A': 0, 'M_B': 0, 'M_C': 0}

    # --- Cálculo devido à Carga Distribuída (q) ---
    if q_kn_cm > 0:
        if tipo_viga == 'Bi-apoiada':
            momentos_q['M_max'] = (q_ult_kn_m * L_m**2) / 8
            momentos_q['M_A'] = (3 * q_ult_kn_m * L_m**2) / 32
            momentos_q['M_B'] = momentos_q['M_max']
            momentos_q['M_C'] = momentos_q['M_A']

            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de q', 'formula': 'M_{max,q} = \\frac{q_{ultima} \\times L^2}{8}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de q', 'formula': 'M_{A,q} = \\frac{3 \\times q_{ultima} \\times L^2}{32}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de q', 'formula': 'M_{B,q} = M_{max,q}', 'valores': {}, 'valor': momentos_q['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de q', 'formula': 'M_{C,q} = M_{A,q}', 'valores': {}, 'valor': momentos_q['M_C']})

        elif tipo_viga == 'Engastada e Livre (Balanço)':
            momentos_q['M_max'] = (q_ult_kn_m * L_m**2) / 2
            momentos_q['M_A'] = (q_ult_kn_m * (0.75 * L_m)**2) / 2
            momentos_q['M_B'] = (q_ult_kn_m * (0.5 * L_m)**2) / 2
            momentos_q['M_C'] = (q_ult_kn_m * (0.25 * L_m)**2) / 2

            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de q', 'formula': 'M_{max,q} = \\frac{q_{ultima} \\times L^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de q', 'formula': 'M_{A,q} = \\frac{q_{ultima} \\times (0,75L)^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de q', 'formula': 'M_{B,q} = \\frac{q_{ultima} \\times (0,5L)^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de q', 'formula': 'M_{C,q} = \\frac{q_{ultima} \\times (0,25L)^2}{2}', 'valores': {'q_{ultima}': q_ult_kn_m, 'L': L_m}, 'valor': momentos_q['M_C']})
        else:
              return 1.0, {'formula_simbolica': 'N/A', 'momentos': {}}

    # --- Cálculo devido à Carga Pontual (P) ---
    if p_load:
        P_kn, x_cm = p_load
        x_m = x_cm / 100.0
        
        if tipo_viga == 'Bi-apoiada':
            a, b = x_m, L_m - x_m
            momentos_p['M_max'] = (P_kn * a * b) / L_m if L_m > 0 else 0
            momentos_p['M_A'] = (P_kn * b * (0.25*L_m)) / L_m if x_m > 0.25*L_m else (P_kn * a * (L_m - 0.25*L_m)) / L_m
            momentos_p['M_B'] = (P_kn * b * (0.5*L_m)) / L_m if x_m > 0.5*L_m else (P_kn * a * (L_m - 0.5*L_m)) / L_m
            momentos_p['M_C'] = (P_kn * b * (0.75*L_m)) / L_m if x_m > 0.75*L_m else (P_kn * a * (L_m - 0.75*L_m)) / L_m
            
            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de P', 'formula': 'M_{max,P} = \\frac{P_{ultima} \\times a \\times b}{L}', 'valores': {'P_{ultima}': P_kn, 'a': a, 'b': b, 'L': L_m}, 'valor': momentos_p['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de P', 'formula': 'M_{A,P} \\text{ (calculado em L/4)}', 'valores': {}, 'valor': momentos_p['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de P', 'formula': 'M_{B,P} \\text{ (calculado em L/2)}', 'valores': {}, 'valor': momentos_p['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de P', 'formula': 'M_{C,P} \\text{ (calculado em 3L/4)}', 'valores': {}, 'valor': momentos_p['M_C']})

        elif tipo_viga == 'Engastada e Livre (Balanço)':
            momentos_p['M_max'] = P_kn * x_m
            momentos_p['M_A'] = P_kn * x_m if x_m < (0.25 * L_m) else P_kn * (x_m - 0.25 * L_m)
            momentos_p['M_B'] = P_kn * x_m if x_m < (0.5 * L_m) else P_kn * (x_m - 0.5 * L_m)
            momentos_p['M_C'] = P_kn * x_m if x_m < (0.75 * L_m) else P_kn * (x_m - 0.75 * L_m)
            detalhes_cb['momentos']['M_max']['components'].append({'desc': 'Componente de P', 'formula': 'M_{max,P} = P_{ultima} \\times x', 'valores': {'P_{ultima}': P_kn, 'x': x_m}, 'valor': momentos_p['M_max']})
            detalhes_cb['momentos']['M_A']['components'].append({'desc': 'Componente de P', 'formula': 'M_{A,P} = P_{ultima} \\times (x-0.25L)', 'valores': {}, 'valor': momentos_p['M_A']})
            detalhes_cb['momentos']['M_B']['components'].append({'desc': 'Componente de P', 'formula': 'M_{B,P} = P_{ultima} \\times (x-0.5L)', 'valores': {}, 'valor': momentos_p['M_B']})
            detalhes_cb['momentos']['M_C']['components'].append({'desc': 'Componente de P', 'formula': 'M_{C,P} = P_{ultima} \\times (x-0.75L)', 'valores': {}, 'valor': momentos_p['M_C']})
        else:
              return 1.0, {'formula_simbolica': 'N/A', 'momentos': {}}

    # --- Soma das componentes e cálculo final do Cb ---
    M_finais = {}
    for m_key in ['M_max', 'M_A', 'M_B', 'M_C']:
        total = sum(c['valor'] for c in detalhes_cb['momentos'][m_key]['components'])
        detalhes_cb['momentos'][m_key]['final_value'] = abs(total)
        M_finais[m_key] = abs(total)

    denominador = (2.5 * M_finais['M_max'] + 3 * M_finais['M_A'] + 4 * M_finais['M_B'] + 3 * M_finais['M_C'])
    Cb = min(12.5 * M_finais['M_max'] / abs(denominador), 3.0) if denominador != 0 else 1.0

    return Cb, detalhes_cb

def calcular_flecha_maxima(tipo_viga, L_cm, E, Ix, q_serv_kn_cm=0, p_serv_load=None):
    delta_q, delta_p = 0, 0
    L = L_cm

    detalhes = {
        'delta_q': {'formula_simbolica': '', 'formula_numerica': '', 'valor': 0, 'unidade': 'cm'},
        'delta_p': {'formula_simbolica': '', 'formula_numerica': '', 'valor': 0, 'unidade': 'cm'},
        'delta_total': 0
    }

    if q_serv_kn_cm > 0:
        q_serv_val = q_serv_kn_cm
        if tipo_viga == 'Bi-apoiada':
            delta_q = (5 * q_serv_val * L**4) / (384 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{5 \\times q_{serv} \\times L^4}{384 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{5 \\times \\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{384 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            delta_q = (q_serv_val * L**4) / (8 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{q_{serv} \\times L^4}{8 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{\\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{8 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Bi-engastada':
            delta_q = (q_serv_val * L**4) / (384 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{q_{serv} \\times L^4}{384 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{\\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{384 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Engastada e Apoiada':
            delta_q = (q_serv_val * L**4) / (185 * E * Ix)
            detalhes['delta_q']['formula_simbolica'] = "\\delta_q = \\frac{q_{serv} \\times L^4}{185 \\times E \\times I_x}"
            detalhes['delta_q']['formula_numerica'] = f"\\frac{{\\mathbf{{{q_serv_val:.4f}}} \\times \\mathbf{{{L:.2f}}}^4}}{{185 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        detalhes['delta_q']['valor'] = delta_q

    if p_serv_load:
        P, x = p_serv_load
        a = x
        b = L - a
        if tipo_viga == 'Bi-apoiada':
            delta_p = (P * a**2 * b**2) / (3 * E * Ix * L)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^2 \\times b^2}{3 \\times E \\times I_x \\times L}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^2 \\times \\mathbf{{{b:.2f}}}^2}}{{3 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}} \\times \\mathbf{{{L:.2f}}}}}"
        elif tipo_viga == 'Engastada e Livre (Balanço)':
            delta_p = (P * a**3) / (3 * E * Ix)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^3}{3 \\times E \\times I_x}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^3}}{{3 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}}}}"
        elif tipo_viga == 'Bi-engastada':
            delta_p = (P * a**3 * b**3) / (3 * E * Ix * L**3)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^3 \\times b^3}{3 \\times E \\times I_x \\times L^3}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^3 \\times \\mathbf{{{b:.2f}}}^3}}{{3 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}} \\times \\mathbf{{{L:.2f}}}^3}}"
        elif tipo_viga == 'Engastada e Apoiada':
            delta_p = (P * a**3 * b**2) / (12 * E * Ix * L**3) * (a + 2*L)
            detalhes['delta_p']['formula_simbolica'] = "\\delta_p = \\frac{P_{serv} \\times a^3 \\times b^2 \\times (a + 2L)}{12 \\times E \\times I_x \\times L^3}"
            detalhes['delta_p']['formula_numerica'] = f"\\frac{{\\mathbf{{{P:.2f}}} \\times \\mathbf{{{a:.2f}}}^3 \\times \\mathbf{{{b:.2f}}}^2 \\times (\\mathbf{{{a:.2f}}} + 2 \\times \\mathbf{{{L:.2f}}})}}{{12 \\times \\mathbf{{{E:.0f}}} \\times \\mathbf{{{Ix:.2f}}} \\times \\mathbf{{{L:.2f}}}^3}}"

        detalhes['delta_p']['valor'] = delta_p

    detalhes['delta_total'] = delta_q + delta_p
    return detalhes

def get_profile_properties(profile_series):
    props = {
        "d": profile_series.get('d (mm)'),
        "bf": profile_series.get('bf (mm)'),
        "tw": profile_series.get('tw (mm)'),
        "tf": profile_series.get('tf (mm)'),
        # h_faces é a distância entre faces internas; d' desconta os raios nos laminados.
        "h_faces": profile_series.get('h (mm)'),
        "h_clear": profile_series.get("d' (mm)", profile_series.get('h (mm)')),
        "Area": profile_series.get('Área (cm2)'),
        "Ix": profile_series.get('Ix (cm4)'),
        "Wx": profile_series.get('Wx (cm3)'),
        "rx": profile_series.get('rx (cm)'),
        "Zx": profile_series.get('Zx (cm3)'),
        "Iy": profile_series.get('Iy (cm4)'),
        "Wy": profile_series.get('Wy (cm3)'),
        "ry": profile_series.get('ry (cm)'),
        "Zy": profile_series.get('Zy (cm3)'),
        "rt": profile_series.get('rt (cm)'),
        "J": profile_series.get('It (cm4)'),
        "Cw": profile_series.get('Cw (cm6)'),
        "Peso": profile_series.get('Massa Linear (kg/m)', profile_series.get('Peso (kg/m)')),
    }
    required_keys = ["d", "bf", "tw", "tf", "h_faces", "h_clear", "Area", "Ix", "Wx", "rx", "Zx", "Iy", "ry", "J", "Cw", "Peso"]
    profile_name = profile_series.get('Bitola (mm x kg/m)', 'Perfil Desconhecido')
    for key in required_keys:
        value = props.get(key)
        if value is None or pd.isna(value) or (isinstance(value, (int, float)) and value <= 0):
            raise ValueError(f"Propriedade ESSENCIAL '{key}' inválida ou nula no Excel para '{profile_name}'. Verifique a planilha.")
    for key in ['d', 'bf', 'tw', 'tf', 'h_faces', 'h_clear']:
        props[key] /= 10.0
    # Alias mantido apenas para blocos legados de apresentação.
    props['h'] = props['h_clear']
    return props

def _calcular_mrdx_flt(props, Lb, Cb, fy, E):
    Zx, ry, Iy, Cw, J, Wx = props['Zx'], props['ry'], props['Iy'], props['Cw'], props['J'], props['Wx']
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Mp = Zx * fy
    detalhes['passos_calculo'].append({
        'desc': 'Momento de Plastificação',
        'formula': 'M_p = Z_x \\times f_y',
        'valores': {'Z_x': Zx, 'f_y': fy},
        'valor': Mp,
        'unidade': 'kN.cm',
        'verif_id': 'Mp'
    })

    lambda_val = Lb / ry if ry > 0 else float('inf')
    detalhes['passos_calculo'].append({
        'desc': 'Índice de Esbeltez (λ = Lb/ry)',
        'formula': '\\lambda = \\frac{{L_b}}{{r_y}}',
        'valores': {'L_b': Lb, 'r_y': ry},
        'valor': lambda_val,
        'verif_id': 'lambda'
    })

    lambda_p = Config.FATOR_LAMBDA_P_FLT * math.sqrt(E / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)',
        'formula': '\\lambda_p = 1,76 \\sqrt{\\frac{{E}}{{f_y}}}',
        'valores': {'E': E, 'f_y': fy},
        'valor': lambda_p,
        'ref': 'Tabela D.1',
        'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        verificacao_texto = f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}"
        conclusao_texto = "SEÇÃO COMPACTA - O regime de flambagem é Plástico."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'REGIME PLÁSTICO',
            'verif_for_calc': 'lambda_p'
        })

        Mrdx = Mp / Config.GAMMA_A1
        detalhes['Mrdx_calc'] = {
            'desc': 'Momento Resistente (Regime Plástico)',
            'formula': 'M_{rd} = \\frac{{M_p}}{{\\gamma_{{a1}}}}',
            'valores': {'M_p': Mp, '\\gamma_{{a1}}': Config.GAMMA_A1},
            'valor': Mrdx,
            'unidade': 'kN.cm',
            'ref': 'Anexo D'
        }
    else:
        verificacao_texto = f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}"
        conclusao_texto = "SEÇÃO NÃO COMPACTA - O regime de flambagem é Inelástico ou Elástico."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'NECESSÁRIO VERIFICAR REGIME',
            'verif_for_calc': 'lambda_p'
        })

        sigma_r = Config.FATOR_SIGMA_R * fy
        detalhes['passos_calculo'].append({
            'desc': 'Tensão Residual (σr)',
            'formula': '\\sigma_r = 0,3 \\times f_y',
            'valores': {'f_y': fy},
            'valor': sigma_r,
            'unidade': 'kN/cm²',
            'verif_id': 'sigma_r'
        })

        Mr = (fy - sigma_r) * Wx
        detalhes['passos_calculo'].append({
            'desc': 'Momento de Escoamento Residual (Mr)',
            'formula': 'M_r = (f_y - \\sigma_r) \\times W_x',
            'valores': {'f_y': fy, '\\sigma_r': sigma_r, 'W_x': Wx},
            'valor': Mr,
            'unidade': 'kN.cm',
            'verif_id': 'Mr'
        })

        beta1 = ((fy - sigma_r) * Wx) / (E * J) if E * J != 0 else 0
        detalhes['passos_calculo'].append({
            'desc': 'Parâmetro β1',
            'formula': '\\beta_1 = \\frac{(f_y - \\sigma_r) \\times W_x}{E \\times J}',
            'valores': {'f_y': fy, '\\sigma_r': sigma_r, 'W_x': Wx, 'E': E, 'J': J},
            'valor': beta1,
            'unidade': '',
            'verif_id': 'beta1'
        })

        lambda_r = float('inf')
        if ry > 0 and beta1 > 0 and J > 0 and Cw > 0 and Iy > 0:
            termo_sqrt1 = 1 + (27 * Cw * (beta1**2) / Iy)
            termo_sqrt2 = 1 + math.sqrt(termo_sqrt1) if termo_sqrt1 >= 0 else 1
            lambda_r = (1.38 * math.sqrt(Iy * J) / (ry * beta1 * J)) * math.sqrt(termo_sqrt2)

        detalhes['passos_calculo'].append({
            'desc': 'Esbeltez Limite Inelástica (λr)',
            'formula': '\\lambda_r = 1,38 \\frac{\\sqrt{I_y \\times J}}{r_y \\times \\beta_1 \\times J} \\sqrt{1 + \\sqrt{1+\\frac{27 \\times C_w \\times \\beta_1^2}{I_y}}}',
            'valores': {'I_y': Iy, 'J': J, 'r_y': ry, '\\beta_1': beta1, 'C_w': Cw},
            'valor': lambda_r,
            'verif_id': 'lambda_r'
        })

        if lambda_val <= lambda_r:
            verificacao_texto = f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}"
            conclusao_texto = "REGIME INELÁSTICO - Flambagem no regime inelástico."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME INELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx_calc = (Cb / Config.GAMMA_A1) * (Mp - (Mp - Mr) * ((lambda_val - lambda_p) / (lambda_r - lambda_p)))
            Mp_gamma = Mp / Config.GAMMA_A1
            Mrdx = min(Mrdx_calc, Mp_gamma)

            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Regime Inelástico)',
                'formula': 'M_{rd,calc} = \\frac{{C_b}}{{\\gamma_{{a1}}}} [M_p - (M_p - M_r) (\\frac{{\\lambda - \\lambda_p}}{{\\lambda_r - \\lambda_p}})]',
                'valores': {'C_b': Cb, '\\gamma_{{a1}}': Config.GAMMA_A1, 'M_p': Mp, 'M_r': Mr,  '\\lambda': lambda_val, '\\lambda_p': lambda_p, '\\lambda_r': lambda_r},
                'valor': Mrdx_calc,
                'unidade': 'kN.cm',
                'ref': 'Anexo D'
            }

            detalhes['verificacao_limite'] = {
                'desc': 'Verificação do Limite de Plastificação',
                'texto': f"""
                    $M_{{rd,calc}} = {Mrdx_calc/100:.2f} \\, kNm$
                    $M_{{p,rd}} = \\frac{{M_p}}{{\\gamma_{{a1}}}} = {Mp_gamma/100:.2f} \\, kNm$
                    $M_{{rd}} = \\min(M_{{rd,calc}}; M_{{p,rd}}) = \\mathbf{{{Mrdx/100:.2f}}} \\, kNm$"""
            }
        else:
            verificacao_texto = f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}"
            conclusao_texto = "REGIME ELÁSTICO - Flambagem no regime elástico."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME ELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mcr = 0
            if Lb**2 > 0 and Iy > 0 and Cw > 0 and J > 0:
                Mcr = ((Cb * (math.pi**2) * E * Iy) / (Lb**2)) * math.sqrt((Cw/Iy) * (1 + (0.039 * J * (Lb**2) / Cw)))

            Mrdx = Mcr / Config.GAMMA_A1

            detalhes['passos_calculo'].append({
                'desc': 'Momento Crítico Elástico (Mcr)',
                'formula': 'M_{cr} = \\frac{{C_b \\times \\pi^2 \\times E \\times I_y}}{{L_b^2}} \\sqrt{{\\frac{{C_w}}{{I_y}}(1 + 0,039 \\times \\frac{{J \\times L_b^2}}{{C_w}})}}',
                'valores': {'C_b': Cb, '\\pi^2': math.pi**2, 'E': E, 'I_y': Iy, 'L_b': Lb, 'C_w': Cw, 'J': J},
                'valor': Mcr,
                'unidade': 'kN.cm',
                'ref': 'Anexo D',
                'verif_id': 'Mcr'
            })

            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Regime Elástico)',
                'formula': 'M_{rd} = \\frac{{M_{{cr}}}}{{\\gamma_{{a1}}}}',
                'valores': {'M_{{cr}}': Mcr, '\\gamma_{{a1}}': Config.GAMMA_A1},
                'valor': Mrdx,
                'unidade': 'kN.cm',
                'ref': 'Anexo D'
            }

    detalhes['Mrdx'] = Mrdx
    return detalhes

def _calcular_mrdx_flm(props, fy, tipo_fabricacao, E):
    bf, tf, Zx, Wx, h, tw = props['bf'], props['tf'], props['Zx'], props['Wx'], props['h'], props['tw']
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Mp = Zx * fy
    detalhes['passos_calculo'].append({
        'desc': 'Momento de Plastificação',
        'formula': 'M_p = Z_x \\times f_y',
        'valores': {'Z_x': Zx, 'f_y': fy},
        'valor': Mp,
        'unidade': 'kN.cm',
        'verif_id': 'Mp'
    })

    lambda_val = (bf / 2) / tf if tf > 0 else float('inf')
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez da Mesa (λ = bf/2tf)',
        'formula': '\\lambda = \\frac{{b_f/2}}{{t_f}}',
        'valores': {'b_f': bf, 't_f': tf},
        'valor': lambda_val,
        'verif_id': 'lambda'
    })

    lambda_p = Config.FATOR_LAMBDA_P_FLM * math.sqrt(E / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)',
        'formula': '\\lambda_p = 0,38 \\sqrt{{\\frac{{E}}{{f_y}}}}',
        'valores': {'E': E, 'f_y': fy},
        'valor': lambda_p,
        'ref': 'Tabela D.1',
        'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        verificacao_texto = f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}"
        conclusao_texto = "MESA COMPACTA - Não ocorre flambagem local da mesa."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'REGIME PLÁSTICO',
            'verif_for_calc': 'lambda_p'
        })

        Mrdx = Mp / Config.GAMMA_A1
        detalhes['Mrdx_calc'] = {
            'desc': 'Momento Resistente (Mesa Compacta)',
            'formula': 'M_{rd} = \\frac{{M_p}}{{\\gamma_{{a1}}}}',
            'valores': {'M_p': Mp, '\\gamma_{{a1}}': Config.GAMMA_A1},
            'valor': Mrdx,
            'unidade': 'kN.cm'
        }
    else:
        verificacao_texto = f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}"
        conclusao_texto = "MESA NÃO COMPACTA - Verificar se é semicompacta ou esbelta."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'NECESSÁRIO VERIFICAR REGIME',
            'verif_for_calc': 'lambda_p'
        })

        sigma_r = Config.FATOR_SIGMA_R * fy
        detalhes['passos_calculo'].append({
            'desc': 'Tensão Residual (σr)',
            'formula': '\\sigma_r = 0,3 \\times f_y',
            'valores': {'f_y': fy},
            'valor': sigma_r,
            'unidade': 'kN/cm²',
            'verif_id': 'sigma_r'
        })

        Mr = (fy - sigma_r) * Wx
        detalhes['passos_calculo'].append({
            'desc': 'Momento de Escoamento Residual (Mr)',
            'formula': 'M_r = (f_y - \\sigma_r) \\times W_x',
            'valores': {'f_y': fy, '\\sigma_r': sigma_r, 'W_x': Wx},
            'valor': Mr,
            'unidade': 'kN.cm',
            'verif_id': 'Mr'
        })

        if tipo_fabricacao == "Laminado":
            lambda_r = Config.FATOR_LAMBDA_R_FLM_LAMINADO * math.sqrt(E / (fy - sigma_r)) if (fy - sigma_r) > 0 else float('inf')
            lambda_r_formula_str = '\\lambda_r = 0,83 \\sqrt{{\\frac{{E}}{{f_y - \\sigma_r}}}}'
            detalhes['passos_calculo'].append({
                'desc': 'Esbeltez Limite Semicompacta (λr) - Laminado',
                'formula': lambda_r_formula_str,
                'valores': {'E': E, 'f_y': fy, '\\sigma_r': sigma_r},
                'valor': lambda_r,
                'ref': 'Tabela D.1',
                'verif_id': 'lambda_r'
            })
        else: # Soldado
            kc_val = 4 / math.sqrt(h/tw) if (h/tw) > 0 else 0.76
            kc = max(0.35, min(kc_val, 0.76))
            detalhes['passos_calculo'].append({
                'desc': 'Parâmetro kc',
                'formula': 'k_c = \\frac{4}{\\sqrt{h/t_w}} \\, (0,35 \\le k_c \\le 0,76)',
                'valores': {'h': h, 't_w': tw},
                'valor': kc,
                'unidade': '',
                'verif_id': 'kc'
            })

            lambda_r = Config.FATOR_LAMBDA_R_FLM_SOLDADO * math.sqrt(E * kc / (fy - sigma_r)) if (fy - sigma_r) > 0 and kc > 0 else float('inf')
            lambda_r_formula_str = '\\lambda_r = 0,95 \\sqrt{{\\frac{E \\times k_c}{{f_y - \\sigma_r}}}}'
            detalhes['passos_calculo'].append({
                'desc': 'Esbeltez Limite Semicompacta (λr) - Soldado',
                'formula': lambda_r_formula_str,
                'valores': {'E': E, 'k_c': kc, 'f_y': fy, '\\sigma_r': sigma_r},
                'valor': lambda_r,
                'ref': 'Tabela D.1',
                'verif_id': 'lambda_r'
            })

        if lambda_val <= lambda_r:
            verificacao_texto = f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}"
            conclusao_texto = "MESA SEMICOMPACTA - O regime é de transição."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME INELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx = (1 / Config.GAMMA_A1) * (Mp - (Mp - Mr) * ((lambda_val - lambda_p) / (lambda_r - lambda_p)))
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Mesa Semicompacta)',
                'formula': 'M_{rd} = \\frac{{1}}{{\\gamma_{{a1}}}} [M_p - (M_p - M_r) (\\frac{{\\lambda - \\lambda_p}}{{\\lambda_r - \\lambda_p}})]',
                'valores': {'\\gamma_{{a1}}': Config.GAMMA_A1, 'M_p': Mp, 'M_r': Mr,  '\\lambda': lambda_val, '\\lambda_p': lambda_p, '\\lambda_r': lambda_r},
                'valor': Mrdx,
                'unidade': 'kN.cm'
            }
        else:
            verificacao_texto = f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}"
            conclusao_texto = "MESA ESBELTA - O regime de flambagem é elástico."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME ELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            if tipo_fabricacao == "Laminado":
                Mcr = (0.69 * E * Wx) / (lambda_val**2) if lambda_val > 0 else 0
                detalhes['passos_calculo'].append({
                    'desc': 'Momento Crítico (Mcr) - Laminado',
                    'formula': 'M_{cr} = \\frac{0,69 \\times E \\times W_x}{\\lambda^2}',
                    'valores': {'E': E, 'W_x': Wx, '\\lambda': lambda_val},
                    'valor': Mcr,
                    'unidade': 'kN.cm',
                    'verif_id': 'Mcr'
                })
            else: # Soldado
                kc = detalhes['passos_calculo'][-2]['valor']
                Mcr = (0.90 * E * kc * Wx) / (lambda_val**2) if lambda_val > 0 else 0
                detalhes['passos_calculo'].append({
                    'desc': 'Momento Crítico (Mcr) - Soldado',
                    'formula': 'M_{cr} = \\frac{0,90 \\times E \\times k_c \\times W_x}{\\lambda^2}',
                    'valores': {'E': E, 'k_c': kc, 'W_x': Wx, '\\lambda': lambda_val},
                    'valor': Mcr,
                    'unidade': 'kN.cm',
                    'verif_id': 'Mcr'
                })

            Mrdx = Mcr / Config.GAMMA_A1
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Mesa Esbelta)',
                'formula': 'M_{rd} = \\frac{M_{cr}}{\\gamma_{a1}}',
                'valores': {'M_{cr}': Mcr, '\\gamma_{a1}': Config.GAMMA_A1},
                'valor': Mrdx,
                'unidade': 'kN.cm'
            }

    detalhes['Mrdx'] = Mrdx
    return detalhes

def _calcular_mrdx_fla(props, fy, E):
    h, tw, Zx, Wx = props['h'], props['tw'], props['Zx'], props['Wx']
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Mp = Zx * fy
    detalhes['passos_calculo'].append({
        'desc': 'Momento de Plastificação',
        'formula': 'M_p = Z_x \\times f_y',
        'valores': {'Z_x': Zx, 'f_y': fy},
        'valor': Mp,
        'unidade': 'kN.cm',
        'verif_id': 'Mp'
    })

    lambda_val = h / tw if tw > 0 else float('inf')
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez da Alma (λ = h/tw)',
        'formula': '\\lambda = \\frac{{h}}{{t_w}}',
        'valores': {'h': h, 't_w': tw},
        'valor': lambda_val,
        'verif_id': 'lambda'
    })

    lambda_p = Config.FATOR_LAMBDA_P_FLA * math.sqrt(E / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)',
        'formula': '\\lambda_p = 3,76 \\sqrt{{\\frac{{E}}{{f_y}}}}',
        'valores': {'E': E, 'f_y': fy},
        'valor': lambda_p,
        'ref': 'Tabela D.1',
        'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        verificacao_texto = f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}"
        conclusao_texto = "ALMA COMPACTA - Não ocorre flambagem local da alma."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'REGIME PLÁSTICO',
            'verif_for_calc': 'lambda_p'
        })

        Mrdx = Mp / Config.GAMMA_A1
        detalhes['Mrdx_calc'] = {
            'desc': 'Momento Resistente (Alma Compacta)',
            'formula': 'M_{rd} = \\frac{{M_p}}{{\\gamma_{{a1}}}}',
            'valores': {'M_p': Mp, '\\gamma_{{a1}}': Config.GAMMA_A1},
            'valor': Mrdx,
            'unidade': 'kN.cm'
        }
    else:
        verificacao_texto = f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}"
        conclusao_texto = "ALMA NÃO COMPACTA - Verificar se é semicompacta ou esbelta."
        detalhes['passos_verificacao'].append({
            'titulo': 'Verificação 1: λ ≤ λp?',
            'texto': verificacao_texto,
            'conclusao': conclusao_texto,
            'regime': 'NECESSÁRIO VERIFICAR REGIME',
            'verif_for_calc': 'lambda_p'
        })

        lambda_r = Config.FATOR_LAMBDA_R_FLA * math.sqrt(E / fy)
        detalhes['passos_calculo'].append({
            'desc': 'Esbeltez Limite Semicompacta (λr)',
            'formula': '\\lambda_r = 5,70 \\sqrt{{\\frac{{E}}{{f_y}}}}',
            'valores': {'E': E, 'f_y': fy},
            'valor': lambda_r,
            'ref': 'Tabela D.1',
            'verif_id': 'lambda_r'
        })

        Mr = fy * Wx
        detalhes['passos_calculo'].append({
            'desc': 'Momento de Escoamento (Mr)',
            'formula': 'M_r = f_y \\times W_x',
            'valores': {'f_y': fy, 'W_x': Wx},
            'valor': Mr,
            'unidade': 'kN.cm',
            'verif_id': 'Mr'
        })

        if lambda_val <= lambda_r:
            verificacao_texto = f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}"
            conclusao_texto = "ALMA SEMICOMPACTA - O regime é de transição."
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME INELÁSTICO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx = (1 / Config.GAMMA_A1) * (Mp - (Mp - Mr) * ((lambda_val - lambda_p) / (lambda_r - lambda_p)))
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Alma Semicompacta)',
                'formula': 'M_{rd} = \\frac{{1}}{{\\gamma_{{a1}}}} [M_p - (M_p - M_r) (\\frac{{\\lambda - \\lambda_p}}{{\\lambda_r - \\lambda_p}})]',
                'valores': {'\\gamma_{{a1}}': Config.GAMMA_A1, 'M_p': Mp, 'M_r': Mr,  '\\lambda': lambda_val, '\\lambda_p': lambda_p, '\\lambda_r': lambda_r},
                'valor': Mrdx,
                'unidade': 'kN.cm'
            }
        else:
            verificacao_texto = f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}"
            conclusao_texto = "ALMA ESBELTA - Regime elástico (Ver Anexo H da NBR 8800)"
            detalhes['passos_verificacao'].append({
                'titulo': 'Verificação 2: λ ≤ λr?',
                'texto': verificacao_texto,
                'conclusao': conclusao_texto,
                'regime': 'REGIME ELÁSTICO - NÃO IMPLEMENTADO',
                'verif_for_calc': 'lambda_r'
            })

            Mrdx = 0
            detalhes['Mrdx_calc'] = {
                'desc': 'Momento Resistente (Alma Esbelta)',
                'formula': 'N/A - Ver Anexo H da NBR 8800',
                'valores': {},
                'valor': Mrdx,
                'unidade': 'kN.cm',
                'ref': 'Perfil com alma esbelta. Ver Anexo H.'
            }

    detalhes['Mrdx'] = Mrdx
    return detalhes

def _calcular_vrd(props, fy, usa_enrijecedores, a_enr, E):
    d, h, tw = props['d'], props['h'], props['tw']
    lambda_val = h / tw if tw > 0 else float('inf')
    detalhes = {'passos_calculo': [], 'passos_verificacao': []}

    Vpl = Config.FATOR_VP * d * tw * fy
    detalhes['passos_calculo'].append({
        'desc': 'Força Cortante de Plastificação', 'formula': 'V_{pl} = 0,60 \\times d \\times t_{w} \\times f_{y}',
        'valores': {'d': d, 't_{w}': tw, 'f_{y}': fy}, 'valor': Vpl, 'unidade': 'kN', 'verif_id': 'Vpl'
    })

    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez da Alma (λ)', 'formula': '\\lambda = \\frac{h}{t_w}',
        'valores': {'h': h, 't_w': tw}, 'valor': lambda_val, 'verif_id': 'lambda'
    })

    if not usa_enrijecedores or a_enr <= 0:
        kv = 5.0
        detalhes['passos_calculo'].append({'desc': 'Fator de Flambagem (kv) - Alma sem enrijecedores', 'formula': 'k_v = 5,0', 'valores': {}, 'valor': kv})
    else:
        a_h_ratio = a_enr / h
        detalhes['passos_calculo'].append({
            'desc': 'Relação de Enrijecedores (a/h)',
            'formula': '\\frac{a}{h}',
            'formula_expandida': f'\\frac{{\\mathbf{{{a_enr:.2f}}}}}{{\\mathbf{{{h:.2f}}}}}',
            'valores': {},
            'valor': a_h_ratio
        })

        check1_pass = a_h_ratio < 3
        detalhes['passos_calculo'].append({
            'type': 'verification', 'desc': 'Verificação do Espaçamento Máximo',
            'lhs_value': a_h_ratio, 'comparator': '<', 'rhs_value': 3, 'passed': check1_pass,
            'conclusion_pass': 'O espaçamento entre enrijecedores atende ao critério inicial.',
            'conclusion_fail': 'O espaçamento é muito grande (a/h ≥ 3), tornando os enrijecedores ineficazes.'
        })

        check2_pass = False
        if check1_pass:
            limite_esbeltez = (260 / lambda_val)**2 if lambda_val > 0 else float('inf')
            detalhes['passos_calculo'].append({
                'desc': 'Limite de Esbeltez da Chapa', 'formula': 'Limite = (\\frac{260}{\\lambda})^2',
                'valores': {'\\lambda': lambda_val}, 'valor': limite_esbeltez
            })
            check2_pass = a_h_ratio < limite_esbeltez
            detalhes['passos_calculo'].append({
                'type': 'verification', 'desc': 'Verificação da Esbeltez da Chapa',
                'lhs_value': a_h_ratio, 'comparator': '<', 'rhs_value': limite_esbeltez, 'passed': check2_pass,
                'conclusion_pass': 'A alma não é esbelta demais para o espaçamento dos enrijecedores.',
                'conclusion_fail': 'A alma é muito esbelta para este espaçamento (a/h ≥ (260/λ)²), tornando os enrijecedores ineficazes.'
            })

        if check1_pass and check2_pass:
            denominador_val = a_h_ratio**2
            kv = 5 + (5 / denominador_val)
            kv_formula_expandida = (f"k_v = 5 + \\frac{{5}}{{(\\frac{{\\mathbf{{{a_enr:.2f}}}}}{{\\mathbf{{{h:.2f}}}}})^2}} = 5 + \\frac{{5}}{{(\\mathbf{{{a_h_ratio:.2f}}})^2}} = 5 + \\frac{{5}}{{\\mathbf{{{denominador_val:.2f}}}}}")
            detalhes['passos_calculo'].append({
                'desc': 'Cálculo Final de kv (Enrijecedores Efetivos)', 'formula': 'k_v = 5 + \\frac{5}{(a/h)^2}', 'formula_expandida': kv_formula_expandida,
                'valores': {'(a/h)': a_h_ratio}, 'valor': kv
            })
        else:
            kv = 5.0
            detalhes['passos_calculo'].append({'desc': 'Cálculo Final de kv (Enrijecedores Ineficazes)', 'formula': 'k_v = 5,0', 'valores': {}, 'valor': kv})

    lambda_p = Config.FATOR_LAMBDA_P_VRD * math.sqrt((kv * E) / fy)
    detalhes['passos_calculo'].append({
        'desc': 'Esbeltez Limite Plástica (λp)', 'formula': '\\lambda_p = 1,10 \\sqrt{\\frac{k_v \\times E}{f_y}}',
        'valores': {'k_v': kv, 'E': E, 'f_y': fy}, 'valor': lambda_p, 'verif_id': 'lambda_p'
    })

    if lambda_val <= lambda_p:
        Vrd = Vpl / Config.GAMMA_A1
        detalhes['passos_verificacao'].append({'titulo': 'Verificação 1: λ ≤ λp?', 'texto': f"λ = {lambda_val:.2f} ≤ λp = {lambda_p:.2f}", 'conclusao': "ESCOAMENTO DA ALMA - Resistência governada pelo escoamento.", 'regime': 'REGIME PLÁSTICO', 'verif_for_calc': 'lambda_p'})
        detalhes['Vrd_calc'] = {'desc': 'Cortante Resistente (Escoamento)', 'formula': 'V_{rd} = \\frac{V_{pl}}{\\gamma_{a1}}', 'valores': {'V_{pl}': Vpl, '\\gamma_{a1}': Config.GAMMA_A1}, 'valor': Vrd, 'unidade': 'kN'}
    else:
        detalhes['passos_verificacao'].append({'titulo': 'Verificação 1: λ ≤ λp?', 'texto': f"λ = {lambda_val:.2f} > λp = {lambda_p:.2f}", 'conclusao': "FLAMBAGEM POR CISALHAMENTO - O regime é Inelástico ou Elástico.", 'regime': 'NECESSÁRIO VERIFICAR REGIME', 'verif_for_calc': 'lambda_p'})
        lambda_r = Config.FATOR_LAMBDA_R_VRD * math.sqrt((kv * E) / fy)
        detalhes['passos_calculo'].append({'desc': 'Esbeltez Limite Inelástica (λr)', 'formula': '\\lambda_r = 1,37 \\sqrt{\\frac{k_v \\times E}{f_y}}', 'valores': {'k_v': kv, 'E': E, 'f_y': fy}, 'valor': lambda_r, 'verif_id': 'lambda_r'})
        if lambda_val <= lambda_r:
            detalhes['passos_verificacao'].append({'titulo': 'Verificação 2: λ ≤ λr?', 'texto': f"λ = {lambda_val:.2f} ≤ λr = {lambda_r:.2f}", 'conclusao': "FLAMBAGEM INELÁSTICA - Regime de transição por cisalhamento.", 'regime': 'REGIME INELÁSTICO', 'verif_for_calc': 'lambda_r'})
            Vrd = (lambda_p / lambda_val) * (Vpl / Config.GAMMA_A1) if lambda_val > 0 else 0
            detalhes['Vrd_calc'] = {'desc': 'Cortante Resistente (Flambagem Inelástica)', 'formula': 'V_{rd} = \\frac{\\lambda_p}{\\lambda} \\times \\frac{V_{pl}}{\\gamma_{a1}}', 'valores': {'\\lambda_p': lambda_p, '\\lambda': lambda_val, 'V_{pl}': Vpl, '\\gamma_{a1}': Config.GAMMA_A1}, 'valor': Vrd, 'unidade': 'kN'}
        else:
            detalhes['passos_verificacao'].append({'titulo': 'Verificação 2: λ ≤ λr?', 'texto': f"λ = {lambda_val:.2f} > λr = {lambda_r:.2f}", 'conclusao': "FLAMBAGEM ELÁSTICA - Regime elástico por cisalhamento.", 'regime': 'REGIME ELÁSTICO', 'verif_for_calc': 'lambda_r'})
            Vrd = (Config.FATOR_VRD_ELASTICO * (lambda_p / lambda_val)**2) * (Vpl / Config.GAMMA_A1) if lambda_val > 0 else 0
            detalhes['Vrd_calc'] = {'desc': 'Cortante Resistente (Flambagem Elástica)', 'formula': 'V_{rd} = 1,24 (\\frac{\\lambda_p}{\\lambda})^2 \\times \\frac{V_{pl}}{\\gamma_{a1}}', 'valores': {'\\lambda_p': lambda_p, '\\lambda': lambda_val, 'V_{pl}': Vpl, '\\gamma_{a1}': Config.GAMMA_A1}, 'valor': Vrd, 'unidade': 'kN'}

    detalhes['Vrd'] = Vrd
    return detalhes

# ==============================================================================
# 3. FUNÇÕES DE GERAÇÃO DE INTERFACE E GRÁFICOS
# ==============================================================================

def create_excel_with_colors(df_list, sheet_names):
    """
    Cria um arquivo Excel com múltiplas abas, aplicando formatação de cores
    baseada na eficiência dos perfis.
    """
    output = io.BytesIO()
    workbook = openpyxl.Workbook()

    # Remova a folha padrão criada automaticamente
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    for df, sheet_name in zip(df_list, sheet_names):
        sheet = workbook.create_sheet(title=sheet_name)

        # Escreva os cabeçalhos
        for col_idx, col_name in enumerate(df.columns, 1):
            sheet.cell(row=1, column=col_idx, value=col_name)

        # Defina os estilos de cores
        fill_fail = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') # Vermelho
        fill_warning_high = PatternFill(start_color='FFEBAE', end_color='FFEBAE', fill_type='solid') # Amarelo escuro (95-100%)
        fill_warning_low = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid') # Amarelo claro (80-95%)
        fill_pass = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid') # Verde
        
        # Escreva os dados e aplique as cores
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Regras de formatação para as colunas de eficiência
                if 'Ef.' in df.columns[col_idx-1]:
                    try:
                        efficiency = float(value)
                        if efficiency > 100.0:
                            cell.fill = fill_fail
                        elif efficiency > 95:
                            cell.fill = fill_warning_high
                        elif efficiency > 80:
                            cell.fill = fill_warning_low
                        else:
                            cell.fill = fill_pass
                    except (ValueError, TypeError):
                        pass

    workbook.save(output)
    output.seek(0)
    return output

# Mantenha esta função como está:
def create_professional_header():
    logo_url = "https://lh3.googleusercontent.com/a/ACg8ocKplHKwLPBbUbVCvtwvTPhn0aboS42tEQxuNtiVPVZ7Xboh1pk=s96-c" # Substitua pela URL da sua logo
    
    st.markdown(f"""
    <div class="pro-header">
        <div class="header-content">
            <img src="{logo_url}" alt="Logo HQ Engenharia">
            <h1 class="gradient-text">Calculadora Estrutural de Perfis</h1>
            <p>Análise de Perfis Metálicos | {Config.NOME_NORMA}</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

def create_metrics_dashboard(input_params):
    """Cria um dashboard com métricas principais do projeto e esforços."""
    st.markdown("### 📊 Parâmetros do Projeto")
    
    msd_value = input_params.get('Msd', 0)
    vsd_value = input_params.get('Vsd', 0)
    cb_value = input_params.get('Cb_projeto', 1.0)

    def br_number(value, decimals):
        formatted = f"{value:,.{decimals}f}"
        return formatted.replace(",", "_").replace(".", ",").replace("_", ".")

    msd_display = br_number(msd_value / 100, 2) if msd_value > 0 else "—"
    vsd_display = br_number(vsd_value, 2) if vsd_value > 0 else "—"
    cb_display = (
        "automático por perfil"
        if input_params.get('cb_modo_auto')
        else br_number(cb_value, 2)
    )

    st.markdown(f"""
    <section class="project-metrics-shell" aria-label="Parâmetros principais do projeto">
        <div class="project-metrics-grid">
            <article class="project-metric-card">
                <div class="project-metric-label">Norma</div>
                <div class="project-metric-value">NBR 8800:2024</div>
                <div class="project-metric-unit">Errata 1:2025</div>
            </article>
            <article class="project-metric-card">
                <div class="project-metric-label">Módulo de elasticidade</div>
                <div class="project-metric-value">{br_number(input_params['E_aco'], 0)}</div>
                <div class="project-metric-unit">kN/cm²</div>
            </article>
            <article class="project-metric-card">
                <div class="project-metric-label">Coeficiente γa1</div>
                <div class="project-metric-value">1,10</div>
                <div class="project-metric-unit">resistência</div>
            </article>
            <article class="project-metric-card">
                <div class="project-metric-label">Vão da viga</div>
                <div class="project-metric-value">{br_number(input_params['L_cm'] / 100, 2)}</div>
                <div class="project-metric-unit">m</div>
            </article>
            <article class="project-metric-card">
                <div class="project-metric-label">Resistência ao escoamento fy</div>
                <div class="project-metric-value">{br_number(input_params['fy_aco'], 1)}</div>
                <div class="project-metric-unit">kN/cm²</div>
            </article>
            <article class="project-metric-card" title="Momento fletor solicitante de cálculo">
                <div class="project-metric-label">Momento solicitante Msd</div>
                <div class="project-metric-value">{msd_display}</div>
                <div class="project-metric-unit">kN·m</div>
            </article>
            <article class="project-metric-card" title="Força cortante solicitante de cálculo">
                <div class="project-metric-label">Cortante solicitante Vsd</div>
                <div class="project-metric-value">{vsd_display}</div>
                <div class="project-metric-unit">kN</div>
            </article>
        </div>
        <div class="project-metrics-context">
            <span><strong>Cb</strong>{cb_display}</span>
            <span><strong>Lb</strong>{br_number(input_params['Lb_projeto'], 2)} cm</span>
            <span><strong>Limite de flecha</strong>L/{input_params['limite_flecha_divisor']:.0f}</span>
        </div>
    </section>
    """, unsafe_allow_html=True)

def style_classic_dataframe(df):
    """Aplica estilização clássica com cores sólidas ao DataFrame."""
    def color_efficiency(val):
        if pd.isna(val) or not isinstance(val, (int, float)): return ''
        if val > 100:   color = '#f8d7da'
        elif val > 95:  color = '#ffeeba'
        elif val > 80:  color = '#fff3cd'
        else:           color = '#d4edda'
        return f'background-color: {color}'

    def style_status(val):
        if val == 'APROVADO':
            return 'background-color: #d4edda; font-weight: bold; color: #155724;'
        elif val == 'REPROVADO':
            return 'background-color: #f8d7da; font-weight: bold; color: #721c24;'
        elif val == 'NÃO VERIFICADO':
            return 'background-color: #fff3cd; font-weight: bold; color: #856404;'
        return ''

    efficiency_cols = [col for col in df.columns if '%' in col]
    
    styled_df = df.style.map(color_efficiency, subset=efficiency_cols)
    
    if 'Status' in df.columns:
        styled_df = styled_df.map(style_status, subset=['Status'])
        
    format_number_2 = lambda val: "—" if pd.isna(val) else f"{val:.2f}"
    format_number_1 = lambda val: "—" if pd.isna(val) else f"{val:.1f}"
    format_dict = {
        "Peso (kg/m)": format_number_2,
        **{col: format_number_1 for col in efficiency_cols},
    }
    return styled_df.format(format_dict)

# Em create_top_profiles_chart(df_approved, top_n=10):
def create_top_profiles_chart(df_approved, top_n=10):
    if df_approved.empty: return None
    df_top = df_approved.head(top_n).sort_values(by='Peso (kg/m)', ascending=False)
    fig = go.Figure(go.Bar(
        y=df_top['Perfil'], x=df_top['Peso (kg/m)'], orientation='h',
        text=[f'{w:.2f} kg/m' for w in df_top['Peso (kg/m)']], textposition='auto',
        marker=dict(color=df_top['Peso (kg/m)'], colorscale='YlOrBr', colorbar=dict(title="Peso")), # Escala de cor Dourada
        hovertemplate='<b>%{y}</b><br>Peso: %{x:.2f} kg/m<extra></extra>'
    ))
    fig.update_layout(
        title={'text': f'🏆 Top {top_n} Perfis Mais Leves (Aprovados)', 'x': 0.5},
        xaxis_title='Peso (kg/m)', yaxis_title='Perfil', 
        template='plotly_dark', # <--- ADICIONE ESTA LINHA
        height=500, 
        margin=dict(l=150),
        paper_bgcolor='rgba(0,0,0,0)', # Fundo do papel transparente
        plot_bgcolor='rgba(0,0,0,0)'   # Fundo do gráfico transparente
    )
    return fig

# Em create_profile_efficiency_chart(perfil_nome, eficiencias):
# Substitua a função create_profile_efficiency_chart inteira por esta:
def create_profile_efficiency_chart(perfil_nome, eficiencias):
    """
    Cria um gráfico de barras comparando as eficiências de um perfil.
    """
    labels = list(eficiencias.keys())
    values = [min(v, 150) if isinstance(v, (int, float)) else 0 for v in eficiencias.values()]
    
    # Cores baseadas na eficiência
    colors = ['#32CD32' if v <= 100 else '#FF4500' for v in values]
    
    fig = go.Figure(data=[
        go.Bar(
            x=labels,
            y=values,
            text=[f'{v:.1f}%' if isinstance(v, (int, float)) else 'N/A' for v in eficiencias.values()],
            textposition='auto',
            marker_color=colors,
            textfont=dict(color='#FFFFFF', size=14, family='Poppins') # Texto dentro da barra
        )
    ])
    
    fig.add_hline(y=100, line_dash="dash", line_color="#fbbd24", # Linha dourada
                  annotation_text="Limite (100%)",
                  annotation_position="bottom right",
                  annotation_font=dict(color='#fbbd24'))

    # ATUALIZAÇÃO COMPLETA DO LAYOUT PARA O TEMA DARK
    fig.update_layout(
        title=dict(
            text=f'Análise de Eficiência para o Perfil: {perfil_nome}',
            font=dict(color='#FFFFFF', size=20, family='Poppins') # Cor do título
        ),
        yaxis_title='Eficiência (%)',
        xaxis_title='Verificação',
        yaxis_range=[0, max(max(values), 100) + 15],
        template='plotly_dark',
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='#94a3b8', family='Inter'), # Cor padrão para todo o texto do gráfico
        xaxis=dict(
            title_font=dict(color='#e2e8f0'), # Cor do título do eixo X
            tickfont=dict(color='#e2e8f0')     # Cor dos labels do eixo X
        ),
        yaxis=dict(
            title_font=dict(color='#e2e8f0'), # Cor do título do eixo Y
            tickfont=dict(color='#e2e8f0')     # Cor dos labels do eixo Y
        )
    )
    return fig
brazilia_tz = pytz.timezone('America/Sao_Paulo')
# Substitua a função create_professional_memorial_html por esta:
def create_professional_memorial_html(perfil_nome, perfil_tipo, resultados, input_details, projeto_info):
    # (O conteúdo da variável 'conteudo_memorial' continua o mesmo)
    conteudo_memorial = f""" 
    <h2>1. Resumo Executivo</h2>
    <div class="result-highlight">{resultados['resumo_html']}</div>
    <h2>2. Dados de Entrada e Solicitações</h2>
    <div class="info-card">
        <h3>2.1. Propriedades do Perfil e Materiais</h3>
        {input_details}
    </div>
    {resultados.get('esforcos_html', '')}
    {resultados.get('cb_calc_html', '')}
    {resultados['passo_a_passo_html']}
    """
    
    # O template HTML agora tem o <h1> com a classe 'gradient-text'
    html_template = f"""
    <!DOCTYPE html><html lang="pt-BR"><head>
        <meta charset="UTF-8"><title>Memorial de Cálculo - {perfil_nome}</title>
        {HTML_TEMPLATE_CSS_PRO}
        <script type="text/javascript" async src="https://cdnjs.cloudflare.com/ajax/libs/mathjax/2.7.7/MathJax.js?config=TeX-MML-AM_CHTML"></script>
    </head><body><div class="container">
        <div class="pro-header">
            <h1><span class="gradient-text">Memorial de Cálculo Estrutural</span></h1>
            <p><strong>{perfil_nome}</strong> ({perfil_tipo})</p>
        </div>
        <div class="info-card">
            <h3>📋 Identificação do Projeto</h3>
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 1rem;">
                <div><strong>Projeto:</strong> {projeto_info['nome']}</div>
                <div><strong>Engenheiro:</strong> {projeto_info['engenheiro']}</div>
                <div><strong>Data:</strong> {projeto_info['data']}</div>
                <div><strong>Revisão:</strong> {projeto_info['revisao']}</div>
            </div>
        </div>
        {conteudo_memorial}
        <div style="text-align: center; margin-top: 3rem; padding-top: 2rem; border-top: 1px solid var(--border); color: var(--text-secondary);">
            <p>Memorial gerado em {datetime.now(brazilia_tz).strftime('%d/%m/%Y às %H:%M')}</p>
        </div>
    </div></body></html>
    """
    return html_template

def _build_verification_block_html(title, solicitante, s_symbol, resistente, r_symbol, eficiencia, status, unit):
    status_class = "pass" if status == "APROVADO" else "fail"
    comp_symbol = "\\le" if status == "APROVADO" else ">"
    return f"""<h4>{title}</h4><div class="formula-block"><p class="formula">$${s_symbol} = {solicitante:.2f} \\, {unit}$$</p><p class="formula">$${r_symbol} = {resistente:.2f} \\, {unit}$$</p><p class="formula">$$\\text{{Verificação: }} {s_symbol} {comp_symbol} {r_symbol}$$</p><p class="formula">$$\\text{{Eficiência}} = \\frac{{{s_symbol}}}{{{r_symbol}}} = \\frac{{{solicitante:.2f}}}{{{resistente:.2f}}} = {eficiencia:.1f}\\%$$</p><div class="final-status {status_class}">{status}</div></div>"""

def _render_cb_calc_section(cb_details, Cb_final, input_mode):
    """Renderiza a seção de cálculo de Cb no memorial com passo a passo detalhado."""
    html = "<h3>2.3. Cálculo do Fator de Modificação Cb</h3>"

    if input_mode == 'Inserir Esforços Manualmente' or not cb_details or 'N/A' in cb_details.get('formula_simbolica', ''):
        html += "<div class='formula-block'>"
        html += f"<h5>Descrição:</h5><p>Fator Cb = **{Cb_final:.2f}** (Valor inserido manualmente ou não aplicável)</p>"
        html += "</div>"
        return html

    # Renderização Detalhada
    html += "<h4>Passo a Passo do Cálculo dos Momentos ($kNm$)</h4>"
    momentos_detalhes = cb_details.get('momentos', {})

    for m_key, m_data in momentos_detalhes.items():
        html += f"<div class='formula-block'>"
        html += f"<h5>Cálculo de ${m_key}$</h5>"
        
        # Renderiza cada componente de carga
        for component in m_data.get('components', []):
            html += _render_calculation_step(component)
        
        # Se houver mais de uma componente, mostra a soma
        if len(m_data.get('components', [])) > 1:
            comp_vals = [c['valor'] for c in m_data['components']]
            sum_str = ' + '.join([f'{abs(v):.2f}' for v in comp_vals])
            html += f"""<h6>Soma das Componentes para ${m_key}$</h6>
                        <p class="formula">$${m_key} = {sum_str} = \\mathbf{{{m_data['final_value']:.2f}}} \\, kNm$$</p>"""
        
        # Mostra o resultado final do momento
        html += f"<div class='final-status pass' style='font-size: 1.1em; padding: 0.75rem;'>Valor Final: ${m_key} = {m_data['final_value']:.2f} \\, kNm$</div>"
        html += "</div>"

    # Renderização do cálculo final do Cb
    html += "<div class='formula-block'>"
    html += f"<h4>Cálculo Final do Fator Cb</h4>"
    html += f"<p class='formula'>$${cb_details['formula_simbolica']}$$</p>"
    
    M_max_val = momentos_detalhes['M_max']['final_value']
    M_A_val = momentos_detalhes['M_A']['final_value']
    M_B_val = momentos_detalhes['M_B']['final_value']
    M_C_val = momentos_detalhes['M_C']['final_value']
    
    numerador_val = 12.5 * M_max_val
    denominador_val = 2.5 * M_max_val + 3 * M_A_val + 4 * M_B_val + 3 * M_C_val
        
    html += f"<h5>Cálculo numérico:</h5>"
    html += f"<p class='formula'>$$C_b = \\frac{{12,5 \\times \\mathbf{{{M_max_val:.2f}}}}}{{|2,5 \\times \\mathbf{{{M_max_val:.2f}}} + 3 \\times \\mathbf{{{M_A_val:.2f}}} + 4 \\times \\mathbf{{{M_B_val:.2f}}} + 3 \\times \\mathbf{{{M_C_val:.2f}}}|}} = \\frac{{{numerador_val:.2f}}}{{{abs(denominador_val):.2f}}}$$</p>"
    html += f"<p class='formula'>$$C_b = \\mathbf{{{Cb_final:.2f}}}$$</p>"
    html += "</div>"
    return html

def _render_esforcos_viga_section(params):
    """Gera a seção do memorial com o cálculo dos esforços solicitantes."""
    # --- AJUSTE: Nova apresentação visual para o modo de entrada manual ---
    if params['input_mode'] == 'Inserir Esforços Manualmente':
        msd_knm = params['Msd'] / 100
        vsd_kn = params['Vsd']
        return f"""
        <h3>2.2. Esforços Solicitantes de Cálculo (ELU)</h3>
        <div class="info-card">
            <div style="display: flex; justify-content: space-around; align-items: stretch; flex-wrap: wrap; gap: 1rem; padding: 1rem 0;">

                <div style="text-align: center; border: 1px solid #e5e7eb; border-radius: 12px; padding: 1.5rem; width: 45%; min-width: 200px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                    <span style="font-size: 2em; line-height: 1;">🔄</span>
                    <h5 style="margin: 0.5rem 0; color: var(--text-secondary); font-weight: 500;">Momento Fletor ($M_{{sd}}$)</h5>
                    <p style="font-size: 2.2em; font-weight: 700; color: var(--primary-color); margin: 0; line-height: 1.2;">
                        {msd_knm:.2f}
                        <span style="font-size: 0.5em; font-weight: 500; color: var(--text-primary);">kNm</span>
                    </p>
                </div>

                <div style="text-align: center; border: 1px solid #e5e7eb; border-radius: 12px; padding: 1.5rem; width: 45%; min-width: 200px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
                    <span style="font-size: 2em; line-height: 1;">✂️</span>
                    <h5 style="margin: 0.5rem 0; color: var(--text-secondary); font-weight: 500;">Força Cortante ($V_{{sd}}$)</h5>
                    <p style="font-size: 2.2em; font-weight: 700; color: var(--primary-color); margin: 0; line-height: 1.2;">
                        {vsd_kn:.2f}
                        <span style="font-size: 0.5em; font-weight: 500; color: var(--text-primary);">kN</span>
                    </p>
                </div>

            </div>
            <p style="text-align: center; font-size: 0.9em; color: var(--text-secondary); margin-top: 1rem;"><i>(Valores inseridos manualmente)</i></p>
        </div>
        """

    # Renderização para o modo de cálculo automático (permanece inalterada)
    html = f"""
    <h3>2.2. Esforços Solicitantes de Cálculo (ELU)</h3>
    <div class="info-card">
        <h4>Cálculo das Cargas de Projeto (Últimas)</h4>
        <div class="formula-block">
            <h5>Largura de Influência (B)</h5>
            <p class="formula">$$B = \\frac{{b_{{esq}}}}{2} + \\frac{{b_{{dir}}}}{2} = \\frac{{ {params['larg_esq_cm']:.2f} }}{2} + \\frac{{ {params['larg_dir_cm']:.2f} }}{2} = \\mathbf{{ {params['larg_inf_total_m']:.2f} \\, m}}$$</p>

            <h5>Carga Distribuída de Serviço (q_serv)</h5>
            <p class="formula">$$q_{{serv}} = Carga_{{area}} \\times B = {params['carga_area']:.2f} \\, kN/m^2 \\times {params['larg_inf_total_m']:.2f} \\, m = \\mathbf{{ {params['q_servico_kn_m']:.2f} \\, kN/m}}$$</p>

            <h5>Carga Distribuída de Cálculo (q_ultima)</h5>
            <p class="formula">$$q_{{ultima}} = q_{{serv}} \\times \\gamma_f = {params['q_servico_kn_m']:.2f} \\times {params['gamma_f']:.2f} = \\mathbf{{ {params['q_ult_kn_cm']*100:.2f} \\, kN/m}}$$</p>
    """
    if params.get('p_load_serv'):
        html += f"""
            <h5>Carga Pontual de Cálculo (P_ultima)</h5>
            <p class="formula">$$P_{{ultima}} = P_{{serv}} \\times \\gamma_f = {params['p_load_serv'][0]:.2f} \\times {params['gamma_f']:.2f} = \\mathbf{{ {params['p_load_ult'][0]:.2f} \\, kN}}$$</p>
        """
    html += """
        </div>
    </div>
    """

    html += f"<h4>Cálculo dos Esforços para Viga {params['tipo_viga']}</h4>"
    html += "<div class='formula-block'>"

    detalhes = params['detalhes_esforcos']

    # Momento
    html += f"<h5>Momento Fletor de Cálculo (Msd)</h5>"
    if detalhes['Msd_q']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Momento devido à carga distribuída (M_q)',
            'formula': detalhes['Msd_q']['formula_simbolica'],
            'valores': {'q_{ultima}': params['q_ult_kn_cm']*100, 'L': params['L_cm']/100},
            'valor': detalhes['Msd_q']['valor'],
            'unidade': 'kN.cm'
        })
    if detalhes['Msd_p']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Momento devido à carga pontual (M_p)',
            'formula': detalhes['Msd_p']['formula_simbolica'],
            'valores': {'P_{ultima}': params['p_load_ult'][0], 'a': params['p_load_ult'][1]/100, 'b': (params['L_cm']-params['p_load_ult'][1])/100, 'L': params['L_cm']/100},
            'valor': detalhes['Msd_p']['valor'],
            'unidade': 'kN.cm'
        })

    # Soma dos momentos
    if (detalhes['Msd_q']['valor'] > 0 or detalhes['Msd_p']['valor'] > 0):
        mom_q = detalhes['Msd_q']['valor'] if detalhes['Msd_q']['valor'] > 0 else 0
        mom_p = detalhes['Msd_p']['valor'] if detalhes['Msd_p']['valor'] > 0 else 0
        if mom_q > 0 and mom_p > 0:
            html += f"<p class='formula'>$$M_{{sd, total}} = {mom_q:.2f} + {mom_p:.2f} = \\mathbf{{{params['Msd']:.2f}}} \\, kN.cm$$</p>"
        else:
            html += f"<p class='formula'>$$M_{{sd}} = \\mathbf{{{params['Msd']:.2f}}} \\, kN.cm$$</p>"

    # Cortante
    html += f"<h5>Força Cortante de Cálculo (Vsd)</h5>"
    if detalhes['Vsd_q']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Cortante devido à carga distribuída (V_q)',
            'formula': detalhes['Vsd_q']['formula_simbolica'],
            'valores': {'q_{ultima}': params['q_ult_kn_cm']*100, 'L': params['L_cm']/100},
            'valor': detalhes['Vsd_q']['valor'],
            'unidade': 'kN'
        })
    if detalhes['Vsd_p']['valor'] > 0:
        html += _render_calculation_step({
            'desc': 'Cortante devido à carga pontual (V_p)',
            'formula': detalhes['Vsd_p']['formula_simbolica'],
            'valores': {'P_{ultima}': params['p_load_ult'][0], 'a': params['p_load_ult'][1]/100, 'b': (params['L_cm']-params['p_load_ult'][1])/100, 'L': params['L_cm']/100},
            'valor': detalhes['Vsd_p']['valor'],
            'unidade': 'kN'
        })

    # Soma dos cortantes
    if (detalhes['Vsd_q']['valor'] > 0 or detalhes['Vsd_p']['valor'] > 0):
        v_q = detalhes['Vsd_q']['valor'] if detalhes['Vsd_q']['valor'] > 0 else 0
        v_p = detalhes['Vsd_p']['valor'] if detalhes['Vsd_p']['valor'] > 0 else 0
        if v_q > 0 and v_p > 0:
            html += f"<p class='formula'>$$V_{{sd, total}} = {v_q:.2f} + {v_p:.2f} = \\mathbf{{{params['Vsd']:.2f}}} \\, kN$$</p>"
        else:
            html += f"<p class='formula'>$$V_{{sd}} = \\mathbf{{{params['Vsd']:.2f}}} \\, kN$$</p>"

    html += "</div>"

    return html

def _verification_status(demand, resistance, applicable=True):
    if not applicable:
        return 0.0, "N/A"
    if resistance is None or resistance <= 0:
        return float('inf'), "NÃO VERIFICADO"
    efficiency = demand / resistance * 100.0
    return efficiency, "APROVADO" if demand <= resistance else "REPROVADO"


def _memorial_2024_html_legacy_summary(bundle):
    """Resumo anterior preservado apenas como referência histórica interna."""
    """Memorial autocontido, com hipóteses, rastreabilidade e resultados por ELU/ELS."""
    flex = bundle['flexure']
    shear = bundle['shear']
    elu = bundle.get('elu_response')
    els = bundle.get('els_response')
    local_rows = ""
    for check in bundle.get('local_checks', []):
        details = check['details']
        state_values = (
            f"Escoamento local: {details['yielding_FRd']:.2f} kN; "
            f"enrugamento: {details['crippling_FRd']:.2f} kN"
        )
        if details.get('sidesway_FRd') is not None:
            state_values += f"; flambagem lateral da alma: {details['sidesway_FRd']:.2f} kN"
        local_rows += (
            f"<tr><td>{check['name']}</td><td>{check['demand']:.2f} kN</td>"
            f"<td>{check['resistance']:.2f} kN</td><td>{check['efficiency']:.1f}%</td>"
            f"<td class=\"{'pass' if check['status'] == 'APROVADO' else 'fail'}\">{check['status']}</td></tr>"
            f"<tr><td colspan='5'><small>{state_values}. Caso 5.7.5: {details['sidesway_case']}.</small></td></tr>"
        )
    if not local_rows:
        local_rows = "<tr><td colspan='5'>Não aplicável ou não disponível para o modo de entrada.</td></tr>"

    analysis_html = ""
    if elu:
        analysis_html = f"""
        <h3>3.1 Análise estrutural e combinação última normal</h3>
        <div class="formula-block">
          <p><strong>Combinação:</strong> {bundle['elu_combination_text']}</p>
          <p>q<sub>d</sub> = {bundle['elu_loads']['q']*100:.3f} kN/m; P<sub>d</sub> = {bundle['elu_loads']['P']:.3f} kN.</p>
          <p>R<sub>A,d</sub> = {elu.reaction_left:.3f} kN; R<sub>B,d</sub> = {elu.reaction_right:.3f} kN.</p>
          <p>M<sub>Sd,max</sub> = {elu.max_moment/100:.3f} kN·m em x = {elu.max_moment_position/100:.3f} m.</p>
          <p>V<sub>Sd,max</sub> = {elu.max_shear:.3f} kN.</p>
          <p><span class="ref-norma">{bundle['elu_loads']['reference']}; análise elástica de primeira ordem.</span></p>
        </div>"""
    else:
        analysis_html = f"""
        <h3>3.1 Esforços fornecidos pelo usuário</h3>
        <div class="formula-block"><p>M<sub>Sd</sub> = {bundle['Msd']/100:.3f} kN·m; V<sub>Sd</sub> = {bundle['Vsd']:.3f} kN.</p>
        <p>Reações, diagrama e forças localizadas não são inferidos no modo manual.</p></div>"""

    flt_value = "N/A" if flex['Mrd_FLT'] is None else f"{flex['Mrd_FLT']/100:.3f} kN·m"
    annex_e = "Sim" if flex['slender_web'] else "Não"
    rupture_html = ""
    if flex.get('Mrd_rupture') is not None:
        rupture_html = (
            f"<p>Furos na mesa tracionada: limite por ruptura = "
            f"{flex['Mrd_rupture']/100:.3f} kN·m; "
            f"condição fu·Afn ≥ Yt·fy·Afg: "
            f"{'atendida' if flex['rupture_condition_ok'] else 'não atendida'} "
            f"(5.4.2.6).</p>"
        )
    scope_items = "".join(f"<li>{item}</li>" for item in bundle.get('scope_notes', []))
    scope_issues = "".join(f"<li>{item}</li>" for item in bundle.get('scope_issues', [])) or "<li>Nenhuma pendência declarada.</li>"

    els_html = "<p>Não calculado no modo manual.</p>"
    if els:
        els_html = f"""
        <div class="formula-block">
          <p><strong>Combinação de serviço:</strong> {bundle['els_combination_text']}</p>
          <p>δ<sub>max</sub> = {els.max_deflection:.4f} cm em x = {els.max_deflection_position/100:.3f} m.</p>
          <p>δ<sub>lim</sub> = {bundle['deflection_limit']:.4f} cm (L/{bundle['deflection_divisor']}; em balanço, L é o dobro do comprimento teórico).</p>
          <p><strong>Resultado:</strong> {bundle['deflection_status']} — {bundle['deflection_efficiency']:.1f}%.</p>
          <p><span class="ref-norma">ABNT NBR 8800:2024, 4.8.7.3 e Anexo B.</span></p>
        </div>"""

    return f"""
    <h2>3. Modelo, ações e combinações</h2>
    {analysis_html}
    <h3>3.2 Fator Cb e trecho destravado</h3>
    <div class="formula-block">
      <p>Lb = {bundle['Lb']/100:.3f} m; Cb = {bundle['Cb']:.4f}.</p>
      <p>{bundle['cb_basis']}</p>
      <p><span class="ref-norma">ABNT NBR 8800:2024, 5.4.2.3 a 5.4.2.5.</span></p>
    </div>
    <h2>4. Resistência à flexão</h2>
    <div class="formula-block">
      <p>Alma esbelta pelo Anexo E: <strong>{annex_e}</strong>; h/tw = {flex['lambda_FLA']:.3f};
      λp = {flex['lambda_p_FLA']:.3f}; λr = {flex['lambda_r_FLA']:.3f}.</p>
      <p>Mpl = {flex['Mpl']/100:.3f} kN·m; limite global 1,50·W·fy/γa1 = {flex['cap_5_4_2_2']/100:.3f} kN·m.</p>
      <p>FLT: {flt_value} ({flex['regime_FLT']}); FLM: {flex['Mrd_FLM']/100:.3f} kN·m ({flex['regime_FLM']});
      FLA/mesa tracionada: {flex['Mrd_FLA_or_tension']/100:.3f} kN·m ({flex['regime_FLA']}).</p>
      {rupture_html}
      <p><strong>MRd governante = {flex['Mrd']/100:.3f} kN·m.</strong></p>
      <p><span class="ref-norma">{flex['reference']}; alternativa χLT de D.2.1 quando aplicável.</span></p>
    </div>
    <h2>5. Resistência ao cisalhamento e enrijecedores</h2>
    <div class="formula-block">
      <p>h/tw = {shear['lambda']:.3f}; kv = {shear['kv']:.3f} ({shear['kv_basis']}).</p>
      <p>λp = {shear['lambda_p']:.3f}; λr = {shear['lambda_r']:.3f}; regime: {shear['regime']}.</p>
      <p><strong>VRd = {shear['Vrd']:.3f} kN.</strong></p>
      <p>Enrijecedores solicitados: {'sim' if shear['stiffener_requested'] else 'não'};
      validados: {'sim' if shear['stiffener_valid'] else 'não/não aplicável'}.</p>
      <p><span class="ref-norma">{shear['reference']}.</span></p>
    </div>
    <h2>6. Forças transversais localizadas</h2>
    <table><thead><tr><th>Local</th><th>FSd</th><th>FRd</th><th>Eficiência</th><th>Status</th></tr></thead>
    <tbody>{local_rows}</tbody></table>
    <p><span class="ref-norma">ABNT NBR 8800:2024, 5.7.3 a 5.7.5. O projeto executivo dos enrijecedores locais e soldas deve atender 5.7.9.</span></p>
    <h2>7. Estado-limite de serviço — deslocamento</h2>
    {els_html}
    <h2>8. Escopo, hipóteses e pendências</h2>
    <div class="info-card"><h4>Hipóteses declaradas</h4><ul>{scope_items}</ul>
    <h4>Pendências que impedem aprovação plena</h4><ul>{scope_issues}</ul>
    <p><strong>Status global: {bundle['status_global']}</strong></p></div>
    """


def _memorial_2024_html(bundle):
    """Renderiza o memorial auditável com os dados do núcleo normativo."""
    return build_memorial_details(bundle)


def perform_all_checks(props, fy_aco, Lb_projeto, Cb_projeto, L_cm, Msd, Vsd, q_serv_kn_cm, p_load_serv, tipo_viga, input_mode, tipo_fabricacao, usa_enrijecedores, a_enr, limite_flecha_divisor, projeto_info, E_aco, detalhado=False, fu_aco=45.0, **kwargs):
    automatic = input_mode == "Calcular a partir de Cargas na Viga"
    scope_issues = list(kwargs.get('unsupported_reasons', []))
    scope_notes = list(kwargs.get('scope_notes', []))
    scope_issues.extend(validate_material(fy_aco, fu_aco))

    self_weight = props['Peso'] * 9.80665 / 100_000.0 if kwargs.get('include_self_weight', True) else 0.0
    elu_response = None
    els_response = None
    elu_loads = None
    els_loads = None
    point_position = kwargs.get('p_pos_cm', p_load_serv[1] if p_load_serv else L_cm / 2.0)
    Cb_final = Cb_projeto
    cb_info = None
    cb_basis = "Cb informado pelo usuário; a origem deve ser registrada no projeto."
    if (
        tipo_viga == 'Engastada e Livre (Balanço)'
        and kwargs.get('cantilever_standard_cb', False)
    ):
        cb_basis = "Cb = 1,0 para a condição de balanço declarada em 5.4.2.3-b."
    elif kwargs.get('cb_source', '').strip():
        cb_basis = f"Cb informado pelo usuário. Origem registrada: {kwargs['cb_source'].strip()}."

    if automatic:
        q_g = kwargs.get('q_g_kn_cm', 0.0)
        q_q = kwargs.get('q_q_kn_cm', q_serv_kn_cm)
        p_g = kwargs.get('p_g_kn', 0.0)
        p_q = kwargs.get('p_q_kn', p_load_serv[0] if p_load_serv else 0.0)
        elu_loads = combine_elu_normal(
            q_g, q_q, self_weight, p_g, p_q,
            gamma_g=kwargs.get('gamma_g', 1.50),
            gamma_q=kwargs.get('gamma_q', 1.50),
            gamma_self_weight=kwargs.get('gamma_self_weight', 1.25),
        )
        elu_response = analyze_beam(
            tipo_viga, L_cm, elu_loads['q'], elu_loads['P'], point_position
        )
        Msd, Vsd = elu_response.max_moment, elu_response.max_shear

        if kwargs.get('cb_modo_auto', False):
            cb_info = calculate_cb_nbr2024(
                elu_response,
                segment_start=kwargs.get('lb_start_cm', 0.0),
                unbraced_length=Lb_projeto,
            )
            Cb_final = cb_info['Cb']
            cb_basis = (
                f"Cálculo automático no trecho x={cb_info['segment_start']/100:.3f} m a "
                f"x={(cb_info['segment_start']+cb_info['Lb'])/100:.3f} m: "
                f"|Mmax|={cb_info['Mmax']/100:.3f}, |MA|={cb_info['MA']/100:.3f}, "
                f"|MB|={cb_info['MB']/100:.3f}, |MC|={cb_info['MC']/100:.3f} kN·m."
            )

        els_loads = combine_els(
            q_g, q_q, self_weight, p_g, p_q,
            combination=kwargs.get('els_combination', 'rare'),
            psi1=kwargs.get('psi1', 0.6),
            psi2=kwargs.get('psi2', 0.4),
        )
        els_response = analyze_beam(
            tipo_viga, L_cm, els_loads['q'], els_loads['P'], point_position,
            E=E_aco, I=props['Ix'],
        )
    elif not kwargs.get('manual_local_checks_confirmed', False):
        scope_issues.append("Modo manual sem reações/forças localizadas e sem verificação ELS.")

    Afg_tension = None
    Afn_tension = None
    if kwargs.get('has_tension_flange_holes', False):
        Afg_tension = props['bf'] * props['tf']
        Afn_tension = Afg_tension * kwargs.get('tension_flange_net_ratio', 1.0)

    flex = flexural_strength_i(
        props, fy_aco, fu_aco, E_aco, Lb_projeto, Cb_final, tipo_fabricacao,
        stiffener_spacing=a_enr if usa_enrijecedores else None,
        flt_applicable=kwargs.get('flt_applicable', True),
        net_tension_flange_area=Afn_tension,
        gross_tension_flange_area=Afg_tension,
    )
    scope_issues.extend(flex['applicability_issues'])

    shear = shear_strength_i(
        props, fy_aco, E_aco,
        stiffener_spacing=a_enr if usa_enrijecedores else None,
        stiffener_width=kwargs.get('stiffener_width'),
        stiffener_thickness=kwargs.get('stiffener_thickness'),
        stiffener_pair=kwargs.get('stiffener_pair', True),
        stiffener_welded_to_web_and_flanges=kwargs.get('stiffener_welded', False),
    )

    flt_eff, flt_status = _verification_status(Msd, flex['Mrd_FLT'], kwargs.get('flt_applicable', True))
    flm_eff, flm_status = _verification_status(Msd, flex['Mrd_FLM'])
    fla_eff, fla_status = _verification_status(Msd, flex['Mrd_FLA_or_tension'])
    rupture_eff, rupture_status = _verification_status(
        Msd, flex['Mrd_rupture'], flex['Mrd_rupture'] is not None
    )
    shear_eff, shear_status = _verification_status(Vsd, shear['Vrd'])

    res_flt = {'Mrdx': flex['Mrd_FLT'] or flex['Mrd'], 'eficiencia': flt_eff, 'status': flt_status, 'core': flex, 'Msd': Msd, 'titulo': 'Flexão — FLT'}
    res_flt.update({
        'rupture_Mrd': flex['Mrd_rupture'],
        'rupture_efficiency': rupture_eff,
        'rupture_status': rupture_status,
    })
    res_flm = {'Mrdx': flex['Mrd_FLM'], 'eficiencia': flm_eff, 'status': flm_status, 'core': flex, 'Msd': Msd, 'titulo': 'Flexão — FLM'}
    res_fla = {'Mrdx': flex['Mrd_FLA_or_tension'], 'eficiencia': fla_eff, 'status': fla_status, 'core': flex, 'Msd': Msd, 'titulo': 'Flexão — FLA/Anexo E'}
    res_cis = {'Vrd': shear['Vrd'], 'eficiencia': shear_eff, 'status': shear_status, 'core': shear, 'Vsd': Vsd}

    local_checks = []
    local_statuses = []
    if automatic and elu_response:
        locations = [
            (
                "Apoio esquerdo", abs(elu_response.reaction_left),
                kwargs.get('bearing_left_cm', 10.0), 0.0, 0.0,
                kwargs.get('support_relative_lateral_restrained', True),
            ),
        ]
        if tipo_viga != 'Engastada e Livre (Balanço)':
            locations.append((
                "Apoio direito", abs(elu_response.reaction_right),
                kwargs.get('bearing_right_cm', 10.0), 0.0, L_cm,
                kwargs.get('support_relative_lateral_restrained', True),
            ))
        if elu_loads['P'] > 0:
            locations.append((
                "Carga pontual", elu_loads['P'], kwargs.get('point_bearing_cm', 10.0),
                min(point_position, L_cm - point_position), point_position,
                kwargs.get('point_relative_lateral_restrained', True),
            ))
        for name, demand, bearing, distance_end, x_load, lateral_restrained in locations:
            local = local_compression_strength(
                props, fy_aco, E_aco, bearing, distance_end, tipo_fabricacao,
                weld_root_or_radius=kwargs.get('weld_root_cm', 0.0),
                lateral_unbraced_length=kwargs.get('local_unbraced_cm', Lb_projeto),
                flange_rotation_restrained=kwargs.get('loaded_flange_rotation_restrained', True),
                relative_lateral_movement_restrained=lateral_restrained,
                moment_at_load=elu_response.moment_at(x_load),
            )
            efficiency, status = _verification_status(demand, local['FRd'])
            local_checks.append({
                'name': name, 'demand': demand, 'resistance': local['FRd'],
                'efficiency': efficiency, 'status': status, 'details': local,
                'position': x_load, 'bearing_length': bearing,
            })
            local_statuses.append(status)

    flecha_max = flecha_limite = eficiencia_flecha = 0.0
    status_flecha = "N/A"
    if els_response:
        absolute_limit = 1.5 if kwargs.get('masonry_on_beam', False) else None
        flecha_limite = deflection_limit(tipo_viga, L_cm, limite_flecha_divisor, absolute_limit)
        flecha_max = els_response.max_deflection
        eficiencia_flecha, status_flecha = _verification_status(flecha_max, flecha_limite)
    res_flecha = {
        'flecha_max': flecha_max, 'flecha_limite': flecha_limite,
        'eficiencia': eficiencia_flecha, 'status': status_flecha,
        'Ix': props['Ix'], 'detalhes': {}, 'divisor': limite_flecha_divisor,
        'response': els_response,
    }

    statuses = [
        flt_status, flm_status, fla_status, rupture_status,
        shear_status, status_flecha,
    ] + local_statuses
    if scope_issues:
        statuses.append("NÃO VERIFICADO")
    status_global = overall_status(statuses)
    for result in (res_flt, res_flm, res_fla, res_cis, res_flecha):
        result['status_global'] = status_global
        result['scope_issues'] = scope_issues
        result['local_checks'] = local_checks

    bundle = {
        'Msd': Msd, 'Vsd': Vsd, 'flexure': flex, 'shear': shear,
        'elu_response': elu_response, 'els_response': els_response,
        'elu_loads': elu_loads, 'els_loads': els_loads,
        'elu_combination_text': kwargs.get('elu_combination_text', '1,50·G + 1,25·PP aço + 1,50·Q'),
        'els_combination_text': kwargs.get('els_combination_text', kwargs.get('els_combination', 'rare')),
        'Cb': Cb_final, 'Lb': Lb_projeto, 'cb_basis': cb_basis, 'cb_info': cb_info,
        'props': props, 'fy': fy_aco, 'fu': fu_aco, 'E': E_aco,
        'fabrication': tipo_fabricacao, 'support': tipo_viga, 'length': L_cm,
        'input_mode': input_mode, 'point_position': point_position,
        'self_weight': self_weight, 'gamma_a1': 1.10, 'gamma_a2': 1.35,
        'influence_left_cm': kwargs.get('larg_esq_cm'),
        'influence_right_cm': kwargs.get('larg_dir_cm'),
        'influence_width_m': kwargs.get('larg_inf_total_m'),
        'g_area': kwargs.get('g_area'), 'q_area': kwargs.get('q_area'),
        'masonry_on_beam': kwargs.get('masonry_on_beam', False),
        'local_checks': local_checks,
        'deflection_limit': flecha_limite, 'deflection_divisor': limite_flecha_divisor,
        'deflection_efficiency': eficiencia_flecha, 'deflection_status': status_flecha,
        'scope_notes': scope_notes, 'scope_issues': scope_issues, 'status_global': status_global,
    }
    passo_a_passo_html = _memorial_2024_html(bundle) if detalhado else ""
    return res_flt, res_flm, res_fla, res_cis, res_flecha, passo_a_passo_html

# Substitua a função build_summary_html por esta versão:
def build_summary_html(Msd, Vsd, res_flt, res_flm, res_fla, res_cisalhamento, res_flecha):
    verificacoes = [
        ('Flexão (FLT)', f"{Msd/100:.2f} kN·m", f"{res_flt['Mrdx']/100:.2f} kN·m" if res_flt['status'] != 'N/A' else 'N/A', res_flt['eficiencia'], res_flt['status']),
        ('Flexão (FLM)', f"{Msd/100:.2f} kN·m", f"{res_flm['Mrdx']/100:.2f} kN·m", res_flm['eficiencia'], res_flm['status']),
        ('Flexão (FLA/Anexo E)', f"{Msd/100:.2f} kN·m", f"{res_fla['Mrdx']/100:.2f} kN·m", res_fla['eficiencia'], res_fla['status']),
        ('Cisalhamento', f"{Vsd:.2f} kN", f"{res_cisalhamento['Vrd']:.2f} kN", res_cisalhamento['eficiencia'], res_cisalhamento['status']),
        ('Flecha (ELS)', f"{res_flecha['flecha_max']:.2f} cm" if res_flecha['status'] != "N/A" else "N/A", f"≤ {res_flecha['flecha_limite']:.2f} cm" if res_flecha['status'] != "N/A" else "N/A", res_flecha['eficiencia'], res_flecha['status'])
    ]
    if res_flt.get('rupture_Mrd') is not None:
        verificacoes.insert(3, (
            'Flexão — ruptura na mesa tracionada',
            f"{Msd/100:.2f} kN·m",
            f"{res_flt['rupture_Mrd']/100:.2f} kN·m",
            res_flt['rupture_efficiency'],
            res_flt['rupture_status'],
        ))
    rows_html = ""
    for nome, sol, res, efic, status in verificacoes:
        # A MUDANÇA ESTÁ AQUI: adiciona a classe 'pass' ou 'fail' ao <td> do status
        status_class = "pass" if status in {"APROVADO", "N/A"} else "fail"
        efic_str = f"{efic:.1f}%" if status != "N/A" and isinstance(efic, (int, float)) and math.isfinite(efic) else "N/A"
        rows_html += f"""<tr><td>{nome}</td><td>{sol}</td><td>{res}</td><td>{efic_str}</td><td class="{status_class}">{status}</td></tr>"""

    # Retorna o HTML da tabela para ser usado no memorial
    return f"""<table class="summary-table">
        <thead><tr><th>Verificação</th><th>Solicitante</th><th>Resistente</th><th>Eficiência</th><th>Status</th></tr></thead>
        <tbody>{rows_html}</tbody>
        <tfoot><tr><th colspan="4">Status global (inclui forças localizadas e escopo)</th><th>{res_flt.get('status_global', 'NÃO VERIFICADO')}</th></tr></tfoot>
    </table>"""

def build_step_by_step_html(L, Msd, Vsd, res_flexao, res_cisalhamento, res_flecha, res_flt, res_flm, res_fla, res_vrd, input_mode):
    html = ""
    html += _render_resistance_calc_section(
        "Flambagem Lateral com Torção (FLT)", Msd, "M_{sd}", "kNm", res_flt, 'Mrdx', "M_{rd}"
    )
    html += _render_resistance_calc_section(
        "Flambagem Local da Mesa (FLM)", Msd, "M_{sd}", "kNm", res_flm, 'Mrdx', "M_{rd}"
    )
    html += _render_resistance_calc_section(
        "Flambagem Local da Alma (FLA)", Msd, "M_{sd}", "kNm", res_fla, 'Mrdx', "M_{rd}"
    )
    html += _build_verification_block_html("Verificação Final à Flexão", Msd/100, "M_{{sd}}", res_flexao['Mrd']/100, "M_{{rd}}", res_flexao['eficiencia'], res_flexao['status'], "kNm")

    html += f"<h3>3.2 Cálculo da Resistência ao Cisalhamento (Vrd)</h3>"
    html += _render_resistance_calc_section(
        "Resistência ao Cisalhamento (VRd)", Vsd, "V_{sd}", "kN", res_vrd, 'Vrd', "V_{rd}"
    )

    if input_mode == "Calcular a partir de Cargas na Viga":
        html += """<h2>4. Verificação de Serviço (ELS)</h2>"""
        html += """<h3>4.1. Cálculo da Flecha Máxima Atuante (δ_max)</h3>"""
        html += "<div class='formula-block'>"

        detalhes_flecha = res_flecha.get('detalhes', {})
        delta_q_details = detalhes_flecha.get('delta_q', {})
        if delta_q_details.get('valor', 0) > 0:
            html += f"<h5>Flecha devido à Carga Distribuída (δ_q)</h5>"
            html += f"<p class='formula'>$${delta_q_details['formula_simbolica']}$$</p>"
            html += f"<p class='formula'>$${delta_q_details['formula_numerica']} = \\mathbf{{{delta_q_details['valor']:.4f}}} \\, cm$$</p>"

        delta_p_details = detalhes_flecha.get('delta_p', {})
        if delta_p_details.get('valor', 0) > 0:
            html += f"<h5>Flecha devido à Carga Pontual (δ_p)</h5>"
            html += f"<p class='formula'>$${delta_p_details['formula_simbolica']}$$</p>"
            html += f"<p class='formula'>$${delta_p_details['formula_numerica']} = \\mathbf{{{delta_p_details['valor']:.4f}}} \\, cm$$</p>"

        html += f"<h5>Flecha Total</h5>"
        q_val = detalhes_flecha.get('delta_q', {}).get('valor', 0)
        p_val = detalhes_flecha.get('delta_p', {}).get('valor', 0)
        html += f"<p class='formula'>$$\\delta_{{max}} = {q_val:.4f} + {p_val:.4f} = \\mathbf{{{res_flecha['flecha_max']:.4f}}} \\, cm$$</p>"
        html += "</div>"

        html += "<h3>4.2. Cálculo da Flecha Limite (δ_lim)</h3>"
        html += f"""<div class="formula-block">
            <p class="formula">$$\\delta_{{lim}} = \\frac{{L}}{{{res_flecha['divisor']}}} = \\frac{{{L:.2f}}}{{{res_flecha['divisor']}}} = \\mathbf{{{res_flecha['flecha_limite']:.2f}}} \\, cm$$</p>
        </div>"""
        
        html += "<h3>4.3. Verificação Final da Flecha</h3>"
        html += _build_verification_block_html("Verificação da Flecha", res_flecha['flecha_max'], "\\delta_{{max}}", res_flecha['flecha_limite'], "\\delta_{{lim}}", res_flecha['eficiencia'], res_flecha['status'], "cm")

    return html

def _render_resistance_calc_section(title, solicitante_val, solicitante_sym, solicitante_unit, details_dict, res_key, res_sym):
    """Renderiza uma seção completa de cálculo de resistência com verificações sequenciais."""
    html = f"<h4>{title}</h4><div class='formula-block'>"

    verificacoes_map = {v.get('verif_for_calc'): v for v in details_dict.get('passos_verificacao', [])}

    for step in details_dict.get('passos_calculo', []):
        if step.get('type') == 'verification':
            html += _render_verification_step(step)
        else:
            html += _render_calculation_step(step)
            
            if step.get('verif_id') in verificacoes_map:
                verificacao_step = verificacoes_map.get(step['verif_id'])
                is_pass = any(s in verificacao_step['conclusao'] for s in ["COMPACTA", "PLÁSTICO", "ESCOAMENTO"])
                status_class = "pass" if is_pass else "fail"
                
                html += f"""
                <div class='verification-step'>
                    <h5>🔍 {verificacao_step['titulo']}</h5>
                    <p><strong>Comparação:</strong> {verificacao_step['texto']}</p>
                    <p class='conclusion {status_class}'>{verificacao_step['conclusao']}</p>
                    <p><strong>Classificação:</strong> {verificacao_step['regime']}</p>
                </div>
                """

    if details_dict.get('Mrdx_calc') or details_dict.get('Vrd_calc'):
        final_calc_key = 'Mrdx_calc' if 'Mrdx_calc' in details_dict else 'Vrd_calc'
        final_calc_info = details_dict[final_calc_key]
        html += _render_calculation_step(final_calc_info)

    if 'verificacao_limite' in details_dict:
        limite_info = details_dict['verificacao_limite']
        html += f"<h5>⚖️ {limite_info['desc']}</h5><div class='verification-step'>{limite_info['texto']}</div>"

    html += "</div>"

    res_val = details_dict.get(res_key, 0)
    solicitante_display = solicitante_val / 100 if solicitante_unit == "kNm" else solicitante_val

    if res_val > 0:
        res_display = res_val / 100 if solicitante_unit == "kNm" else res_val
        eficiencia = (solicitante_val / res_val) * 100
        status = "APROVADO" if eficiencia <= 100.0 else "REPROVADO"
        html += _build_verification_block_html(f"✅ Verificação Final - {title}", solicitante_display, solicitante_sym, res_display, res_sym, eficiencia, status, solicitante_unit)
    else:
        html += "<div class='final-status fail'>❌ REPROVADO (Resistência nula ou não implementada)</div>"

    return html

def _render_verification_step(step_dict):
    """Renderiza um bloco de verificação (comparação) no memorial."""
    desc = step_dict.get('desc', 'Verificação')
    lhs_val = step_dict.get('lhs_value', 0)
    rhs_val = step_dict.get('rhs_value', 0)
    comparator = step_dict.get('comparator', '<')
    passed = step_dict.get('passed', False)

    status_text = "ATENDE" if passed else "NÃO ATENDE"
    status_class = "pass" if passed else "fail"
    conclusion = step_dict.get('conclusion_pass', '') if passed else step_dict.get('conclusion_fail', '')

    html = f"""
    <h5><span style="font-size: 0.8em; color: #6b7280;">Passo de Verificação</span><br>📊 {desc}</h5>
    <div class="verification-step">
        <p class="formula" style="font-size: 1.2em;">
            $$ {lhs_val:.2f} \\, {comparator} \\, {rhs_val:.2f} $$
        </p>
        <div class="final-status {status_class}" style="font-size: 1.1em; padding: 0.75rem; text-transform: none;">Resultado: {status_text}</div>
        <p style="text-align: center; margin-top: 1rem;"><b>Conclusão:</b> {conclusion}</p>
    </div>
    """
    return html

def _render_calculation_step(step_dict):
    """
    Renderiza um passo de cálculo de forma padronizada e explícita:
    1. Fórmula Simbólica
    2. Fórmula Numérica com substituições = Resultado
    """
    desc = step_dict.get('desc', 'Cálculo')
    formula_simbolica = step_dict.get('formula', '')
    valores = step_dict.get('valores', {})
    formula_expandida = step_dict.get('formula_expandida')
    valor_final = step_dict.get('valor')
    unidade = step_dict.get('unidade', '')
    ref = f"<p class='ref-norma'>{step_dict.get('ref', '')}</p>" if step_dict.get('ref') else ""
    nota = step_dict.get('nota', '')

    valor_final_str = ""
    if isinstance(valor_final, (int, float)):
        if valor_final == float('inf'):
            valor_final_str = "\\infty"
        else:
            if abs(valor_final) > 0.01 or valor_final == 0:
                valor_final_str = f"{valor_final:.2f}"
            else:
                valor_final_str = f"{valor_final:.4f}"
    elif valor_final is not None:
        valor_final_str = str(valor_final)

    formula_numerica_final = None
    if formula_expandida:
        formula_numerica_final = formula_expandida
    elif valores:
        formula_numerica = formula_simbolica
        for var, val_num in valores.items():
            if isinstance(val_num, (int, float)):
                if abs(val_num) > 0.01 or val_num == 0: val_str = f"{val_num:.2f}"
                else: val_str = f"{val_num:.3f}"
            else:
                val_str = str(val_num)
            formula_numerica = formula_numerica.replace(var, f"\\mathbf{{{val_str}}}")
        formula_numerica_final = formula_numerica
    
    titulo = f"<h6>{desc}</h6>" if '(' in desc else f"<h5>📏 {desc}</h5>"
    
    html_output = f"""{titulo}
                    <p class="formula">$${formula_simbolica}$$</p>"""
    
    if formula_numerica_final:
        html_output += f"""<p class="formula">$${formula_numerica_final} = \\mathbf{{{valor_final_str}}} \\, {unidade}$$</p>"""
    elif valor_final is not None:
          html_output += f"""<p class="formula">$${formula_simbolica} = \\mathbf{{{valor_final_str}}} \\, {unidade}$$</p>"""

    html_output += f"{ref}{nota}"
    
    return html_output

# ==============================================================================
# 5. APLICAÇÃO PRINCIPAL STREAMLIT (REESTRUTURADA)
# ==============================================================================

def main():
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = None
    if 'detailed_analysis_html' not in st.session_state:
        st.session_state.detailed_analysis_html = None
    if 'analysis_mode' not in st.session_state:
        st.session_state.analysis_mode = "batch"
    if 'profile_efficiency_chart' not in st.session_state:
        st.session_state.profile_efficiency_chart = None

    all_sheets = load_data_from_local_file()
    if not all_sheets:
        st.stop()
    
    # ADICIONE A CHAMADA DA FUNÇÃO AQUI
    create_navigation_buttons()
    
    st.markdown(HTML_TEMPLATE_CSS_PRO, unsafe_allow_html=True)
    create_professional_header()

    with st.sidebar:
        st.markdown("## ⚙️ Configuração do Projeto")
        with st.expander("📋 Identificação do Projeto", expanded=False):
            projeto_nome = st.text_input("Nome do Projeto", "Análise Estrutural")
            engenheiro = st.text_input("Engenheiro Responsável", " ")
            data_projeto = st.date_input("Data", datetime.now())
            revisao = st.text_input("Revisão", "00")
        
        st.markdown("---")
        st.markdown("### 🏗️ Modelo Estrutural")
        tipo_viga = st.selectbox("🔗 Tipo de Viga:", ('Bi-apoiada', 'Engastada e Livre (Balanço)', 'Bi-engastada', 'Engastada e Apoiada'), key='tipo_viga')
        L_cm = st.number_input("📏 Comprimento (L, cm)", 10.0, value=500.0, step=10.0, key='L_cm')
        
        st.markdown("---")
        st.markdown("### ⚖️ Carregamento")
        input_mode = st.radio("Método de entrada:", ("Calcular a partir de Cargas na Viga", "Inserir Esforços Manualmente"), key='input_mode')

        Msd, Vsd, q_serv_kn_cm, p_load_serv = 0.0, 0.0, 0.0, None
        q_g_kn_cm = q_q_kn_cm = p_g_kn = p_q_kn = 0.0
        larg_esq_cm = larg_dir_cm = larg_inf_total_m = 0.0
        g_area = q_area = 0.0
        p_pos_cm = L_cm / 2.0
        point_bearing_cm = 10.0
        manual_local_checks_confirmed = False
        include_self_weight = True
        gamma_g, gamma_q, gamma_self_weight = 1.50, 1.50, 1.25
        psi1, psi2 = 0.6, 0.4
        els_combination = 'rare'
        els_combination_text = 'Rara: G + Q principal'
        detalhes_esforcos_memorial = {'input_mode': input_mode, 'Msd': Msd, 'Vsd': Vsd, 'L_cm': L_cm}

        if input_mode == "Calcular a partir de Cargas na Viga":
            with st.container(border=True):
                st.subheader("Ações características")
                larg_esq_cm = st.number_input("Largura da laje à esquerda (cm)", 0.0, value=200.0, step=10.0, key='larg_esq_cm')
                larg_dir_cm = st.number_input("Largura da laje à direita (cm)", 0.0, value=200.0, step=10.0, key='larg_dir_cm')
                larg_inf_total_m = (larg_esq_cm + larg_dir_cm) / 200.0
                st.info(f"Largura de influência B = {larg_inf_total_m:.2f} m")
                g_area = st.number_input("Ação permanente adicional Gk (kN/m²)", 0.0, value=1.5, step=0.25, key='g_area')
                q_area = st.number_input("Ação variável Qk (kN/m²)", 0.0, value=3.0, step=0.25, key='q_area')
                q_g_kn_cm = g_area * larg_inf_total_m / 100.0
                q_q_kn_cm = q_area * larg_inf_total_m / 100.0
                q_serv_kn_cm = q_g_kn_cm + q_q_kn_cm

                st.subheader("Força pontual localizada")
                add_p_load = st.checkbox("Adicionar força pontual", key='add_p_load')
                if add_p_load:
                    p_g_kn = st.number_input("Parcela permanente Pg,k (kN)", 0.0, value=0.0, step=1.0, key='p_g_kn')
                    p_q_kn = st.number_input("Parcela variável Pq,k (kN)", 0.0, value=10.0, step=1.0, key='p_q_kn')
                    p_pos_cm = st.number_input("Posição x desde a esquerda (cm)", 0.0, max_value=L_cm, value=L_cm/2, key='p_pos_cm')
                    point_bearing_cm = st.number_input("Comprimento de atuação ℓn da força (cm)", 0.1, value=10.0, step=0.5, key='point_bearing_cm')
                    p_load_serv = (p_g_kn + p_q_kn, p_pos_cm)

                with st.expander("Combinações ELU/ELS", expanded=False):
                    include_self_weight = st.checkbox("Incluir peso próprio de cada perfil", value=True, key='include_self_weight')
                    gamma_g_label = st.selectbox(
                        "γg da ação permanente adicional — Tabela 1",
                        ("1,50 — elementos construtivos em geral", "1,40 — industrializados com adições in loco", "1,35 — moldados in loco", "1,30 — pré-moldados/madeira/industrializados", "1,25 — aço/equipamentos"),
                        key='gamma_g_label'
                    )
                    gamma_g = float(gamma_g_label[:4].replace(',', '.'))
                    gamma_q_label = st.selectbox(
                        "γq da ação variável principal — Tabela 1",
                        (
                            "1,50 — demais ações variáveis/ações agrupadas",
                            "1,40 — vento",
                            "1,20 — temperatura atmosférica ou ação truncada",
                        ),
                        key='gamma_q_label',
                    )
                    gamma_q = float(gamma_q_label[:4].replace(',', '.'))
                    categoria_psi = st.selectbox(
                        "Categoria da ação variável — Tabela 2",
                        ("Industrial/comercial/escritórios/público", "Residencial de acesso restrito", "Biblioteca/arquivo/depósito/oficina/garagem/cobertura"),
                        key='categoria_psi'
                    )
                    psi_values = {
                        "Industrial/comercial/escritórios/público": (0.7, 0.6, 0.4),
                        "Residencial de acesso restrito": (0.5, 0.4, 0.3),
                        "Biblioteca/arquivo/depósito/oficina/garagem/cobertura": (0.8, 0.7, 0.6),
                    }
                    _, psi1, psi2 = psi_values[categoria_psi]
                    els_label = st.selectbox(
                        "Combinação usada no deslocamento",
                        ("Rara: G + Q principal", "Frequente: G + ψ1 Q", "Quase permanente: G + ψ2 Q", "Somente parcela variável δ3: Q"),
                        key='els_label'
                    )
                    els_map = {
                        "Rara: G + Q principal": 'rare',
                        "Frequente: G + ψ1 Q": 'frequent',
                        "Quase permanente: G + ψ2 Q": 'quasi_permanent',
                        "Somente parcela variável δ3: Q": 'variable_only',
                    }
                    els_combination = els_map[els_label]
                    els_combination_text = els_label
        else:
            with st.container(border=True):
                st.warning("No modo manual, informe esforços já combinados. Sem reações e cargas de serviço, o aplicativo não pode auditar 5.7 nem o ELS.")
                Msd = st.number_input("Momento solicitante Msd (kN·m)", 0.0, value=100.0, key='msd_input') * 100.0
                Vsd = st.number_input("Força cortante Vsd (kN)", 0.0, value=50.0, key='vsd_input')
                manual_local_checks_confirmed = st.checkbox("Confirmo que forças localizadas e ELS foram verificados externamente", value=False, key='manual_external_checks')
            detalhes_esforcos_memorial = {'input_mode': input_mode, 'Msd': Msd, 'Vsd': Vsd, 'L_cm': L_cm}

        st.markdown("---")
        st.markdown("### 🔩 Material e estabilidade lateral")
        material = st.selectbox("Aço estrutural (valores nominais)", ("ASTM A572 Grau 50", "ASTM A36", "Personalizado"), key='material')
        if material == "ASTM A572 Grau 50":
            fy_aco, fu_aco = 34.5, 45.0
        elif material == "ASTM A36":
            fy_aco, fu_aco = 25.0, 40.0
        else:
            fy_aco = st.number_input("fy nominal (kN/cm²)", 1.0, 45.0, 34.5, 0.5, key='fy_aco_custom')
            fu_aco = st.number_input("fu nominal (kN/cm²)", 1.0, 80.0, 45.0, 0.5, key='fu_aco_custom')
        E_aco_input = st.number_input("Módulo de elasticidade E (kN/cm²)", 1_000.0, value=20_000.0, step=100.0, key='E_aco_input')
        st.caption(f"fy = {fy_aco:.2f} kN/cm²; fu = {fu_aco:.2f} kN/cm²; fu/fy = {fu_aco/fy_aco:.3f}")
        for material_issue in validate_material(fy_aco, fu_aco):
            st.error(material_issue)
        material_qualified = st.checkbox(
            "Aço com qualificação estrutural assegurada e requisitos medidos de 4.6.2.2.1 atendidos",
            value=material != "Personalizado",
            key='material_qualified',
        )

        has_tension_flange_holes = st.checkbox(
            "Há furos para parafusos na mesa tracionada",
            value=False,
            key='has_tension_flange_holes',
        )
        tension_flange_net_ratio = 1.0
        if has_tension_flange_holes:
            tension_flange_net_ratio = st.number_input(
                "Relação Afn/Afg da mesa tracionada",
                min_value=0.01,
                max_value=1.00,
                value=0.85,
                step=0.01,
                key='tension_flange_net_ratio',
            )

        flt_condition = st.selectbox(
            "Condição da mesa comprimida",
            ("Trechos com contenções discretas — verificar FLT", "Contenção lateral contínua eficaz — FLT não aplicável"),
            key='flt_condition'
        )
        flt_applicable = flt_condition.startswith("Trechos")
        Lb_projeto = st.number_input("Comprimento destravado Lb (cm)", 1.0, max_value=L_cm, value=L_cm, step=1.0, key='Lb_projeto')
        lb_start_cm = st.number_input("Início do trecho destravado x0 (cm)", 0.0, max_value=max(L_cm-Lb_projeto, 0.0), value=0.0, step=1.0, key='lb_start_cm')
        load_height = st.selectbox(
            "Posição das forças transversais no trecho destravado",
            ("Semialtura da seção", "Abaixo da semialtura — adoção conservadora da semialtura", "Acima da semialtura sem contenção — exige análise de estabilidade"),
            key='load_height'
        )
        cantilever_standard_cb = True
        if tipo_viga == 'Engastada e Livre (Balanço)':
            cantilever_standard_cb = st.checkbox("Empenamento impedido no apoio e extremidade livre sem restrição lateral/torcional", value=True, key='cantilever_standard_cb')
        auto_cb_allowed = (
            input_mode == "Calcular a partir de Cargas na Viga"
            and flt_applicable
            and not load_height.startswith("Acima")
            and tipo_viga != 'Engastada e Livre (Balanço)'
        )
        cb_modo_auto = st.checkbox("Calcular Cb pelo diagrama no trecho Lb", value=auto_cb_allowed, disabled=not auto_cb_allowed, key='cb_modo_auto')
        Cb_projeto = 1.0
        cb_source = ""
        if tipo_viga == 'Engastada e Livre (Balanço)' and cantilever_standard_cb:
            Cb_projeto = 1.0
            st.info("Cb = 1,0 conforme 5.4.2.3-b para a condição declarada do balanço.")
        elif not cb_modo_auto and flt_applicable:
            Cb_projeto = st.number_input("Cb adotado", 0.1, 10.0, 1.0, step=0.05, key='Cb_projeto')
            cb_source = st.text_input("Origem do Cb manual (análise/procedimento)", key='cb_source')
        detalhes_cb_memorial = None

        with st.container(border=True):
            st.subheader("Enrijecedores transversais para cisalhamento")
            usa_enrijecedores = st.checkbox("Considerar enrijecedores", key='usa_enrijecedores')
            a_enr = 0.0
            stiffener_width = stiffener_thickness = None
            stiffener_welded = False
            if usa_enrijecedores:
                a_enr = st.number_input("Espaçamento a (cm)", 1.0, value=100.0, step=1.0, key='a_enr')
                stiffener_width = st.number_input("Largura de cada chapa bs (cm)", 0.1, value=10.0, step=0.5, key='stiffener_width')
                stiffener_thickness = st.number_input("Espessura ts (cm)", 0.1, value=0.8, step=0.1, key='stiffener_thickness')
                stiffener_welded = st.checkbox("Par de enrijecedores soldado à alma e às mesas conforme 5.4.3.1.3-a", value=False, key='stiffener_welded')

        with st.container(border=True):
            st.subheader("Forças localizadas — apoios e carga pontual")
            bearing_left_cm = st.number_input("Comprimento de apoio esquerdo ℓn (cm)", 0.1, value=10.0, step=0.5, key='bearing_left_cm')
            bearing_right_cm = st.number_input("Comprimento de apoio direito ℓn (cm)", 0.1, value=10.0, step=0.5, key='bearing_right_cm')
            support_relative_lateral_restrained = st.checkbox(
                "Nos apoios, o deslocamento lateral relativo entre as mesas é impedido",
                value=True,
                key='support_lateral_restrained',
            )
            point_relative_lateral_restrained = st.checkbox("Na carga pontual, o deslocamento lateral relativo entre mesas é impedido", value=False, key='point_lateral_restrained')
            loaded_flange_rotation_restrained = st.checkbox("A rotação da mesa carregada é impedida", value=False, key='loaded_flange_rotation')
            local_unbraced_cm = st.number_input("Comprimento destravado local ℓ (cm)", 1.0, value=Lb_projeto, step=1.0, key='local_unbraced_cm')
            weld_root_cm = st.number_input("Raiz do filete mesa–alma em perfis soldados (cm; 0 = conservador)", 0.0, value=0.0, step=0.1, key='weld_root_cm')
            support_torsion_restrained = st.checkbox("Apoios impedem rotação torcional ou a alma é ligada a outro elemento", value=True, key='support_torsion_restrained')

        st.markdown("---")
        st.markdown("### 📐 Estado-limite de serviço")
        service_category = st.selectbox("Categoria de deslocamento — Tabela B.1", ("Viga de piso — L/350", "Viga de cobertura — L/250", "Terça/travessa — L/250", "Viga que suporta pilar — L/500", "Personalizado"), key='service_category')
        service_divisors = {"Viga de piso — L/350": 350, "Viga de cobertura — L/250": 250, "Terça/travessa — L/250": 250, "Viga que suporta pilar — L/500": 500}
        limite_flecha_divisor = service_divisors.get(service_category)
        if limite_flecha_divisor is None:
            limite_flecha_divisor = st.number_input("Divisor personalizado x em L/x", 1.0, value=350.0, step=10.0, key='custom_deflection_divisor')
        masonry_on_beam = st.checkbox("Há alvenaria solidarizada sobre ou sob a viga (limite adicional 15 mm)", value=False, key='masonry_on_beam')

        unsupported_reasons = []
        with st.expander("Triagem de aplicabilidade obrigatória", expanded=False):
            has_axial_torsion = st.checkbox("Há força axial, torção ou flexão biaxial", key='scope_axial_torsion')
            has_web_openings = st.checkbox("Há aberturas na alma", key='scope_openings')
            has_fatigue = st.checkbox("Há mais de 20.000 ciclos relevantes de tensão (fadiga)", key='scope_fatigue')
            has_vibration = st.checkbox("O piso/sistema é suscetível a vibrações", key='scope_vibration')
            outside_ambient_scope = st.checkbox("Há incêndio, sismo ou perfil formado a frio", key='scope_external')
        if has_axial_torsion: unsupported_reasons.append("Interação com força axial/torção/flexão biaxial não verificada por este módulo.")
        if has_web_openings: unsupported_reasons.append("Aberturas na alma exigem o Anexo F e não foram modeladas.")
        if has_fatigue: unsupported_reasons.append("Fadiga aplicável: detalhamento e faixa de tensões do Anexo H não informados.")
        if has_vibration: unsupported_reasons.append("Vibrações aplicáveis: avaliação do Anexo I não realizada.")
        if outside_ambient_scope: unsupported_reasons.append("Situação fora do escopo à temperatura ambiente de perfis laminados/soldados.")
        if not material_qualified:
            unsupported_reasons.append(
                "Qualificação estrutural do aço e requisitos efetivamente medidos de 4.6.2.2.1 não confirmados."
            )
        if not support_torsion_restrained: unsupported_reasons.append("5.7.8 exige enrijecedores nos apoios/extremidades declarados sem restrição torcional e com alma livre.")
        standard_cantilever_cb = (
            tipo_viga == 'Engastada e Livre (Balanço)' and cantilever_standard_cb
        )
        if flt_applicable and not cb_modo_auto and not standard_cantilever_cb and not cb_source.strip():
            unsupported_reasons.append(
                "Cb manual sem origem documentada por análise de estabilidade ou procedimento técnico aceito."
            )

    projeto_info = {'nome': projeto_nome, 'engenheiro': engenheiro, 'data': data_projeto.strftime('%d/%m/%Y'), 'revisao': revisao}
    scope_notes = [
        "Viga prismática I/H duplamente simétrica, carregada no plano da alma e fletida no eixo forte.",
        "Análise elástica de primeira ordem; vinculações ideais selecionadas pelo usuário.",
        "Ações gravitacionais estáticas; valores característicos devem vir das normas de ações aplicáveis.",
        f"Material nominal: {material}; fy={fy_aco:.2f} kN/cm²; fu={fu_aco:.2f} kN/cm².",
    ]
    input_params = {
        'tipo_viga': tipo_viga, 'L_cm': L_cm, 'input_mode': input_mode, 'Msd': Msd, 'Vsd': Vsd,
        'q_serv_kn_cm': q_serv_kn_cm, 'p_load_serv': p_load_serv, 'q_g_kn_cm': q_g_kn_cm, 'q_q_kn_cm': q_q_kn_cm,
        'larg_esq_cm': larg_esq_cm, 'larg_dir_cm': larg_dir_cm,
        'larg_inf_total_m': larg_inf_total_m, 'g_area': g_area, 'q_area': q_area,
        'p_g_kn': p_g_kn, 'p_q_kn': p_q_kn, 'p_pos_cm': p_pos_cm,
        'fy_aco': fy_aco, 'fu_aco': fu_aco, 'E_aco': E_aco_input,
        'has_tension_flange_holes': has_tension_flange_holes,
        'tension_flange_net_ratio': tension_flange_net_ratio,
        'Lb_projeto': Lb_projeto, 'lb_start_cm': lb_start_cm, 'Cb_projeto': Cb_projeto,
        'flt_applicable': flt_applicable, 'cb_modo_auto': cb_modo_auto, 'cb_source': cb_source,
        'cantilever_standard_cb': cantilever_standard_cb,
        'detalhes_esforcos_memorial': detalhes_esforcos_memorial, 'detalhes_cb_memorial': detalhes_cb_memorial,
        'usa_enrijecedores': usa_enrijecedores, 'a_enr': a_enr,
        'stiffener_width': stiffener_width, 'stiffener_thickness': stiffener_thickness,
        'stiffener_pair': True, 'stiffener_welded': stiffener_welded,
        'bearing_left_cm': bearing_left_cm, 'bearing_right_cm': bearing_right_cm,
        'support_relative_lateral_restrained': support_relative_lateral_restrained,
        'point_bearing_cm': point_bearing_cm, 'point_relative_lateral_restrained': point_relative_lateral_restrained,
        'loaded_flange_rotation_restrained': loaded_flange_rotation_restrained,
        'local_unbraced_cm': local_unbraced_cm, 'weld_root_cm': weld_root_cm,
        'limite_flecha_divisor': limite_flecha_divisor, 'masonry_on_beam': masonry_on_beam,
        'include_self_weight': include_self_weight, 'gamma_g': gamma_g, 'gamma_q': gamma_q,
        'gamma_self_weight': gamma_self_weight, 'els_combination': els_combination,
        'els_combination_text': els_combination_text, 'psi1': psi1, 'psi2': psi2,
        'elu_combination_text': f'{gamma_g:.2f}·G + {gamma_self_weight:.2f}·PP aço + {gamma_q:.2f}·Q',
        'manual_local_checks_confirmed': manual_local_checks_confirmed,
        'unsupported_reasons': unsupported_reasons, 'scope_notes': scope_notes,
        'projeto_info': projeto_info,
    }

    create_metrics_dashboard(input_params)

    # st.markdown("### 🎯 Modo de Análise") # <- Esta linha foi removida
    st.subheader("🎯 Modo de Análise") # <- Substituída por esta
    
    col1, col2 = st.columns(2)
    if col1.button("📊 Análise em Lote e Otimização", use_container_width=True, type="secondary"):
        st.session_state.analysis_mode = "batch"
    if col2.button("📋 Memorial Detalhado de Perfil", use_container_width=True, type="secondary"):
        st.session_state.analysis_mode = "detailed"

    if st.session_state.analysis_mode == "batch":
        # st.header("📊 Análise em Lote") # <- Esta linha foi removida
        st.subheader("📊 Análise em Lote") # <- Substituída por esta
        
        if st.button("🚀 Iniciar Análise Otimizada", type="primary", use_container_width=True):
            run_batch_analysis(all_sheets, input_params)
        
        if st.session_state.analysis_results is not None:
            df_all_results = st.session_state.analysis_results
            
            tabs = st.tabs([PROFILE_TYPE_MAP.get(name, name) for name in all_sheets.keys()])
            for i, sheet_name in enumerate(all_sheets.keys()):
                with tabs[i]:
                    df_type = df_all_results[df_all_results['Tipo'] == sheet_name].drop(columns=['Tipo'])
                    df_aprovados_cat = df_type[df_type['Status'] == 'APROVADO'].copy().sort_values(by='Peso (kg/m)')
                    df_reprovados_cat = df_type[df_type['Status'] == 'REPROVADO'].copy().sort_values(by='Peso (kg/m)')
                    df_pendentes_cat = df_type[df_type['Status'] == 'NÃO VERIFICADO'].copy().sort_values(by='Peso (kg/m)')

                    # Botão de download para todos os resultados em uma aba
                    if not df_type.empty:
                        df_total = pd.concat([df_aprovados_cat, df_reprovados_cat, df_pendentes_cat])
                        excel_data = create_excel_with_colors([df_total], [f"{sheet_name}_Resultados"])
                        st.download_button(
                            label=f"📥 Baixar todos os resultados ({sheet_name}) em XLSX",
                            data=excel_data,
                            file_name=f"resultados_{sheet_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                    if not df_aprovados_cat.empty:
                        st.plotly_chart(create_top_profiles_chart(df_aprovados_cat), use_container_width=True)
                        with st.expander(f"Ver todos os {len(df_aprovados_cat)} perfis aprovados"):
                            st.dataframe(style_classic_dataframe(df_aprovados_cat), use_container_width=True)
                    else:
                        st.info("Nenhum perfil aprovado nesta categoria.")

                    if not df_reprovados_cat.empty:
                        with st.expander(f"Ver os {len(df_reprovados_cat)} perfis reprovados"):
                            st.dataframe(style_classic_dataframe(df_reprovados_cat), use_container_width=True)

                    if not df_pendentes_cat.empty:
                        with st.expander(f"Ver os {len(df_pendentes_cat)} perfis não verificados"):
                            st.dataframe(style_classic_dataframe(df_pendentes_cat), use_container_width=True)

    elif st.session_state.analysis_mode == "detailed":
        # st.header("📋 Memorial Detalhado") # <- Esta linha foi removida
        st.subheader("📋 Memorial Detalhado") # <- Substituída por esta
        
        display_names = [PROFILE_TYPE_MAP.get(name, name) for name in all_sheets.keys()]
        reverse_name_map = {v: k for k, v in PROFILE_TYPE_MAP.items()}

        col1, col2 = st.columns(2)
        selected_display_name = col1.selectbox("Selecione o Tipo de Perfil:", display_names)
        sheet_name = reverse_name_map.get(selected_display_name, selected_display_name)
        df_selecionado = all_sheets[sheet_name]
        perfil_selecionado_nome = col2.selectbox("Selecione o Perfil Específico:", df_selecionado['Bitola (mm x kg/m)'])

        if st.button("📄 Gerar Memorial Completo", type="primary", use_container_width=True):
            run_detailed_analysis(df_selecionado, perfil_selecionado_nome, selected_display_name, input_params)

        if st.session_state.detailed_analysis_html:
            with st.expander("📊 Resumo Visual da Análise", expanded=True):
                st.plotly_chart(st.session_state.profile_efficiency_chart, use_container_width=True)

            with st.expander("📄 Visualização do Memorial", expanded=True):
                # O memorial é deliberadamente extenso; a altura maior evita que
                # o leitor precise alternar entre duas barras de rolagem a cada etapa.
                st.components.v1.html(st.session_state.detailed_analysis_html, height=9000, width=2500, scrolling=True)
            
            st.download_button(
                label="📥 Baixar Memorial em HTML",
                data=st.session_state.detailed_analysis_html.encode('utf-8'),
                file_name=f"Memorial_{perfil_selecionado_nome.replace(' ', '_')}.html",
                mime="text/html",
                use_container_width=True
            )

def run_detailed_analysis(df, perfil_nome, perfil_tipo_display, input_params):
    with st.spinner(f"Gerando análise completa para {perfil_nome}..."):
        try:
            perfil_series = df[df['Bitola (mm x kg/m)'] == perfil_nome].iloc[0]
            props = get_profile_properties(perfil_series)

            tipo_fabricacao = "Soldado" if "Soldado" in perfil_tipo_display else "Laminado"
            # A análise, as combinações e Cb são agora documentados no bloco normativo 2024.
            esforcos_html = ""
            cb_calc_html = ""

            res_flt, res_flm, res_fla, res_cis, res_flecha, passo_a_passo = perform_all_checks(
                props=props, detalhado=True, tipo_fabricacao=tipo_fabricacao, **input_params
            )
            
            eficiencias = {
                "FLT": res_flt['eficiencia'],
                "FLM": res_flm['eficiencia'],
                "FLA": res_fla['eficiencia'],
                "Cisalhamento": res_cis['eficiencia'],
                "Flecha": res_flecha['eficiencia'],
            }
            if res_flt.get('rupture_status') != 'N/A':
                eficiencias["Ruptura da mesa"] = res_flt['rupture_efficiency']
            st.session_state.profile_efficiency_chart = create_profile_efficiency_chart(perfil_nome, eficiencias)
            
            resumo_html = build_summary_html(res_flt['Msd'], res_cis['Vsd'], res_flt, res_flm, res_fla, res_cis, res_flecha)
            resultados = {'resumo_html': resumo_html, 'passo_a_passo_html': passo_a_passo, 'esforcos_html': esforcos_html, 'cb_calc_html': cb_calc_html}
            
            html_content = create_professional_memorial_html(
                perfil_nome, perfil_tipo_display, resultados,
                f"""
                
                <div style="text-align: left;">
                    <p><strong>Módulo de Elasticidade (E):</strong> {input_params['E_aco']:.2f} kN/cm²</p>
                    <p><strong>Tensão de Escoamento (fy):</strong> {input_params['fy_aco']:.2f} kN/cm²</p>
                    <p><strong>Tensão de Ruptura (fu):</strong> {input_params['fu_aco']:.2f} kN/cm²</p>
                    <p><strong>Altura total (d):</strong> {perfil_series.get('d (mm)'):.2f} mm</p>
                    <p><strong>Largura da Mesa (bf):</strong> {perfil_series.get('bf (mm)'):.2f} mm</p>
                    <p><strong>Espessura da Alma (tw):</strong> {perfil_series.get('tw (mm)'):.2f} mm</p>
                    <p><strong>Espessura da Mesa (tf):</strong> {perfil_series.get('tf (mm)'):.2f} mm</p>
                    <p><strong>Distância entre faces internas (h):</strong> {perfil_series.get('h (mm)'):.2f} mm</p>
                    <p><strong>Altura livre da alma (d′):</strong> {perfil_series.get("d' (mm)"):.2f} mm</p>
                    <p><strong>Área (A):</strong> {perfil_series.get('Área (cm2)'):.2f} cm²</p>
                    <p><strong>Inércia Ix:</strong> {perfil_series.get('Ix (cm4)'):.2f} cm⁴</p>
                    <p><strong>Módulo de Seção Elástico (Wx):</strong> {perfil_series.get('Wx (cm3)'):.2f} cm³</p>
                    <p><strong>Raio de Giração (rx):</strong> {props.get('rx'):.2f} cm</p>
                    <p><strong>Módulo de Seção Plástico (Zx):</strong> {props.get('Zx'):.2f} cm³</p>
                    <p><strong>Inércia Iy:</strong> {props.get('Iy'):.2f} cm⁴</p>
                    <p><strong>Raio de Giração (ry):</strong> {props.get('ry'):.2f} cm</p>
                    <p><strong>Constante de Torção (J):</strong> {props.get('J'):.2f} cm⁴</p>
                    <p><strong>Constante de Empenamento (Cw):</strong> {props.get('Cw'):.2f} cm⁶</p>
                </div>
                """, input_params['projeto_info']
            )
            st.session_state.detailed_analysis_html = html_content
        except Exception as e:
            st.error(f"❌ Ocorreu um erro: {e}")

def run_batch_analysis(all_sheets, input_params):
    all_results = []
    progress_bar = st.progress(0, text="Analisando perfis...")
    total_perfis = sum(len(df) for df in all_sheets.values())
    perfis_processados = 0
    
    for sheet_name, df in all_sheets.items():
        tipo_fabricacao_auto = PROFILE_FABRICATION_MAP.get(sheet_name, "Laminado")
        
        for _, row in df.iterrows():
            perfis_processados += 1
            progress_bar.progress(perfis_processados / total_perfis, text=f"Analisando: {row['Bitola (mm x kg/m)']}")
            try:
                props = get_profile_properties(row)
                res_flt, res_flm, res_fla, res_cis, res_flecha, _ = perform_all_checks(
                    props=props, tipo_fabricacao=tipo_fabricacao_auto, **input_params
                )
                
                status_geral = res_flt.get('status_global', 'NÃO VERIFICADO')
                local_efficiencies = [item['efficiency'] for item in res_cis.get('local_checks', [])]
                max_local_efficiency = max(local_efficiencies, default=0.0)
                
                all_results.append({
                    'Tipo': sheet_name, 'Perfil': row['Bitola (mm x kg/m)'],
                    'Peso (kg/m)': props.get('Peso', 0), 'Status': status_geral,
                    'Ef. FLT (%)': res_flt['eficiencia'], 'Ef. FLM (%)': res_flm['eficiencia'],
                    'Ef. FLA (%)': res_fla['eficiencia'], 'Ef. Cisalhamento (%)': res_cis['eficiencia'],
                    'Ef. Ruptura Mesa (%)': (
                        res_flt['rupture_efficiency']
                        if res_flt.get('rupture_status') != 'N/A' else None
                    ),
                    'Ef. Forças Locais (%)': max_local_efficiency,
                    'Ef. Flecha (%)': res_flecha['eficiencia']
                })
            except (ValueError, KeyError) as exc:
                all_results.append({
                    'Tipo': sheet_name, 'Perfil': row.get('Bitola (mm x kg/m)', 'N/D'),
                    'Peso (kg/m)': row.get('Massa Linear (kg/m)', 0), 'Status': 'NÃO VERIFICADO',
                    'Observação': str(exc),
                })
    progress_bar.empty()
    st.session_state.analysis_results = pd.DataFrame(all_results) if all_results else pd.DataFrame()

if __name__ == '__main__':
    main()
